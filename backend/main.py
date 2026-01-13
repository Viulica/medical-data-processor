import os
import sys
import logging
import gc
import signal
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# Set environment variables to limit threading and prevent resource exhaustion
os.environ['OPENBLAS_NUM_THREADS'] = '12'
os.environ['OMP_NUM_THREADS'] = '12'
os.environ['MKL_NUM_THREADS'] = '12'

# Configure logging first
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Import export utilities
from export_utils import save_dataframe_dual_format, convert_csv_to_xlsx

def ensure_csv_file(input_file_path: str, output_file_path: str = None) -> str:
    """
    Convert XLSX/XLS file to CSV if needed, or return the CSV path as-is.
    
    Args:
        input_file_path: Path to input file (CSV, XLSX, or XLS)
        output_file_path: Optional output path. If not provided, uses input path with .csv extension
    
    Returns:
        str: Path to CSV file (either original or converted)
    """
    input_path = Path(input_file_path)
    
    # If already CSV, return as-is
    if input_path.suffix.lower() == '.csv':
        return str(input_path)
    
    # If XLSX or XLS, convert to CSV
    if input_path.suffix.lower() in ('.xlsx', '.xls'):
        if output_file_path is None:
            output_file_path = str(input_path.with_suffix('.csv'))
        
        try:
            # Read Excel file
            df = pd.read_excel(input_file_path, dtype=str)
            # Save as CSV
            df.to_csv(output_file_path, index=False)
            logger.info(f"Converted {input_file_path} to {output_file_path}")
            return output_file_path
        except Exception as e:
            raise Exception(f"Failed to convert Excel file to CSV: {str(e)}")
    
    # Unknown format
    raise Exception(f"Unsupported file format: {input_path.suffix}")

def kill_process_tree(process):
    """Kill a process and all its children"""
    try:
        import psutil
        parent = psutil.Process(process.pid)
        children = parent.children(recursive=True)
        
        # Kill children first
        for child in children:
            try:
                child.kill()
                logger.info(f"Killed child process {child.pid}")
            except psutil.NoSuchProcess:
                pass
        
        # Kill parent
        process.kill()
        process.wait()
        logger.info(f"Killed parent process {process.pid}")
    except Exception as e:
        logger.error(f"Error killing process tree: {e}")
        # Fallback to simple kill
        try:
            process.kill()
            process.wait()
        except:
            pass

# Try to import required modules
try:
    from fastapi import FastAPI, UploadFile, File, Form, HTTPException, BackgroundTasks, Body
    from fastapi.middleware.cors import CORSMiddleware
    from fastapi.responses import FileResponse
    from starlette.background import BackgroundTask
    logger.info("✅ FastAPI imported successfully")
except ImportError as e:
    logger.error(f"❌ Failed to import FastAPI: {e}")
    raise e

try:
    import zipfile
    import tempfile
    import shutil
    import uuid
    from pathlib import Path
    from typing import List
    import subprocess
    import json
    import time
    from typing import List, Optional
    import pandas as pd
    logger.info("✅ Standard library modules imported successfully")
except ImportError as e:
    logger.error(f"❌ Failed to import standard library module: {e}")
    raise e



# Global port configuration - this ensures consistency
PORT = int(os.environ.get('PORT', 8000))

app = FastAPI(title="PDF Processing API", version="1.0.0")

@app.on_event("startup")
async def startup_event():
    """Log when the application starts"""
    logger.info("🚀 Medical Data Processor API starting up...")
    logger.info(f"📡 Port: {PORT}")
    logger.info(f"🔑 Google API Key: {'Set' if os.environ.get('GOOGLE_API_KEY') else 'Not set'}")
    logger.info(f"🔑 OCR.space API Key: {'Set' if os.environ.get('OCRSPACE_API_KEY') else 'Not set'}")
    logger.info(f"🔑 OpenRouter API Key: {'Set' if os.environ.get('OPENROUTER_API_KEY') else 'Not set'}")
    logger.info(f"🔑 OpenAI API Key: {'Set' if os.environ.get('OPENAI_API_KEY') else 'Not set'} (fallback for OpenRouter)")
    logger.info(f"🌍 Environment: {os.environ.get('RAILWAY_ENVIRONMENT', 'development')}")
    logger.info(f"🏗️ Railway Project: {os.environ.get('RAILWAY_PROJECT_ID', 'unknown')}")
    logger.info(f"🚂 Railway Service: {os.environ.get('RAILWAY_SERVICE_NAME', 'unknown')}")
    logger.info(f"🔧 PORT env var: {os.environ.get('PORT', 'not set')}")
    logger.info(f"🔧 GLOBAL PORT: {PORT}")
    
    # Initialize database schema
    try:
        from db_utils import init_database
        init_database()
        logger.info("✅ Database schema initialized")
    except Exception as e:
        logger.error(f"❌ Failed to initialize database: {e}")
    
    # Schedule periodic cleanup of expired unified results (every 6 hours)
    try:
        import asyncio
        from db_utils import delete_expired_unified_results
        
        async def periodic_cleanup():
            while True:
                try:
                    await asyncio.sleep(6 * 60 * 60)  # Wait 6 hours
                    deleted_count = delete_expired_unified_results()
                    if deleted_count > 0:
                        logger.info(f"Cleaned up {deleted_count} expired unified results")
                except Exception as e:
                    logger.error(f"Error in periodic cleanup: {e}")
        
        # Start cleanup task in background
        asyncio.create_task(periodic_cleanup())
        logger.info("✅ Started periodic cleanup task for expired unified results")
    except Exception as e:
        logger.warning(f"Failed to start cleanup task: {e}")

# Add CORS middleware
# Get allowed origins from environment or use defaults
allowed_origins = os.environ.get(
    "CORS_ORIGINS",
    "https://medical-data-processor.vercel.app,http://localhost:8080,http://localhost:3000"
).split(",")

app.add_middleware(
    CORSMiddleware,
    allow_origins=[origin.strip() for origin in allowed_origins],  # Explicitly allow frontend origins
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Global storage for job status
job_status = {}

class ProcessingJob:
    def __init__(self, job_id: str):
        self.job_id = job_id
        self.status = "pending"
        self.progress = 0
        self.message = "Job created"
        self.result_file = None
        self.result_file_xlsx = None  # Store XLSX version
        self.result_file_json = None  # Store JSON version
        self.error = None
        self.metadata = {}  # Store additional data like reasoning

def process_pdfs_background(job_id: str, zip_path: str, excel_path: str, n_pages: int, excel_filename: str, model: str = "gemini-flash-latest", worktracker_group: str = None, worktracker_batch: str = None, extract_csn: bool = False):
    """Background task to process PDFs"""
    import os
    
    job = job_status[job_id]
    
    try:
        job.status = "processing"
        job.message = "Unzipping files..."
        job.progress = 10
        
        # Create temporary directory for processing
        temp_dir = Path(f"/tmp/processing_{job_id}")
        temp_dir.mkdir(exist_ok=True)
        
        # Unzip files
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(temp_dir / "input")
        
        job.message = "Files unzipped, starting processing..."
        job.progress = 20
        
        # Copy Excel file to temp directory
        excel_dest = temp_dir / "instructions" / excel_filename
        excel_dest.parent.mkdir(exist_ok=True)
        shutil.copy2(excel_path, excel_dest)
        
        # Set up environment for the processing script
        env = os.environ.copy()
        env['PYTHONPATH'] = str(Path(__file__).parent / "current")
        # Limit OpenBLAS threads to prevent resource exhaustion
        env['OPENBLAS_NUM_THREADS'] = '12'
        env['OMP_NUM_THREADS'] = '12'
        env['MKL_NUM_THREADS'] = '12'
        
        # Run the processing script
        script_path = Path(__file__).parent / "current" / "2-extract_info.py"
        
        job.message = f"Processing PDFs (extracting first {n_pages} pages per patient)..."
        job.progress = 30
        
        # Run the script with subprocess (with timeout)
        process = None
        try:
            cmd = [
                sys.executable, str(script_path),
                str(temp_dir / "input"),
                str(excel_dest),
                str(n_pages),  # n_pages parameter
                "7",  # max_workers
                model  # model parameter
            ]
            
            # Add worktracker fields if provided
            if worktracker_group:
                cmd.append(worktracker_group)
            else:
                cmd.append("")  # Empty string if not provided
                
            if worktracker_batch:
                cmd.append(worktracker_batch)
            else:
                cmd.append("")  # Empty string if not provided
            
            # Add extract_csn flag
            if extract_csn:
                cmd.append("true")
            else:
                cmd.append("false")
            
            # Add empty progress_file (not needed for sequential processing)
            cmd.append("")
            
            process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, cwd=temp_dir, env=env)
            
            # Wait for process with timeout (30 minutes max)
            stdout, stderr = process.communicate(timeout=1800)
            
            # Log stdout and stderr for debugging
            logger.info(f"Script stdout: {stdout}")
            if stderr:
                logger.info(f"Script stderr: {stderr}")
            
            if process.returncode != 0:
                raise Exception(f"Processing failed: {stderr}")
        except subprocess.TimeoutExpired:
            # Kill the process and all its children if it times out
            if process:
                logger.warning(f"Process timed out, killing process tree for PID {process.pid}")
                kill_process_tree(process)
                gc.collect()  # Force cleanup after killing process
            raise Exception("Processing timed out after 30 minutes")
        except Exception as e:
            # Ensure process is terminated on any error
            if process and process.poll() is None:
                logger.warning(f"Process failed, killing process tree for PID {process.pid}")
                kill_process_tree(process)
                gc.collect()  # Force cleanup after killing process
            raise e
        
        job.message = "Processing complete, preparing results..."
        job.progress = 80
        
        # Find the generated CSV file
        extracted_dir = temp_dir / "extracted"
        csv_files = list(extracted_dir.glob("*.csv"))
        
        if not csv_files:
            # Provide detailed diagnostic information
            input_dir = temp_dir / "input"
            input_exists = input_dir.exists()
            input_files = list(input_dir.rglob("*")) if input_exists else []
            extracted_exists = extracted_dir.exists()
            
            error_msg = f"No CSV file generated. Diagnostics:\n"
            error_msg += f"- Input dir exists: {input_exists}\n"
            error_msg += f"- Input files: {[f.name for f in input_files[:10]]}\n"  # Show first 10 files
            error_msg += f"- Extracted dir exists: {extracted_exists}\n"
            error_msg += f"- Script stdout (last 500 chars): {stdout[-500:] if stdout else 'empty'}\n"
            error_msg += f"- Script stderr (last 500 chars): {stderr[-500:] if stderr else 'empty'}"
            logger.error(error_msg)
            raise Exception("No CSV file generated - check logs for details")
        
        # Copy the CSV to a permanent location and create XLSX version
        result_base = Path(f"/tmp/results/{job_id}")
        result_base.parent.mkdir(exist_ok=True)
        
        # Copy CSV
        result_file = result_base.with_suffix('.csv')
        shutil.copy2(csv_files[0], result_file)
        
        # Create XLSX version
        try:
            import pandas as pd
            df = pd.read_csv(result_file, dtype=str)
            result_file_xlsx = result_base.with_suffix('.xlsx')
            df.to_excel(result_file_xlsx, index=False, engine='openpyxl')
            job.result_file_xlsx = str(result_file_xlsx)
            logger.info(f"Created XLSX version: {result_file_xlsx}")
        except Exception as e:
            logger.warning(f"Could not create XLSX version: {e}")
        
        job.result_file = str(result_file)
        job.status = "completed"
        job.progress = 100
        job.message = "Processing completed successfully!"
        
        # Clean up temporary files
        shutil.rmtree(temp_dir)
        os.unlink(zip_path)
        os.unlink(excel_path)
        
        # Force garbage collection to free memory
        gc.collect()
        logger.info(f"Job {job_id}: Memory cleaned up after processing")
        
    except Exception as e:
        job.status = "failed"
        job.error = str(e)
        job.message = f"Processing failed: {str(e)}"
        logger.error(f"Job {job_id} failed: {str(e)}")
        
        # Clean up memory even on failure
        gc.collect()

def split_pdf_background(job_id: str, pdf_paths: List[str], filter_string: str, detection_shift: int = 0):
    """Background task to split multiple PDFs using the existing detection script"""
    job = job_status[job_id]
    
    try:
        job.status = "processing"
        job.message = f"Starting PDF splitting for {len(pdf_paths)} file(s)..."
        job.progress = 10
        
        # Use existing input and output folders
        current_dir = Path(__file__).parent / "current"
        input_dir = current_dir / "input"
        output_dir = current_dir / "output"
        
        # Create input and output folders if they don't exist
        input_dir.mkdir(exist_ok=True)
        output_dir.mkdir(exist_ok=True)
        
        # Clear output folder first
        if output_dir.exists():
            for file in output_dir.glob("*.pdf"):
                file.unlink()
        
        # Copy all PDFs to input folder
        input_pdf_paths = []
        for idx, pdf_path in enumerate(pdf_paths):
            pdf_filename = f"{Path(pdf_path).stem}_{idx}{Path(pdf_path).suffix}"
            input_pdf_path = input_dir / pdf_filename
            shutil.copy2(pdf_path, input_pdf_path)
            
            # Verify the file was copied
            if not input_pdf_path.exists():
                raise Exception(f"Failed to copy PDF to input folder: {input_pdf_path}")
            
            input_pdf_paths.append(input_pdf_path)
            logger.info(f"PDF {idx+1}/{len(pdf_paths)} copied to input folder: {input_pdf_path}")
        
        logger.info(f"Input folder now contains: {[f.name for f in input_dir.glob('*.pdf')]}")
        
        job.message = f"PDFs copied ({len(pdf_paths)} file(s)), running split script..."
        job.progress = 30
        
        # Set up environment for the splitting script
        env = os.environ.copy()
        env['PYTHONPATH'] = str(current_dir)
        # Limit OpenBLAS threads to prevent resource exhaustion
        env['OPENBLAS_NUM_THREADS'] = '12'
        env['OMP_NUM_THREADS'] = '12'
        env['MKL_NUM_THREADS'] = '12'
        
        # Original script path
        script_path = current_dir / "1-split_pdf_by_detections.py"
        
        # Create a temporary script with custom filter string and shift
        temp_script_path = current_dir / f"temp_split_{job_id}.py"
        
        # Read the original script
        with open(script_path, 'r') as f:
            script_content = f.read()
        
        # Replace the filter strings with the custom one
        custom_filter_lines = f'FILTER_STRINGS = ["{filter_string}"]\n'
        custom_shift_line = f'DETECTION_SHIFT = {detection_shift}\n'
        
        # Find and replace the FILTER_STRINGS and DETECTION_SHIFT lines
        lines = script_content.split('\n')
        new_lines = []
        filter_replaced = False
        shift_replaced = False
        
        for line in lines:
            if line.strip().startswith('FILTER_STRINGS = ') and not filter_replaced:
                new_lines.append(custom_filter_lines)
                filter_replaced = True
            elif line.strip().startswith('DETECTION_SHIFT = ') and not shift_replaced:
                new_lines.append(custom_shift_line)
                shift_replaced = True
            else:
                new_lines.append(line)
        
        # If DETECTION_SHIFT wasn't found, add it after FILTER_STRINGS
        if not shift_replaced and filter_replaced:
            # Find where FILTER_STRINGS was inserted and add shift after it
            for i, line in enumerate(new_lines):
                if line.strip().startswith('FILTER_STRINGS = '):
                    new_lines.insert(i + 1, custom_shift_line)
                    break
        
        # Write the modified script
        with open(temp_script_path, 'w') as f:
            f.write('\n'.join(new_lines))
        
        shift_msg = f" (shift: {detection_shift})" if detection_shift != 0 else ""
        job.message = f"Splitting PDF into sections using filter: '{filter_string}'{shift_msg}..."
        job.progress = 50
        
        # Run the modified script
        logger.info(f"Running split script with custom filter: {filter_string}")
        logger.info(f"Working directory: {current_dir}")
        logger.info(f"Input folder contents: {list(input_dir.glob('*.pdf'))}")
        
        process = None
        try:
            process = subprocess.Popen([
                sys.executable, str(temp_script_path)
            ], stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, cwd=current_dir, env=env)
            
            # Wait for process with timeout (20 minutes max for splitting)
            stdout, stderr = process.communicate(timeout=1200)
            
            logger.info(f"Script stdout: {stdout}")
            logger.info(f"Script stderr: {stderr}")
            
            if process.returncode != 0:
                raise Exception(f"Splitting failed: {stderr}")
        except subprocess.TimeoutExpired:
            # Kill the process and all its children if it times out
            if process:
                logger.warning(f"PDF splitting timed out, killing process tree for PID {process.pid}")
                kill_process_tree(process)
                gc.collect()  # Force cleanup after killing process
            raise Exception("PDF splitting timed out after 20 minutes")
        except Exception as e:
            # Ensure process is terminated on any error
            if process and process.poll() is None:
                logger.warning(f"PDF splitting failed, killing process tree for PID {process.pid}")
                kill_process_tree(process)
                gc.collect()  # Force cleanup after killing process
            raise e
        
        job.message = "Creating ZIP archive of split PDFs..."
        job.progress = 80
        
        # Create ZIP file with all split PDFs
        zip_path = Path(f"/tmp/results/{job_id}_split_pdfs.zip")
        zip_path.parent.mkdir(exist_ok=True)
        
        # Find all PDF files created by the script
        pdf_files = list(output_dir.glob("*.pdf"))
        
        if not pdf_files:
            raise Exception("No PDF files were created by the splitting script")
        
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            for pdf_file in pdf_files:
                zipf.write(pdf_file, pdf_file.name)
        
        job.result_file = str(zip_path)
        job.status = "completed"
        job.progress = 100
        job.message = f"PDF splitting completed successfully! Created {len(pdf_files)} sections from {len(pdf_paths)} PDF(s)."
        
        # Clean up uploaded files, input folder, and temp script
        for pdf_path in pdf_paths:
            if os.path.exists(pdf_path):
                os.unlink(pdf_path)
                logger.info(f"Cleaned up uploaded file: {pdf_path}")
        
        for input_pdf_path in input_pdf_paths:
            if input_pdf_path.exists():
                input_pdf_path.unlink()
                logger.info(f"Cleaned up input file: {input_pdf_path}")
        
        # Clean up temp script
        if temp_script_path.exists():
            temp_script_path.unlink()
            logger.info(f"Cleaned up temp script: {temp_script_path}")
        
        # Force garbage collection to free memory
        gc.collect()
        logger.info(f"Job {job_id}: Memory cleaned up after PDF splitting")
        
    except Exception as e:
        job.status = "failed"
        job.error = str(e)
        job.message = f"PDF splitting failed: {str(e)}"
        logger.error(f"Split job {job_id} failed: {str(e)}")
        
        # Clean up on error
        try:
            for pdf_path in pdf_paths:
                if os.path.exists(pdf_path):
                    os.unlink(pdf_path)
                    logger.info(f"Cleaned up uploaded file on error: {pdf_path}")
            if 'input_pdf_paths' in locals():
                for input_pdf_path in input_pdf_paths:
                    if input_pdf_path.exists():
                        input_pdf_path.unlink()
                        logger.info(f"Cleaned up input file on error: {input_pdf_path}")
            if 'temp_script_path' in locals() and temp_script_path.exists():
                temp_script_path.unlink()
                logger.info(f"Cleaned up temp script on error: {temp_script_path}")
        except Exception as cleanup_error:
            logger.error(f"Cleanup error: {cleanup_error}")
        
        # Clean up memory even on failure
        gc.collect()

def split_pdf_gemini_background(job_id: str, pdf_paths: List[str], filter_string: str, batch_size: int = 5, model: str = "gemini-3-flash-preview", max_workers: int = 12):
    """Background task to split multiple PDFs using new splitting method"""
    job = job_status[job_id]
    
    try:
        job.status = "processing"
        job.message = f"Starting PDF splitting for {len(pdf_paths)} file(s)..."
        job.progress = 10
        
        # Use existing output folder
        current_dir = Path(__file__).parent / "current"
        output_dir = current_dir / "output"
        
        # Create output folder if it doesn't exist
        output_dir.mkdir(exist_ok=True)
        
        # Clear output folder first
        if output_dir.exists():
            for file in output_dir.glob("*.pdf"):
                file.unlink()
        
        job.message = f"Analyzing {len(pdf_paths)} PDF(s) with Gemini (filter: '{filter_string}')..."
        job.progress = 30
        
        # Set up environment for the script
        env = os.environ.copy()
        env['PYTHONPATH'] = str(current_dir)
        # Copy API keys
        if 'GOOGLE_API_KEY' in os.environ:
            env['GOOGLE_API_KEY'] = os.environ['GOOGLE_API_KEY']
        
        # Limit OpenBLAS threads to prevent resource exhaustion
        env['OPENBLAS_NUM_THREADS'] = '12'
        env['OMP_NUM_THREADS'] = '12'
        env['MKL_NUM_THREADS'] = '12'
        
        # Path to Gemini splitting script
        script_path = current_dir / "1-split_pdf_gemini.py"
        
        if not script_path.exists():
            raise Exception(f"Gemini splitting script not found: {script_path}")
        
        job.message = f"Analyzing PDF pages..."
        job.progress = 40
        
        # Import and call the function directly
        import sys
        import importlib.util
        spec = importlib.util.spec_from_file_location("split_pdf_gemini", script_path)
        split_module = importlib.util.module_from_spec(spec)
        sys.modules["split_pdf_gemini"] = split_module
        spec.loader.exec_module(split_module)
        split_pdf_with_gemini = split_module.split_pdf_with_gemini
        
        logger.info(f"Running Gemini split function with filter: {filter_string}")
        logger.info(f"Input PDFs: {len(pdf_paths)} file(s)")
        logger.info(f"Output folder: {output_dir}")
        logger.info(f"Batch size: {batch_size}")
        logger.info(f"Model: {model}")
        logger.info(f"Max workers: {max_workers}")
        
        # Process each PDF
        total_created = 0
        filter_strings = [filter_string]
        
        for idx, pdf_path in enumerate(pdf_paths):
            job.message = f"Gemini processing PDF {idx+1}/{len(pdf_paths)}..."
            job.progress = 30 + int((idx / len(pdf_paths)) * 50)  # 30-80% for processing
            
            logger.info(f"Processing PDF {idx+1}/{len(pdf_paths)}: {pdf_path}")
            created_count = split_pdf_with_gemini(
                pdf_path, str(output_dir), filter_strings, batch_size, model, max_workers
            )
            
            if created_count is None:
                logger.warning(f"Gemini splitting returned None for PDF {idx+1}, continuing...")
            else:
                total_created += created_count or 0
                logger.info(f"PDF {idx+1} created {created_count} sections")
        
        if total_created == 0:
            raise Exception("Gemini splitting failed - no sections created from any PDF")
        
        job.message = "Creating ZIP archive of split PDFs..."
        job.progress = 85
        
        # Create ZIP file with all split PDFs
        zip_path = Path(f"/tmp/results/{job_id}_split_pdfs_gemini.zip")
        zip_path.parent.mkdir(exist_ok=True)
        
        # Find all PDF files created by the script
        pdf_files = list(output_dir.glob("*.pdf"))
        
        if not pdf_files:
            raise Exception("No PDF files were created by the splitting script. This could mean no matching pages were found.")
        
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            for pdf_file in pdf_files:
                zipf.write(pdf_file, pdf_file.name)
        
        job.result_file = str(zip_path)
        job.status = "completed"
        job.progress = 100
        job.message = f"Splitting completed successfully! Created {len(pdf_files)} sections from {len(pdf_paths)} PDF(s)."
        
        # Clean up uploaded files
        for pdf_path in pdf_paths:
            if os.path.exists(pdf_path):
                os.unlink(pdf_path)
                logger.info(f"Cleaned up input file: {pdf_path}")
        
        # Force garbage collection to free memory
        gc.collect()
        logger.info(f"Job {job_id}: Memory cleaned up after Gemini PDF splitting")
        
    except Exception as e:
        job.status = "failed"
        job.error = str(e)
        job.message = f"PDF splitting failed: {str(e)}"
        logger.error(f"Split job {job_id} failed: {str(e)}")
        
        # Clean up on error
        try:
            for pdf_path in pdf_paths:
                if os.path.exists(pdf_path):
                    os.unlink(pdf_path)
                    logger.info(f"Cleaned up input file on error: {pdf_path}")
        except Exception as cleanup_error:
            logger.error(f"Cleanup error: {cleanup_error}")
        
        # Clean up memory even on failure
        gc.collect()

def split_pdf_gemini_prompt_background(job_id: str, pdf_paths: List[str], custom_prompt: str, batch_size: int = 5, model: str = "gemini-3-flash-preview", max_workers: int = 12, detection_shift: int = 0):
    """Background task to split multiple PDFs using Gemini with custom prompt - PARALLEL processing"""
    job = job_status[job_id]

    try:
        job.status = "processing"
        job.message = f"Starting parallel PDF splitting for {len(pdf_paths)} file(s)..."
        job.progress = 10

        # Use job-specific output folder
        current_dir = Path(__file__).parent / "current"
        output_dir = current_dir / "output" / job_id  # Job-specific folder

        # Create output folder if it doesn't exist
        output_dir.mkdir(parents=True, exist_ok=True)

        shift_msg = f" (shift: {detection_shift})" if detection_shift != 0 else ""
        job.message = f"Analyzing {len(pdf_paths)} PDF(s) with Gemini prompt{shift_msg}..."
        job.progress = 20

        # Set up environment for the script
        env = os.environ.copy()
        env['PYTHONPATH'] = str(current_dir)
        # Copy API keys
        if 'GOOGLE_API_KEY' in os.environ:
            env['GOOGLE_API_KEY'] = os.environ['GOOGLE_API_KEY']

        # Path to Gemini prompt splitting script
        script_path = current_dir / "1-split_pdf_gemini_prompt.py"

        if not script_path.exists():
            raise Exception(f"Gemini prompt splitting script not found: {script_path}")

        # Import and call the function directly
        import sys
        import importlib.util
        spec = importlib.util.spec_from_file_location("split_pdf_gemini_prompt", script_path)
        split_module = importlib.util.module_from_spec(spec)
        sys.modules["split_pdf_gemini_prompt"] = split_module
        spec.loader.exec_module(split_module)
        split_pdf_with_gemini_prompt = split_module.split_pdf_with_gemini_prompt

        logger.info(f"Running parallel Gemini prompt split function")
        logger.info(f"Input PDFs: {len(pdf_paths)} file(s)")
        logger.info(f"Output folder: {output_dir} (job-specific)")
        logger.info(f"Batch size: {batch_size}")
        logger.info(f"Model: {model}")
        logger.info(f"Max workers per PDF: {max_workers}")
        logger.info(f"Parallel PDFs: {min(20, len(pdf_paths))} concurrent")
        logger.info(f"Detection shift: {detection_shift}")
        logger.info(f"Prompt preview: {custom_prompt[:100]}...")

        job.message = f"Processing {len(pdf_paths)} PDFs with up to 20 parallel Gemini requests..."
        job.progress = 30

        # Process PDFs in parallel (up to 20 concurrent)
        from concurrent.futures import ThreadPoolExecutor, as_completed
        total_created = 0
        successful_pdfs = 0
        failed_pdfs = 0

        # Thread-safe progress tracking
        import threading
        progress_lock = threading.Lock()
        completed_pdfs = 0

        def update_progress(pdf_idx, pdf_name, created_count):
            nonlocal completed_pdfs, total_created, successful_pdfs, failed_pdfs
            with progress_lock:
                completed_pdfs += 1
                if created_count is not None:
                    total_created += created_count or 0
                    successful_pdfs += 1
                else:
                    failed_pdfs += 1

                progress_pct = 30 + int((completed_pdfs / len(pdf_paths)) * 60)  # 30-90% for processing
                job.progress = progress_pct
                job.message = f"Processed {completed_pdfs}/{len(pdf_paths)} PDFs (Success: {successful_pdfs}, Failed: {failed_pdfs})"

                logger.info(f"PDF {pdf_idx+1}/{len(pdf_paths)} ({pdf_name}): {created_count or 0} sections created")

        # Process PDFs in parallel with up to 20 concurrent threads
        pdf_workers = min(20, len(pdf_paths))  # Up to 20 parallel PDFs
        with ThreadPoolExecutor(max_workers=pdf_workers) as executor:
            # Submit all PDF processing tasks
            future_to_pdf = {}
            for idx, pdf_path in enumerate(pdf_paths):
                future = executor.submit(
                    split_pdf_with_gemini_prompt,
                    pdf_path, str(output_dir), custom_prompt, batch_size, model, max_workers, detection_shift
                )
                future_to_pdf[future] = (idx, pdf_path)

            # Collect results as they complete
            for future in as_completed(future_to_pdf):
                idx, pdf_path = future_to_pdf[future]
                pdf_name = Path(pdf_path).name

                try:
                    created_count = future.result()
                    update_progress(idx, pdf_name, created_count)

                except Exception as e:
                    logger.error(f"Exception processing PDF {idx+1} ({pdf_name}): {str(e)}")
                    update_progress(idx, pdf_name, None)

        if total_created == 0:
            raise Exception(f"Gemini prompt splitting failed - no sections created from {len(pdf_paths)} PDF(s)")

        logger.info(f"Parallel processing complete: {successful_pdfs} successful, {failed_pdfs} failed, {total_created} total sections")

        job.message = "Creating ZIP archive of split PDFs..."
        job.progress = 95

        # Create ZIP file with all split PDFs
        zip_path = Path(f"/tmp/results/{job_id}_split_pdfs_gemini_prompt.zip")
        zip_path.parent.mkdir(exist_ok=True)

        # Find all PDF files created by the script
        pdf_files = list(output_dir.glob("*.pdf"))

        if not pdf_files:
            raise Exception("No PDF files were created by the splitting script. This could mean no matching pages were found.")

        with zipfile.ZipFile(zip_path, 'w') as zipf:
            for pdf_file in pdf_files:
                zipf.write(pdf_file, pdf_file.name)

        job.result_file = str(zip_path)
        job.status = "completed"
        job.progress = 100
        job.message = f"Splitting completed successfully! Created {len(pdf_files)} sections from {len(pdf_paths)} PDF(s)."

        # Clean up uploaded files
        for pdf_path in pdf_paths:
            if os.path.exists(pdf_path):
                try:
                    os.unlink(pdf_path)
                except:
                    pass

        logger.info(f"Parallel Gemini prompt splitting completed successfully for job {job_id}")

    except Exception as e:
        logger.error(f"Gemini prompt splitting error: {str(e)}")
        job.status = "failed"
        job.error = str(e)
        job.message = f"Error: {str(e)}"

        # Clean up uploaded files even on failure
        for pdf_path in pdf_paths:
            if os.path.exists(pdf_path):
                try:
                    os.unlink(pdf_path)
                except:
                    pass

        # Clean up memory even on failure
        gc.collect()

def split_pdf_ocrspace_background(job_id: str, pdf_paths: List[str], filter_string: str, max_workers: int = 7, detection_shift: int = 0):
    """Background task to split multiple PDFs using OCR.space API"""
    job = job_status[job_id]
    
    try:
        job.status = "processing"
        job.message = f"Starting PDF splitting with OCR.space for {len(pdf_paths)} file(s)..."
        job.progress = 10
        
        # Use job-specific output folder (prevents conflicts with concurrent users)
        current_dir = Path(__file__).parent / "current"
        output_dir = current_dir / "output" / job_id  # Job-specific folder
        
        # Create output folder if it doesn't exist
        output_dir.mkdir(parents=True, exist_ok=True)
        
        shift_msg = f" (shift: {detection_shift})" if detection_shift != 0 else ""
        job.message = f"Analyzing {len(pdf_paths)} PDF(s) with OCR.space API (filter: '{filter_string}'{shift_msg})..."
        job.progress = 10
        
        # Path to OCR.space splitting script
        script_path = current_dir / "1-split_pdf_ocrspace.py"
        
        if not script_path.exists():
            raise Exception(f"OCR.space splitting script not found: {script_path}")
        
        job.message = f"OCR processing PDF pages..."
        job.progress = 20
        
        # Import and call the function directly (like Gemini) for progress tracking
        import sys
        import importlib.util
        spec = importlib.util.spec_from_file_location("split_pdf_ocrspace", script_path)
        split_module = importlib.util.module_from_spec(spec)
        sys.modules["split_pdf_ocrspace"] = split_module
        spec.loader.exec_module(split_module)
        split_pdf_with_ocrspace = split_module.split_pdf_with_ocrspace
        
        logger.info(f"Running OCR.space split function with filter: {filter_string}, shift: {detection_shift}")
        logger.info(f"Input PDFs: {len(pdf_paths)} file(s)")
        logger.info(f"Output folder: {output_dir} (job-specific)")
        logger.info(f"Max workers: {max_workers}")
        
        # Process each PDF
        total_created = 0
        filter_strings = [filter_string]
        
        for idx, pdf_path in enumerate(pdf_paths):
            # Progress callback function - maps OCR progress (0-100%) to job progress (20-90%)
            # OCR processing: 20-90% (70% range)
            # ZIP creation: 90-100% (10% range)
            # Use a lambda with default argument to capture idx correctly
            def make_progress_callback(pdf_idx, total_pdfs):
                def progress_callback(completed, total, message):
                    if total > 0:
                        # Calculate progress within this PDF's portion
                        pdf_progress = (pdf_idx / total_pdfs) * 70  # Progress through all PDFs
                        ocr_progress_pct = int((completed / total) * (70 / total_pdfs))  # Progress for this PDF
                        job.progress = 20 + int(pdf_progress) + ocr_progress_pct  # Map to 20-90% range
                        job.message = f"OCR.space: PDF {pdf_idx+1}/{total_pdfs} - {message} ({completed}/{total} pages)"
                return progress_callback
            
            logger.info(f"Processing PDF {idx+1}/{len(pdf_paths)}: {pdf_path}")
            created_count = split_pdf_with_ocrspace(
                pdf_path, str(output_dir), filter_strings, max_workers, False, make_progress_callback(idx, len(pdf_paths)), detection_shift
            )
            
            if created_count is None:
                logger.warning(f"OCR.space splitting returned None for PDF {idx+1}, continuing...")
            else:
                total_created += created_count or 0
                logger.info(f"PDF {idx+1} created {created_count} sections")
        
        if total_created == 0:
            raise Exception("OCR.space splitting failed - no sections created from any PDF")
        
        job.message = "Creating ZIP archive of split PDFs..."
        job.progress = 90
        
        # Create ZIP file with all split PDFs
        zip_path = Path(f"/tmp/results/{job_id}_split_pdfs_ocrspace.zip")
        zip_path.parent.mkdir(exist_ok=True)
        
        # Find all PDF files created by the script
        pdf_files = list(output_dir.glob("*.pdf"))
        
        if not pdf_files:
            raise Exception("No PDF files were created by the splitting script. This could mean no matching pages were found.")
        
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            for pdf_file in pdf_files:
                zipf.write(pdf_file, pdf_file.name)
        
        job.result_file = str(zip_path)
        job.status = "completed"
        job.progress = 100
        job.message = f"Splitting completed successfully! Created {len(pdf_files)} sections from {len(pdf_paths)} PDF(s)."
        
        # Clean up uploaded files
        for pdf_path in pdf_paths:
            if os.path.exists(pdf_path):
                os.unlink(pdf_path)
                logger.info(f"Cleaned up input file: {pdf_path}")
        
        # Force garbage collection to free memory
        gc.collect()
        logger.info(f"Job {job_id}: Memory cleaned up after OCR.space PDF splitting")
        
    except Exception as e:
        job.status = "failed"
        job.error = str(e)
        job.message = f"PDF splitting failed: {str(e)}"
        logger.error(f"Split job {job_id} failed: {str(e)}")
        
        # Clean up on error
        try:
            for pdf_path in pdf_paths:
                if os.path.exists(pdf_path):
                    os.unlink(pdf_path)
                    logger.info(f"Cleaned up input file on error: {pdf_path}")
        except Exception as cleanup_error:
            logger.error(f"Cleanup error: {cleanup_error}")
        
        # Clean up memory even on failure
        gc.collect()

def predict_cpt_background(job_id: str, csv_path: str, client: str = "uni"):
    """Background task to predict CPT codes from CSV procedures"""
    job = job_status[job_id]
    
    try:
        job.status = "processing"
        job.message = "Starting CPT code prediction..."
        job.progress = 10
        
        # Import required modules for CPT prediction
        import pandas as pd
        from concurrent.futures import ThreadPoolExecutor, as_completed
        from google import genai
        from google.genai import types
        
        # Setup clients
        # Use JSON string from environment variable (Option 1)
        import json
        import tempfile
        credentials_json = os.environ.get("GOOGLE_CREDENTIALS_JSON")
        if credentials_json:
            with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as f:
                f.write(credentials_json)
                os.environ["GOOGLE_APPLICATION_CREDENTIALS"] = f.name
            
            genai_client = genai.Client(
                vertexai=True,
                project="835764687231",
                location="us-central1",
            )
        else:
            # Fallback to API key method
            genai_client = genai.Client(vertexai=True, api_key="AQ.Ab8RN6LnO1TE5YbcCw1PLVGe2qxhL7TuOVtVm3GnhXndEM0nsw")
        
        fallback_client = genai.Client(vertexai=True, api_key="AQ.Ab8RN6LnO1TE5YbcCw1PLVGe2qxhL7TuOVtVm3GnhXndEM0nsw")
        
        # Client model mapping
        client_models = {
            "uni": "projects/835764687231/locations/us-central1/endpoints/4576355411292061696",
            "sio-stl": "projects/835764687231/locations/us-central1/endpoints/6830407024790994944",
            "gap-fin": "projects/835764687231/locations/us-central1/endpoints/8077904121572622336",
            "apo-utp": "projects/835764687231/locations/us-central1/endpoints/8135325016821596160"
        }
        
        custom_model = client_models.get(client, client_models["uni"])
        fallback_model = "gemini-3-pro-preview"
        
        logger.info(f"Using client model for {client}: {custom_model}")
        
        generate_content_config = types.GenerateContentConfig(
            temperature=1,
            top_p=1,
            seed=0,
            max_output_tokens=50,
            safety_settings=[
                types.SafetySetting(category="HARM_CATEGORY_HATE_SPEECH", threshold="OFF"),
                types.SafetySetting(category="HARM_CATEGORY_DANGEROUS_CONTENT", threshold="OFF"),
                types.SafetySetting(category="HARM_CATEGORY_SEXUALLY_EXPLICIT", threshold="OFF"),
                types.SafetySetting(category="HARM_CATEGORY_HARASSMENT", threshold="OFF"),
            ],
            thinking_config=types.ThinkingConfig(thinking_budget=0),
        )
        
        def format_asa_code(code):
            """Format ASA code with leading zeros based on length"""
            if not code or not isinstance(code, str):
                return code
            
            # Clean the code - remove any non-numeric characters except leading zeros
            cleaned_code = code.strip()
            
            # Extract only numeric characters
            numeric_code = ''.join(filter(str.isdigit, cleaned_code))
            
            if not numeric_code:
                return code  # Return original if no numbers found
            
            # Apply formatting rules
            if len(numeric_code) == 4:
                # 4 numbers: add 1 leading zero
                return f"0{numeric_code}"
            elif len(numeric_code) == 3:
                # 3 numbers: add 2 leading zeros
                return f"00{numeric_code}"
            else:
                # Return as is for other lengths
                return numeric_code
        
        def apply_colonoscopy_correction(row, predicted_code, insurances_df):
            """
            Apply colonoscopy-specific correction rules based on procedure type,
            insurance plan, and polyp findings.
            
            Args:
                row: DataFrame row containing procedure information
                predicted_code: The AI-predicted CPT code
                insurances_df: DataFrame containing insurance information
            
            Returns:
                Corrected CPT code or original predicted code if no correction applies
            """
            try:
                # Get the required fields from the row
                is_colonoscopy = str(row.get('is_colonoscopy', '')).strip().upper() == 'TRUE'
                colonoscopy_is_screening = str(row.get('colonoscopy_is_screening', '')).strip().upper() == 'TRUE'
                is_upper_endonoscopy = str(row.get('is_upper_endonoscopy', '')).strip().upper() == 'TRUE'
                polyps_found = str(row.get('Polyps found', '')).strip().upper() == 'FOUND'
                primary_mednet_code = str(row.get('Primary Mednet Code', '')).strip()
                
                # Priority rule: if both upper endoscopy and colonoscopy, always return 00813
                if is_upper_endonoscopy and is_colonoscopy:
                    logger.info(f"Colonoscopy correction: Both upper endoscopy and colonoscopy detected -> 00813")
                    return "00813"
                
                # Only proceed if it's a colonoscopy
                if not is_colonoscopy:
                    return predicted_code
                
                # Check if insurance is Medicare
                is_medicare = False
                if primary_mednet_code and not insurances_df.empty:
                    # Find the insurance plan by MedNet Code
                    insurance_match = insurances_df[insurances_df['MedNet Code'].astype(str).str.strip() == primary_mednet_code]
                    if not insurance_match.empty:
                        insurance_plan = str(insurance_match.iloc[0].get('Insurance Plan', '')).strip()
                        if 'Medicare' in insurance_plan or 'MEDICARE' in insurance_plan or 'medicare' in insurance_plan:
                            is_medicare = True
                            logger.info(f"Colonoscopy correction: Medicare insurance detected ({insurance_plan})")
                
                # Apply correction rules
                if is_medicare:
                    # MEDICARE rules
                    if colonoscopy_is_screening and not polyps_found:
                        logger.info(f"Colonoscopy correction: Medicare + screening + no polyps -> 00812")
                        return "00812"
                    elif colonoscopy_is_screening and polyps_found:
                        logger.info(f"Colonoscopy correction: Medicare + screening + polyps found -> 00811")
                        return "00811"
                    else:  # not screening (polyps don't matter)
                        logger.info(f"Colonoscopy correction: Medicare + not screening -> 00811")
                        return "00811"
                else:
                    # NOT MEDICARE rules
                    if colonoscopy_is_screening and not polyps_found:
                        logger.info(f"Colonoscopy correction: Non-Medicare + screening + no polyps -> 00812")
                        return "00812"
                    elif colonoscopy_is_screening and polyps_found:
                        logger.info(f"Colonoscopy correction: Non-Medicare + screening + polyps found -> 00812")
                        return "00812"
                    else:  # not screening (polyps don't matter)
                        logger.info(f"Colonoscopy correction: Non-Medicare + not screening -> 00811")
                        return "00811"
                
            except Exception as e:
                logger.warning(f"Colonoscopy correction failed: {str(e)}, returning original prediction")
                return predicted_code

        def get_prediction_and_review(procedure, extracted_description=None, retries=5):
            """Two-stage prediction: 1) Custom model predicts, 2) Gemini Flash reviews"""
            
            # Check if input is invalid (None, empty, or less than 3 characters)
            procedure_str = str(procedure) if procedure is not None else ""
            procedure_str = procedure_str.strip()
            
            if not procedure_str or len(procedure_str) < 3:
                # Skip custom model and go directly to fallback for invalid input
                initial_prediction = None
                model_source = None
                failure_reason = f"Input too short or empty (length: {len(procedure_str)})"
            else:
                # Stage 1: Get initial prediction from custom model
                initial_prediction = None
                model_source = None
                failure_reason = None
            
            prompt = f'For this procedure: "{procedure_str}" give me the most appropriate anesthesia CPT code'
            last_error = "Unknown error"

            # Try custom model first (only if input is valid)
            if initial_prediction is None and procedure_str and len(procedure_str) >= 3:
                for attempt in range(retries):
                    try:
                        response = genai_client.models.generate_content(
                            model=custom_model,
                            contents=[types.Content(role="user", parts=[{"text": prompt}])],
                            config=generate_content_config
                        )
                        result = response.text
                        if result and result.strip().startswith("0"):
                            initial_prediction = format_asa_code(result.strip())
                            model_source = "base_model"
                            break
                        else:
                            failure_reason = f"Base model returned invalid format: '{result.strip()}'"
                    except Exception as e:
                        last_error = str(e)
                        failure_reason = f"Base model error (attempt {attempt + 1}/{retries}): {str(e)}"

            # Determine what to use for fallback model: ALWAYS prefer extracted description if available
            fallback_procedure_str = None
            if extracted_description:
                extracted_str = str(extracted_description) if extracted_description is not None else ""
                extracted_str = extracted_str.strip()
                if extracted_str and len(extracted_str) >= 3:
                    fallback_procedure_str = extracted_str
                    logger.info(f"Fallback model will use 'Procedure Description Extracted'")
            
            # If no extracted description available, use original procedure description
            if not fallback_procedure_str:
                fallback_procedure_str = procedure_str
            
            # Track if base model returned 01926 (to trigger fallback)
            base_returned_01926 = (initial_prediction == "01926")
            
            # Special case: If base model returned 01926, trigger fallback with extracted description
            if base_returned_01926:
                if fallback_procedure_str and fallback_procedure_str != procedure_str:
                    logger.info(f"Base model returned 01926, switching to extracted description and using fallback model")
                elif not fallback_procedure_str or len(fallback_procedure_str) < 3:
                    failure_reason = f"Base model returned 01926, but extracted description is invalid or unavailable"
                    # Keep 01926 if no valid extracted description - don't trigger fallback
                else:
                    logger.info(f"Base model returned 01926, using fallback model with original procedure")
                
                # Reset to trigger fallback (only if we have valid fallback procedure)
                if fallback_procedure_str and len(fallback_procedure_str) >= 3:
                    initial_prediction = None

            # Fallback to production model if custom model failed OR returned 01926
            if not initial_prediction:
                try:
                    # Use extracted description for fallback if available, otherwise use original
                    fallback_prompt = f'For this procedure: "{fallback_procedure_str}" give me the most appropriate anesthesia CPT code'
                    
                    # Add web search to fallback model
                    fallback_tools = [
                        types.Tool(googleSearch=types.GoogleSearch()),
                    ]
                    fallback_config = types.GenerateContentConfig(
                        tools=fallback_tools
                    )
                    
                    response = fallback_client.models.generate_content(
                        model=fallback_model,
                        contents=[types.Content(role="user", parts=[{"text": fallback_prompt + "Only answer with the code, absolutely nothing else, no other text."}])],
                        config=fallback_config
                    )
                    fallback_result = response.text
                    if fallback_result:
                        initial_prediction = format_asa_code(fallback_result.strip())
                        # Track if this was triggered by 01926
                        if base_returned_01926:
                            model_source = "fallback_01926_triggered"
                        else:
                            model_source = "fallback"
                    else:
                        # If fallback fails and base returned 01926, return 01926
                        if base_returned_01926:
                            initial_prediction = "01926"
                            model_source = "base_model"
                            failure_reason = f"Base model returned 01926, fallback model returned empty response"
                        else:
                            return f"Prediction failed: empty response", "error", f"Fallback model returned empty response. Base model failure: {failure_reason}"
                except Exception as e:
                    # If fallback errors and base returned 01926, return 01926
                    if base_returned_01926:
                        initial_prediction = "01926"
                        model_source = "base_model"
                        failure_reason = f"Base model returned 01926, fallback model error: {str(e)}"
                    else:
                        return f"Prediction failed: {str(e)}", "error", f"Fallback model error: {str(e)}. Base model failure: {failure_reason}"

            # Stage 2: Review with custom instructions (if available)
            try:
                # Load custom instructions
                instructions_file = Path("data/cpt_instructions.json")
                custom_instructions = ""
                if instructions_file.exists():
                    with open(instructions_file, "r", encoding="utf-8") as f:
                        data = json.load(f)
                        custom_instructions = data.get("instructions", "").strip()

                # If no custom instructions, return initial prediction with model source
                if not custom_instructions:
                    return initial_prediction, model_source, failure_reason

                # Create review prompt
                review_prompt = f"""You are a medical coding reviewer. Your job is to REVIEW and potentially CORRECT a CPT code prediction, but ONLY if the custom medical coder instructions specifically apply to this case.

PROCEDURE DESCRIPTION: "{procedure}"
PREDICTED CPT CODE: "{initial_prediction}"

CUSTOM MEDICAL CODER INSTRUCTIONS:
{custom_instructions}

IMPORTANT RULES:
1. Your job is NOT to predict a new code - it's to REVIEW and potentially CORRECT the existing prediction
2. ONLY change the code if the custom instructions specifically mention this type of procedure OR this specific code
3. If the custom instructions don't apply to this specific procedure or code, return the original predicted code unchanged
4. If you do make a correction, return only the corrected code with proper formatting (leading zeros)
5. Return ONLY the CPT code, nothing else

What is your final code decision?


answer ONLY with the code, nothing else"""

                # Use Gemini Flash for review with thinking enabled
                review_config = types.GenerateContentConfig(
                    temperature=0.3,  # Lower temperature for more consistent reviews
                    top_p=0.9,
                    max_output_tokens=50,
                    thinking_config=types.ThinkingConfig(
                        thinking_level="HIGH",
                    ),
                )
                
                review_response = fallback_client.models.generate_content(
                    model="gemini-3-flash-preview",
                    contents=[types.Content(role="user", parts=[{"text": review_prompt}])],
                    config=review_config
                )
                
                reviewed_result = review_response.text.strip()
                
                # Format the reviewed result
                if reviewed_result and reviewed_result.replace("0", "").replace("1", "").replace("2", "").replace("3", "").replace("4", "").replace("5", "").replace("6", "").replace("7", "").replace("8", "").replace("9", "") == "":
                    # It's a numeric code, format it
                    final_result = format_asa_code(reviewed_result)
                    # If review changed the code, mark as reviewed
                    if final_result != initial_prediction:
                        return final_result, "reviewed", failure_reason
                    else:
                        return final_result, model_source, failure_reason
                else:
                    # Not a clean numeric code, return original prediction
                    return initial_prediction, model_source, failure_reason

            except Exception as e:
                logger.warning(f"Review stage failed: {str(e)}, returning initial prediction")
                return initial_prediction, model_source, failure_reason
        
        job.message = "Reading CSV file..."
        job.progress = 20
        
        # Read CSV file with dtype=str to preserve leading zeros in MedNet codes
        try:
            df = pd.read_csv(csv_path, encoding='utf-8', dtype=str)
        except UnicodeDecodeError:
            try:
                df = pd.read_csv(csv_path, encoding='latin-1', dtype=str)
            except Exception as e:
                raise Exception(f"Could not read CSV with utf-8 or latin-1 encoding: {e}")

        if "Procedure Description" not in df.columns:
            raise Exception("CSV file missing 'Procedure Description' column")
        
        # Check if "Procedure Description Extracted" column exists
        has_extracted = "Procedure Description Extracted" in df.columns
        if has_extracted:
            logger.info("Found 'Procedure Description Extracted' column - will use for 01926 fallback")
        
        # Load insurances.csv for colonoscopy correction
        job.message = "Loading insurance data for colonoscopy corrections..."
        insurances_df = pd.DataFrame()
        try:
            insurances_path = Path(__file__).parent / "insurances.csv"
            if insurances_path.exists():
                insurances_df = pd.read_csv(insurances_path, dtype=str)
                logger.info(f"Loaded {len(insurances_df)} insurance records for colonoscopy correction")
            else:
                logger.warning(f"insurances.csv not found at {insurances_path}, colonoscopy corrections may not work properly")
        except Exception as e:
            logger.warning(f"Failed to load insurances.csv: {str(e)}, colonoscopy corrections may not work properly")

        job.message = f"Processing {len(df)} procedures..."
        job.progress = 30
        
        # Process predictions with threading
        predictions = [None] * len(df)
        model_sources = [None] * len(df)
        failure_reasons = [None] * len(df)
        
        with ThreadPoolExecutor(max_workers=5) as executor:
            # Pass extracted description if available
            if has_extracted:
                futures = {
                    executor.submit(
                        get_prediction_and_review, 
                        proc, 
                        df.iloc[i].get("Procedure Description Extracted", None)
                    ): i 
                    for i, proc in enumerate(df["Procedure Description"])
                }
            else:
                futures = {executor.submit(get_prediction_and_review, proc): i for i, proc in enumerate(df["Procedure Description"])}
            
            completed = 0
            for future in as_completed(futures):
                idx = futures[future]
                result = future.result()
                if isinstance(result, tuple) and len(result) == 3:
                    predictions[idx], model_sources[idx], failure_reasons[idx] = result
                elif isinstance(result, tuple) and len(result) == 2:
                    # Handle old format for backward compatibility
                    predictions[idx], model_sources[idx] = result
                    failure_reasons[idx] = None
                else:
                    # Handle single return value
                    predictions[idx] = result
                    model_sources[idx] = "unknown"
                    failure_reasons[idx] = None
                completed += 1
                job.progress = 30 + int((completed / len(df)) * 50)
                job.message = f"Processed {completed}/{len(df)} procedures..."

        job.message = "Applying colonoscopy corrections..."
        job.progress = 85
        
        # Apply colonoscopy corrections to predictions (only for uni client)
        corrected_predictions = []
        correction_applied = []
        for idx, row in df.iterrows():
            original_prediction = predictions[idx]
            # Only apply colonoscopy correction for uni client
            if client == "uni":
                corrected_code = apply_colonoscopy_correction(row, original_prediction, insurances_df)
            else:
                corrected_code = original_prediction
            corrected_predictions.append(corrected_code)
            # Track if correction was applied
            correction_applied.append("Yes" if corrected_code != original_prediction else "No")
        
        job.message = "Adding predictions to CSV..."
        job.progress = 90
        
        # Insert predictions, model sources, and failure reasons into dataframe
        insert_index = df.columns.get_loc("Procedure Description") + 1
        df.insert(insert_index, "ASA Code", corrected_predictions)
        df.insert(insert_index + 1, "Procedure Code", corrected_predictions)
        df.insert(insert_index + 2, "Model Source", model_sources)
        df.insert(insert_index + 3, "Colonoscopy Correction Applied", correction_applied)
        df.insert(insert_index + 4, "Base Model Failure Reason", failure_reasons)
        
        # Save result in both formats
        result_base = Path(f"/tmp/results/{job_id}_with_codes")
        result_base.parent.mkdir(exist_ok=True)
        
        result_csv_path, result_xlsx_path = save_dataframe_dual_format(df, result_base)
        
        job.result_file = result_csv_path
        job.result_file_xlsx = result_xlsx_path
        job.status = "completed"
        job.progress = 100
        job.message = f"CPT prediction completed! Processed {len(df)} procedures."
        
        # Clean up input file
        os.unlink(csv_path)
        
        # Force garbage collection to free memory
        gc.collect()
        logger.info(f"Job {job_id}: Memory cleaned up after CPT prediction")
        
    except Exception as e:
        job.status = "failed"
        job.error = str(e)
        job.message = f"CPT prediction failed: {str(e)}"
        logger.error(f"CPT prediction job {job_id} failed: {str(e)}")
        
        # Clean up memory even on failure
        gc.collect()

def predict_cpt_custom_background(job_id: str, csv_path: str, confidence_threshold: float = 0.5):
    """Background task to predict CPT codes using custom TAN-ESC model"""
    job = job_status[job_id]
    
    try:
        job.status = "processing"
        job.message = "Starting custom model CPT prediction..."
        job.progress = 10
        
        # Import the custom prediction function
        import sys
        custom_coding_path = Path(__file__).parent / "custom-coding"
        sys.path.insert(0, str(custom_coding_path))
        
        from predict import predict_codes_api
        
        job.message = f"Loading TAN-ESC model and processing {csv_path}..."
        job.progress = 30
        
        # Create output file path
        result_base = Path(f"/tmp/results/{job_id}_with_codes")
        result_base.parent.mkdir(exist_ok=True)
        
        # Run the prediction using the custom model
        # Model files should be in backend/custom-coding/ directory
        model_dir = custom_coding_path
        
        job.message = "Making predictions with TAN-ESC model..."
        job.progress = 50
        
        # Save to CSV first
        result_file_csv = result_base.with_suffix('.csv')
        success = predict_codes_api(
            input_file=csv_path,
            output_file=str(result_file_csv),
            model_dir=str(model_dir),
            confidence_threshold=confidence_threshold
        )
        
        if not success:
            raise Exception("Custom model prediction failed")
        
        job.message = "Prediction complete, preparing results..."
        job.progress = 90
        
        # Verify output file exists
        if not result_file_csv.exists():
            raise Exception("Output file was not created")
        
        # Create XLSX version
        try:
            result_file_xlsx = convert_csv_to_xlsx(result_file_csv, result_base.with_suffix('.xlsx'))
            job.result_file_xlsx = result_file_xlsx
        except Exception as e:
            logger.warning(f"Could not create XLSX version: {e}")
        
        job.result_file = str(result_file_csv)
        job.status = "completed"
        job.progress = 100
        job.message = f"TAN-ESC prediction completed successfully!"
        
        # Clean up input file
        os.unlink(csv_path)
        
        # Force garbage collection to free memory
        gc.collect()
        logger.info(f"Job {job_id}: Memory cleaned up after custom model prediction")
        
    except Exception as e:
        job.status = "failed"
        job.error = str(e)
        job.message = f"Custom model prediction failed: {str(e)}"
        logger.error(f"Custom model prediction job {job_id} failed: {str(e)}")
        
        # Clean up memory even on failure
        gc.collect()

def predict_cpt_general_background(job_id: str, csv_path: str, model: str = "gpt5", max_workers: int = 50, custom_instructions: str = None):
    """Background task to predict CPT codes using OpenAI general model"""
    job = job_status[job_id]
    
    try:
        job.status = "processing"
        job.message = "Starting OpenAI general model CPT prediction..."
        job.progress = 10
        
        # Import the general prediction function
        import sys
        general_coding_path = Path(__file__).parent / "general-coding"
        sys.path.insert(0, str(general_coding_path))
        
        from predict_general import predict_codes_general_api
        
        job.message = f"Processing with OpenAI {model} model..."
        job.progress = 30
        
        # Create output file path
        result_base = Path(f"/tmp/results/{job_id}_with_codes")
        result_base.parent.mkdir(exist_ok=True)
        
        # Progress callback to update job status
        def progress_callback(completed, total, message):
            job.progress = 30 + int((completed / total) * 60)
            job.message = message
        
        # Save to CSV first
        result_file_csv = result_base.with_suffix('.csv')
        success = predict_codes_general_api(
            input_file=csv_path,
            output_file=str(result_file_csv),
            model=model,
            api_key=os.getenv("OPENAI_API_KEY"),
            max_workers=max_workers,
            progress_callback=progress_callback
        )
        
        if not success:
            raise Exception("OpenAI general model prediction failed")
        
        job.message = "Prediction complete, preparing results..."
        job.progress = 90
        
        # Verify output file exists
        if not result_file_csv.exists():
            raise Exception("Output file was not created")
        
        # Create XLSX version
        try:
            result_file_xlsx = convert_csv_to_xlsx(result_file_csv, result_base.with_suffix('.xlsx'))
            job.result_file_xlsx = result_file_xlsx
        except Exception as e:
            logger.warning(f"Could not create XLSX version: {e}")
        
        job.result_file = str(result_file_csv)
        job.status = "completed"
        job.progress = 100
        job.message = f"OpenAI general model prediction completed successfully!"
        
        # Clean up input file
        os.unlink(csv_path)
        
        # Force garbage collection to free memory
        gc.collect()
        logger.info(f"Job {job_id}: Memory cleaned up after general model prediction")
        
    except Exception as e:
        job.status = "failed"
        job.error = str(e)
        job.message = f"OpenAI general model prediction failed: {str(e)}"
        logger.error(f"General model prediction job {job_id} failed: {str(e)}")
        
        # Clean up memory even on failure
        gc.collect()

def predict_cpt_from_pdfs_background(job_id: str, zip_path: str, n_pages: int = 1, model: str = "openai/gpt-5.2:online", max_workers: int = 50, custom_instructions: str = None):
    """Background task to predict CPT codes from PDF images using OpenAI vision model"""
    job = job_status[job_id]
    
    try:
        job.status = "processing"
        job.message = "Extracting PDFs from ZIP archive..."
        job.progress = 10
        
        # Create temporary directory for processing
        temp_dir = Path(f"/tmp/processing_{job_id}")
        temp_dir.mkdir(exist_ok=True)
        
        # Unzip files
        pdfs_dir = temp_dir / "pdfs"
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(pdfs_dir)
        
        # Log what was extracted for debugging
        extracted_files = list(pdfs_dir.rglob("*"))
        pdf_files_extracted = [f for f in extracted_files if f.is_file() and f.suffix.lower() == '.pdf']
        logger.info(f"Extracted {len(extracted_files)} total items, {len(pdf_files_extracted)} PDF files")
        if pdf_files_extracted:
            logger.info(f"PDF files found: {[f.name for f in pdf_files_extracted]}")
        
        job.message = f"Starting image-based CPT prediction (analyzing {n_pages} page(s) per PDF)..."
        job.progress = 20
        
        # Import the general prediction function
        import sys
        general_coding_path = Path(__file__).parent / "general-coding"
        sys.path.insert(0, str(general_coding_path))
        
        from predict_general import predict_codes_from_pdfs_api, is_gemini_model
        
        # Check if using Gemini model
        using_gemini = is_gemini_model(model)
        
        if using_gemini:
            job.message = f"Processing PDFs with Gemini {model} vision model..."
        else:
            job.message = f"Processing PDFs with OpenRouter {model} vision model..."
        job.progress = 30
        
        # Create output file path
        result_base = Path(f"/tmp/results/{job_id}_with_codes")
        result_base.parent.mkdir(exist_ok=True)
        
        # Progress callback to update job status
        def progress_callback(completed, total, message):
            job.progress = 30 + int((completed / total) * 60)
            job.message = message
        
        # Save to CSV first
        result_file_csv = result_base.with_suffix('.csv')
        # Use GOOGLE_API_KEY for Gemini models, OPENROUTER_API_KEY for others
        if using_gemini:
            api_key = os.getenv("GOOGLE_API_KEY")
        else:
            api_key = os.getenv("OPENROUTER_API_KEY") 
        success = predict_codes_from_pdfs_api(
            pdf_folder=str(temp_dir / "pdfs"),
            output_file=str(result_file_csv),
            n_pages=n_pages,
            model=model,
            api_key=api_key,
            max_workers=max_workers,
            progress_callback=progress_callback,
            custom_instructions=custom_instructions
        )
        
        if not success:
            if using_gemini:
                raise Exception("Gemini vision model prediction failed")
            else:
                raise Exception("OpenRouter vision model prediction failed")
        
        job.message = "Prediction complete, preparing results..."
        job.progress = 90
        
        # Verify output file exists
        if not result_file_csv.exists():
            raise Exception("Output file was not created")
        
        # Create XLSX version
        try:
            result_file_xlsx = convert_csv_to_xlsx(result_file_csv, result_base.with_suffix('.xlsx'))
            job.result_file_xlsx = result_file_xlsx
        except Exception as e:
            logger.warning(f"Could not create XLSX version: {e}")
        
        job.result_file = str(result_file_csv)
        job.status = "completed"
        job.progress = 100
        job.message = f"Vision-based CPT prediction completed successfully!"
        
        # Clean up input files
        os.unlink(zip_path)
        shutil.rmtree(temp_dir)
        
        # Force garbage collection to free memory
        gc.collect()
        logger.info(f"Job {job_id}: Memory cleaned up after vision-based prediction")
        
    except Exception as e:
        job.status = "failed"
        job.error = str(e)
        job.message = f"Vision-based prediction failed: {str(e)}"
        logger.error(f"Vision-based prediction job {job_id} failed: {str(e)}")
        
        # Clean up memory even on failure
        gc.collect()

def predict_icd_from_pdfs_background(job_id: str, zip_path: str, n_pages: int = 1, model: str = "openai/gpt-5.2:online", max_workers: int = 50, custom_instructions: str = None):
    """Background task to predict ICD codes from PDF images using OpenAI vision model"""
    job = job_status[job_id]
    
    try:
        job.status = "processing"
        job.message = "Extracting PDFs from ZIP archive..."
        job.progress = 10
        
        # Create temporary directory for processing
        temp_dir = Path(f"/tmp/processing_{job_id}")
        temp_dir.mkdir(exist_ok=True)
        
        # Unzip files
        pdfs_dir = temp_dir / "pdfs"
        with zipfile.ZipFile(zip_path, 'r') as zip_ref:
            zip_ref.extractall(pdfs_dir)
        
        # Log what was extracted for debugging
        extracted_files = list(pdfs_dir.rglob("*"))
        pdf_files_extracted = [f for f in extracted_files if f.is_file() and f.suffix.lower() == '.pdf']
        logger.info(f"Extracted {len(extracted_files)} total items, {len(pdf_files_extracted)} PDF files")
        if pdf_files_extracted:
            logger.info(f"PDF files found: {[f.name for f in pdf_files_extracted]}")
        
        job.message = f"Starting image-based ICD prediction (analyzing {n_pages} page(s) per PDF)..."
        job.progress = 20
        
        # Import the general prediction function
        import sys
        general_coding_path = Path(__file__).parent / "general-coding"
        sys.path.insert(0, str(general_coding_path))
        
        from predict_general import predict_icd_codes_from_pdfs_api, is_gemini_model
        
        # Check if using Gemini model
        using_gemini = is_gemini_model(model)
        
        if using_gemini:
            job.message = f"Processing PDFs with Gemini {model} vision model..."
        else:
            job.message = f"Processing PDFs with OpenRouter {model} vision model..."
        job.progress = 30
        
        # Create output file path
        result_base = Path(f"/tmp/results/{job_id}_with_icd_codes")
        result_base.parent.mkdir(exist_ok=True)
        
        # Progress callback to update job status
        def progress_callback(completed, total, message):
            job.progress = 30 + int((completed / total) * 60)
            job.message = message
        
        # Save to CSV first
        result_file_csv = result_base.with_suffix('.csv')
        # Use GOOGLE_API_KEY for Gemini models, OPENROUTER_API_KEY for others
        if using_gemini:
            api_key = os.getenv("GOOGLE_API_KEY")
        else:
            api_key = os.getenv("OPENROUTER_API_KEY") 
        success = predict_icd_codes_from_pdfs_api(
            pdf_folder=str(temp_dir / "pdfs"),
            output_file=str(result_file_csv),
            n_pages=n_pages,
            model=model,
            api_key=api_key,
            max_workers=max_workers,
            progress_callback=progress_callback,
            custom_instructions=custom_instructions
        )
        
        if not success:
            if using_gemini:
                raise Exception("Gemini vision model ICD prediction failed")
            else:
                raise Exception("OpenRouter vision model ICD prediction failed")
        
        job.message = "Prediction complete, preparing results..."
        job.progress = 90
        
        # Verify output file exists
        if not result_file_csv.exists():
            raise Exception("Output file was not created")
        
        # Create XLSX version
        try:
            result_file_xlsx = convert_csv_to_xlsx(result_file_csv, result_base.with_suffix('.xlsx'))
            job.result_file_xlsx = result_file_xlsx
        except Exception as e:
            logger.warning(f"Could not create XLSX version: {e}")
        
        job.result_file = str(result_file_csv)
        job.status = "completed"
        job.progress = 100
        job.message = f"Vision-based ICD prediction completed successfully!"
        
        # Clean up input files
        os.unlink(zip_path)
        shutil.rmtree(temp_dir)
        
        # Force garbage collection to free memory
        gc.collect()
        logger.info(f"Job {job_id}: Memory cleaned up after vision-based ICD prediction")
        
    except Exception as e:
        job.status = "failed"
        job.error = str(e)
        job.message = f"Vision-based ICD prediction failed: {str(e)}"
        logger.error(f"Vision-based ICD prediction job {job_id} failed: {str(e)}")
        
        # Clean up memory even on failure
        gc.collect()

@app.post("/upload")
async def upload_files(
    background_tasks: BackgroundTasks,
    zip_file: UploadFile = File(...),
    excel_file: UploadFile = File(None),
    template_id: int = Form(None),
    n_pages: int = Form(..., ge=1, le=50),  # Validate page count between 1-50
    model: str = Form(default="gemini-2.5-flash"),  # Model parameter with default
    worktracker_group: str = Form(None),  # Optional worktracker group field
    worktracker_batch: str = Form(None),  # Optional worktracker batch field
    extract_csn: str = Form(None)  # Optional extract CSN flag
):
    """Upload ZIP file and either Excel instructions file or template ID"""
    
    try:
        # Validate that either excel_file or template_id is provided
        if not excel_file and not template_id:
            raise HTTPException(status_code=400, detail="Either excel_file or template_id must be provided")
        
        if excel_file and template_id:
            raise HTTPException(status_code=400, detail="Provide either excel_file or template_id, not both")
        
        excel_filename = None
        if excel_file:
            excel_filename = excel_file.filename
            logger.info(f"Received upload request - zip: {zip_file.filename}, excel: {excel_file.filename}, pages: {n_pages}, model: {model}")
        else:
            logger.info(f"Received upload request - zip: {zip_file.filename}, template_id: {template_id}, pages: {n_pages}, model: {model}")
        
        # Validate file types
        if not zip_file.filename.endswith('.zip'):
            raise HTTPException(status_code=400, detail="First file must be a ZIP archive")
        
        if excel_file and not excel_file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Excel file must be .xlsx or .xls")
        
        # Generate job ID
        job_id = str(uuid.uuid4())
        job = ProcessingJob(job_id)
        job_status[job_id] = job
        
        logger.info(f"Created job {job_id}")
        
        # Save ZIP file
        zip_path = f"/tmp/{job_id}_archive.zip"
        with open(zip_path, "wb") as f:
            shutil.copyfileobj(zip_file.file, f)
        
        # Handle Excel file or template
        excel_path = f"/tmp/{job_id}_instructions.xlsx"
        
        if excel_file:
            # Save uploaded Excel file
            with open(excel_path, "wb") as f:
                shutil.copyfileobj(excel_file.file, f)
            logger.info(f"Files saved - zip: {zip_path}, excel: {excel_path}")
        else:
            # Fetch template from database and create Excel file
            from db_utils import get_template
            template = get_template(template_id=template_id)
            
            if not template:
                raise HTTPException(status_code=404, detail=f"Template {template_id} not found")
            
            # Extract field definitions from template
            fields = template['template_data'].get('fields', [])
            
            if not fields:
                raise HTTPException(status_code=400, detail="Template has no field definitions")
            
            # Create DataFrame in the expected format
            data = {}
            for field in fields:
                data[field['name']] = [
                    field.get('description', ''),
                    field.get('location', ''),
                    field.get('output_format', ''),
                    'YES' if field.get('priority', False) else 'NO'
                ]
            
            import pandas as pd
            df = pd.DataFrame(data)
            df.to_excel(excel_path, index=False, engine='openpyxl')
            
            excel_filename = f"{template['name']}.xlsx"
            logger.info(f"Created Excel from template - zip: {zip_path}, template: {template['name']}, excel: {excel_path}")
        
        # Parse extract_csn flag
        extract_csn_flag = extract_csn and extract_csn.lower() == "true"
        
        # Start background processing
        background_tasks.add_task(
            process_pdfs_background, 
            job_id, 
            zip_path, 
            excel_path, 
            n_pages, 
            excel_filename, 
            model,
            worktracker_group,
            worktracker_batch,
            extract_csn_flag
        )
        
        logger.info(f"Background task started for job {job_id}")
        
        return {"job_id": job_id, "message": "Files uploaded and processing started"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

@app.get("/status/{job_id}")
async def get_job_status(job_id: str):
    """Get the status of a processing job"""
    if job_id not in job_status:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job = job_status[job_id]
    response = {
        "job_id": job_id,
        "status": job.status,
        "progress": job.progress,
        "message": job.message,
        "error": job.error
    }
    # Include metadata if present (e.g., reasoning for PDF splitting)
    if hasattr(job, 'metadata') and job.metadata:
        response["metadata"] = job.metadata
    # Include result if present (e.g., provider mapping output)
    if hasattr(job, 'result') and job.result:
        response["result"] = job.result
    return response

@app.delete("/cleanup/{job_id}")
async def cleanup_job(job_id: str):
    """Clean up a completed job and its files"""
    if job_id not in job_status:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job = job_status[job_id]
    
    # Clean up CSV result file if it exists
    if job.result_file and os.path.exists(job.result_file):
        try:
            os.unlink(job.result_file)
            logger.info(f"Cleaned up CSV result file: {job.result_file}")
        except Exception as e:
            logger.error(f"Failed to clean up CSV result file: {e}")
    
    # Clean up XLSX result file if it exists
    if job.result_file_xlsx and os.path.exists(job.result_file_xlsx):
        try:
            os.unlink(job.result_file_xlsx)
            logger.info(f"Cleaned up XLSX result file: {job.result_file_xlsx}")
        except Exception as e:
            logger.error(f"Failed to clean up XLSX result file: {e}")
    
    # Clean up JSON result file if it exists
    if hasattr(job, 'result_file_json') and job.result_file_json and os.path.exists(job.result_file_json):
        try:
            os.unlink(job.result_file_json)
            logger.info(f"Cleaned up JSON result file: {job.result_file_json}")
        except Exception as e:
            logger.error(f"Failed to clean up JSON result file: {e}")
    
    # Remove job from memory
    del job_status[job_id]
    logger.info(f"Cleaned up job {job_id} from memory")
    
    return {"message": f"Job {job_id} cleaned up successfully"}

@app.delete("/cleanup-all")
async def cleanup_all_jobs():
    """Clean up all completed jobs and their files"""
    cleaned_count = 0
    total_size_freed = 0
    
    for job_id, job in list(job_status.items()):
        if job.status in ["completed", "failed"]:
            # Clean up CSV result file if it exists
            if job.result_file and os.path.exists(job.result_file):
                try:
                    file_size = os.path.getsize(job.result_file)
                    os.unlink(job.result_file)
                    total_size_freed += file_size
                    logger.info(f"Cleaned up CSV result file: {job.result_file}")
                except Exception as e:
                    logger.error(f"Failed to clean up CSV result file: {e}")
            
            # Clean up XLSX result file if it exists
            if job.result_file_xlsx and os.path.exists(job.result_file_xlsx):
                try:
                    file_size = os.path.getsize(job.result_file_xlsx)
                    os.unlink(job.result_file_xlsx)
                    total_size_freed += file_size
                    logger.info(f"Cleaned up XLSX result file: {job.result_file_xlsx}")
                except Exception as e:
                    logger.error(f"Failed to clean up XLSX result file: {e}")
            
            # Remove job from memory
            del job_status[job_id]
            cleaned_count += 1
    
    logger.info(f"Cleaned up {cleaned_count} jobs, freed {total_size_freed} bytes")
    return {
        "message": f"Cleaned up {cleaned_count} jobs",
        "jobs_cleaned": cleaned_count,
        "bytes_freed": total_size_freed
    }

@app.post("/save-instructions")
async def save_instructions(request: dict):
    """Save custom coding instructions to JSON file"""
    try:
        instructions = request.get("instructions", "")
        
        # Create data directory if it doesn't exist
        data_dir = Path("data")
        data_dir.mkdir(exist_ok=True)
        
        # Save instructions to JSON file
        instructions_file = data_dir / "cpt_instructions.json"
        with open(instructions_file, "w", encoding="utf-8") as f:
            json.dump({"instructions": instructions}, f, ensure_ascii=False, indent=2)
        
        logger.info(f"Saved custom instructions: {len(instructions)} characters")
        return {"message": "Instructions saved successfully"}
        
    except Exception as e:
        logger.error(f"Failed to save instructions: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to save instructions: {str(e)}")

@app.get("/load-instructions")
async def load_instructions():
    """Load custom coding instructions from JSON file"""
    try:
        instructions_file = Path("data/cpt_instructions.json")
        
        if not instructions_file.exists():
            return {"instructions": ""}
        
        with open(instructions_file, "r", encoding="utf-8") as f:
            data = json.load(f)
        
        instructions = data.get("instructions", "")
        logger.info(f"Loaded custom instructions: {len(instructions)} characters")
        return {"instructions": instructions}
        
    except Exception as e:
        logger.error(f"Failed to load instructions: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Failed to load instructions: {str(e)}")

@app.post("/predict-cpt")
async def predict_cpt(
    background_tasks: BackgroundTasks,
    csv_file: UploadFile = File(...),
    client: str = Form(default="uni")
):
    """Upload a CSV or XLSX file to predict CPT codes for procedures"""
    
    try:
        logger.info(f"Received CPT prediction request - file: {csv_file.filename}, client: {client}")
        
        # Validate file type
        if not csv_file.filename.endswith(('.csv', '.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="File must be a CSV or XLSX")
        
        # Generate job ID
        job_id = str(uuid.uuid4())
        job = ProcessingJob(job_id)
        job_status[job_id] = job
        
        logger.info(f"Created CPT prediction job {job_id}")
        
        # Save uploaded file
        input_path = f"/tmp/{job_id}_input{Path(csv_file.filename).suffix}"
        
        with open(input_path, "wb") as f:
            shutil.copyfileobj(csv_file.file, f)
        
        logger.info(f"File saved - path: {input_path}")
        
        # Convert to CSV if needed
        csv_path = ensure_csv_file(input_path, f"/tmp/{job_id}_input.csv")
        
        # Clean up original file if it was converted
        if csv_path != input_path and os.path.exists(input_path):
            os.unlink(input_path)
        
        logger.info(f"CSV ready - path: {csv_path}")
        
        # Start background processing
        background_tasks.add_task(predict_cpt_background, job_id, csv_path, client)
        
        logger.info(f"Background CPT prediction task started for job {job_id}")
        
        return {"job_id": job_id, "message": "File uploaded and CPT prediction started"}
        
    except Exception as e:
        logger.error(f"CPT prediction upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

@app.post("/predict-cpt-custom")
async def predict_cpt_custom(
    background_tasks: BackgroundTasks,
    csv_file: UploadFile = File(...),
    confidence_threshold: float = Form(default=0.5)
):
    """Upload a CSV or XLSX file to predict CPT codes using custom TAN-ESC model"""
    
    try:
        logger.info(f"Received custom model CPT prediction request - file: {csv_file.filename}, threshold: {confidence_threshold}")
        
        # Validate file type
        if not csv_file.filename.endswith(('.csv', '.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="File must be a CSV or XLSX")
        
        # Generate job ID
        job_id = str(uuid.uuid4())
        job = ProcessingJob(job_id)
        job_status[job_id] = job
        
        logger.info(f"Created custom model CPT prediction job {job_id}")
        
        # Save uploaded file
        input_path = f"/tmp/{job_id}_custom_input{Path(csv_file.filename).suffix}"
        
        with open(input_path, "wb") as f:
            shutil.copyfileobj(csv_file.file, f)
        
        logger.info(f"File saved - path: {input_path}")
        
        # Convert to CSV if needed
        csv_path = ensure_csv_file(input_path, f"/tmp/{job_id}_custom_input.csv")
        
        # Clean up original file if it was converted
        if csv_path != input_path and os.path.exists(input_path):
            os.unlink(input_path)
        
        logger.info(f"CSV ready - path: {csv_path}")
        
        # Start background processing with custom model
        background_tasks.add_task(predict_cpt_custom_background, job_id, csv_path, confidence_threshold)
        
        logger.info(f"Background custom model CPT prediction task started for job {job_id}")
        
        return {"job_id": job_id, "message": "File uploaded and TAN-ESC prediction started"}
        
    except Exception as e:
        logger.error(f"Custom model CPT prediction upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

@app.post("/predict-cpt-general")
async def predict_cpt_general(
    background_tasks: BackgroundTasks,
    csv_file: UploadFile = File(...),
    model: str = Form(default="gpt5"),
    max_workers: int = Form(default=50),
    custom_instructions: Optional[str] = Form(default=None),
    instruction_template_id: Optional[int] = Form(default=None)
):
    """Upload a CSV or XLSX file to predict CPT codes using OpenAI general model"""
    
    try:
        logger.info(f"Received general model CPT prediction request - file: {csv_file.filename}, model: {model}, workers: {max_workers}")
        
        # Fetch instruction template if provided
        if instruction_template_id:
            from db_utils import get_prediction_instruction
            template = get_prediction_instruction(instruction_id=instruction_template_id)
            if template:
                custom_instructions = template['instructions_text']
                logger.info(f"Using instruction template '{template['name']}' for general CPT prediction")
            else:
                logger.warning(f"Instruction template {instruction_template_id} not found, proceeding without custom instructions")
        
        # Validate file type
        if not csv_file.filename.endswith(('.csv', '.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="File must be a CSV or XLSX")
        
        # Generate job ID
        job_id = str(uuid.uuid4())
        job = ProcessingJob(job_id)
        job_status[job_id] = job
        
        logger.info(f"Created general model CPT prediction job {job_id}")
        
        # Save uploaded file
        input_path = f"/tmp/{job_id}_general_input{Path(csv_file.filename).suffix}"
        
        with open(input_path, "wb") as f:
            shutil.copyfileobj(csv_file.file, f)
        
        logger.info(f"File saved - path: {input_path}")
        
        # Convert to CSV if needed
        csv_path = ensure_csv_file(input_path, f"/tmp/{job_id}_general_input.csv")
        
        # Clean up original file if it was converted
        if csv_path != input_path and os.path.exists(input_path):
            os.unlink(input_path)
        
        logger.info(f"CSV ready - path: {csv_path}")
        
        # Start background processing with general model
        background_tasks.add_task(predict_cpt_general_background, job_id, csv_path, model, max_workers, custom_instructions)
        
        logger.info(f"Background general model CPT prediction task started for job {job_id}")
        
        return {"job_id": job_id, "message": f"File uploaded and OpenAI {model} prediction started"}
        
    except Exception as e:
        logger.error(f"General model CPT prediction upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

@app.post("/predict-cpt-from-pdfs")
async def predict_cpt_from_pdfs(
    background_tasks: BackgroundTasks,
    zip_file: UploadFile = File(...),
    n_pages: int = Form(default=1, ge=1, le=50),
    model: str = Form(default="openai/gpt-5.2:online"),
    max_workers: int = Form(default=50),
    custom_instructions: Optional[str] = Form(default=None),
    instruction_template_id: Optional[int] = Form(default=None)
):
    """Upload a ZIP file containing PDFs to predict CPT codes using OpenAI vision model"""
    
    try:
        logger.info(f"Received vision-based CPT prediction request - zip: {zip_file.filename}, pages: {n_pages}, model: {model}, workers: {max_workers}")
        
        # Validate file type
        if not zip_file.filename.endswith('.zip'):
            raise HTTPException(status_code=400, detail="File must be a ZIP archive")
        
        # Fetch instruction template if provided
        if instruction_template_id:
            from db_utils import get_prediction_instruction
            template = get_prediction_instruction(instruction_id=instruction_template_id)
            if template:
                custom_instructions = template['instructions_text']
                logger.info(f"Using instruction template '{template['name']}' for CPT prediction")
            else:
                logger.warning(f"Instruction template {instruction_template_id} not found, proceeding without custom instructions")
        
        # Generate job ID
        job_id = str(uuid.uuid4())
        job = ProcessingJob(job_id)
        job_status[job_id] = job
        
        logger.info(f"Created vision-based CPT prediction job {job_id}")
        
        # Save uploaded ZIP
        zip_path = f"/tmp/{job_id}_pdfs.zip"
        
        with open(zip_path, "wb") as f:
            shutil.copyfileobj(zip_file.file, f)
        
        logger.info(f"ZIP saved - path: {zip_path}")
        
        # Start background processing with vision model
        background_tasks.add_task(predict_cpt_from_pdfs_background, job_id, zip_path, n_pages, model, max_workers, custom_instructions)
        
        logger.info(f"Background vision-based CPT prediction task started for job {job_id}")
        
        return {"job_id": job_id, "message": f"ZIP uploaded and OpenAI {model} vision prediction started"}
        
    except Exception as e:
        logger.error(f"Vision-based CPT prediction upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

@app.post("/predict-icd-from-pdfs")
async def predict_icd_from_pdfs(
    background_tasks: BackgroundTasks,
    zip_file: UploadFile = File(...),
    n_pages: int = Form(default=1, ge=1, le=50),
    model: str = Form(default="openai/gpt-5"),
    max_workers: int = Form(default=50),
    custom_instructions: Optional[str] = Form(default=None),
    instruction_template_id: Optional[int] = Form(default=None)
):
    """Upload a ZIP file containing PDFs to predict ICD codes using OpenAI vision model"""
    
    try:
        logger.info(f"Received vision-based ICD prediction request - zip: {zip_file.filename}, pages: {n_pages}, model: {model}, workers: {max_workers}")
        
        # Validate file type
        if not zip_file.filename.endswith('.zip'):
            raise HTTPException(status_code=400, detail="File must be a ZIP archive")
        
        # Fetch instruction template if provided
        if instruction_template_id:
            from db_utils import get_prediction_instruction
            template = get_prediction_instruction(instruction_id=instruction_template_id)
            if template:
                custom_instructions = template['instructions_text']
                logger.info(f"Using instruction template '{template['name']}' for ICD prediction")
            else:
                logger.warning(f"Instruction template {instruction_template_id} not found, proceeding without custom instructions")
        
        # Generate job ID
        job_id = str(uuid.uuid4())
        job = ProcessingJob(job_id)
        job_status[job_id] = job
        
        logger.info(f"Created vision-based ICD prediction job {job_id}")
        
        # Save uploaded ZIP
        zip_path = f"/tmp/{job_id}_icd_pdfs.zip"
        
        with open(zip_path, "wb") as f:
            shutil.copyfileobj(zip_file.file, f)
        
        logger.info(f"ZIP saved - path: {zip_path}")
        
        # Start background processing with vision model
        background_tasks.add_task(predict_icd_from_pdfs_background, job_id, zip_path, n_pages, model, max_workers, custom_instructions)
        
        logger.info(f"Background vision-based ICD prediction task started for job {job_id}")
        
        return {"job_id": job_id, "message": f"ZIP uploaded and OpenAI {model} vision ICD prediction started"}
        
    except Exception as e:
        logger.error(f"Vision-based ICD prediction upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

@app.post("/split-pdf")
async def split_pdf(
    background_tasks: BackgroundTasks,
    pdf_files: List[UploadFile] = File(...),
    filter_string: str = Form(..., description="Text to search for in PDF pages for splitting"),
    detection_shift: int = Form(default=0, description="Shift detections by N pages (positive = down, negative = up)")
):
    """Upload one or more PDF files to split into sections (OCR-based method)"""
    
    try:
        logger.info(f"Received PDF split request - {len(pdf_files)} PDF file(s), shift: {detection_shift}")
        
        # Validate all files are PDFs
        pdf_paths = []
        job_id = str(uuid.uuid4())
        job = ProcessingJob(job_id)
        job_status[job_id] = job
        
        logger.info(f"Created split job {job_id}")
        
        # Save all uploaded PDFs
        for idx, pdf_file in enumerate(pdf_files):
            if not pdf_file.filename.endswith('.pdf'):
                raise HTTPException(status_code=400, detail=f"File {pdf_file.filename} must be a PDF")
            
            pdf_path = f"/tmp/{job_id}_input_{idx}.pdf"
            with open(pdf_path, "wb") as f:
                shutil.copyfileobj(pdf_file.file, f)
            pdf_paths.append(pdf_path)
            logger.info(f"PDF {idx+1}/{len(pdf_files)} saved - path: {pdf_path}")
        
        # Start background processing with all PDFs
        background_tasks.add_task(split_pdf_background, job_id, pdf_paths, filter_string, detection_shift)
        
        logger.info(f"Background split task started for job {job_id} with {len(pdf_files)} PDF(s)")
        
        return {"job_id": job_id, "message": f"{len(pdf_files)} PDF(s) uploaded and splitting started"}
        
    except Exception as e:
        logger.error(f"PDF split upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

@app.post("/split-pdf-gemini")
async def split_pdf_gemini(
    background_tasks: BackgroundTasks,
    pdf_files: List[UploadFile] = File(...),
    filter_string: str = Form(..., description="Text to search for in PDF pages for splitting"),
    batch_size: int = Form(default=5, description="Number of pages to process per API call (1-50)"),
    model: str = Form(default="gemini-2.5-pro", description="Gemini model to use")
):
    """Upload one or more PDF files to split into sections using new splitting method"""
    
    try:
        logger.info(f"Received new PDF split request - {len(pdf_files)} PDF file(s)")
        
        # Validate all files are PDFs
        for pdf_file in pdf_files:
            if not pdf_file.filename.endswith('.pdf'):
                raise HTTPException(status_code=400, detail=f"File {pdf_file.filename} must be a PDF")
        
        # Validate batch size
        if batch_size < 1 or batch_size > 50:
            raise HTTPException(status_code=400, detail="Batch size must be between 1 and 50")
        
        # Validate model
        valid_models = ["gemini-2.5-pro", "gemini-2.5-flash", "gemini-3-flash-preview", "gemini-3-pro-preview", "gemini-flash-latest"]
        if model not in valid_models:
            raise HTTPException(status_code=400, detail=f"Invalid model. Must be one of: {', '.join(valid_models)}")
        
        # Configuration
        max_workers = 12
        
        logger.info(f"Using batch_size: {batch_size}, model: {model}")
        
        # Generate job ID
        job_id = str(uuid.uuid4())
        job = ProcessingJob(job_id)
        job_status[job_id] = job
        
        logger.info(f"Created Gemini split job {job_id}")
        
        # Save all uploaded PDFs
        pdf_paths = []
        for idx, pdf_file in enumerate(pdf_files):
            pdf_path = f"/tmp/{job_id}_input_{idx}.pdf"
            with open(pdf_path, "wb") as f:
                shutil.copyfileobj(pdf_file.file, f)
            pdf_paths.append(pdf_path)
            logger.info(f"PDF {idx+1}/{len(pdf_files)} saved - path: {pdf_path}")
        
        # Start background processing with new splitting method
        background_tasks.add_task(split_pdf_gemini_background, job_id, pdf_paths, filter_string, batch_size, model, max_workers)
        
        logger.info(f"Background split task started for job {job_id} with {len(pdf_files)} PDF(s)")
        
        return {"job_id": job_id, "message": f"{len(pdf_files)} PDF(s) uploaded and splitting started"}
        
    except Exception as e:
        logger.error(f"Gemini PDF split upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

@app.post("/split-pdf-ocrspace")
async def split_pdf_ocrspace(
    background_tasks: BackgroundTasks,
    pdf_files: List[UploadFile] = File(...),
    filter_string: str = Form(..., description="Text to search for in PDF pages for splitting"),
    detection_shift: int = Form(default=0, description="Shift detections by N pages (positive = down, negative = up)")
):
    """Upload one or more PDF files to split using OCR.space API - fastest and most reliable method"""
    
    try:
        logger.info(f"Received OCR.space PDF split request - {len(pdf_files)} PDF file(s), filter: {filter_string}, shift: {detection_shift}")
        
        # Validate all files are PDFs
        for pdf_file in pdf_files:
            if not pdf_file.filename.endswith('.pdf'):
                raise HTTPException(status_code=400, detail=f"File {pdf_file.filename} must be a PDF")
        
        # Fixed configuration
        max_workers = 7  # Optimal speed while staying under API limit (7 concurrent requests)
        
        # Generate job ID
        job_id = str(uuid.uuid4())
        job = ProcessingJob(job_id)
        job_status[job_id] = job
        
        logger.info(f"Created OCR.space split job {job_id}")
        
        # Save all uploaded PDFs
        pdf_paths = []
        for idx, pdf_file in enumerate(pdf_files):
            pdf_path = f"/tmp/{job_id}_input_{idx}.pdf"
            with open(pdf_path, "wb") as f:
                shutil.copyfileobj(pdf_file.file, f)
            pdf_paths.append(pdf_path)
            logger.info(f"PDF {idx+1}/{len(pdf_files)} saved - path: {pdf_path}")
        
        # Start background processing with OCR.space method
        background_tasks.add_task(split_pdf_ocrspace_background, job_id, pdf_paths, filter_string, max_workers, detection_shift)
        
        logger.info(f"Background OCR.space split task started for job {job_id} with {len(pdf_files)} PDF(s)")
        
        return {"job_id": job_id, "message": f"{len(pdf_files)} PDF(s) uploaded and OCR.space splitting started"}
        
    except Exception as e:
        logger.error(f"OCR.space PDF split upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

@app.post("/split-pdf-gemini-prompt")
async def split_pdf_gemini_prompt(
    background_tasks: BackgroundTasks,
    pdf_files: List[UploadFile] = File(...),
    custom_prompt: str = Form(..., description="Custom prompt describing what to look for on pages"),
    batch_size: int = Form(default=5, description="Number of pages to process per API call (1-50)"),
    model: str = Form(default="gemini-3-flash-preview", description="Gemini model to use"),
    detection_shift: int = Form(default=0, description="Shift detections by N pages (positive = down, negative = up)")
):
    """Upload one or more PDF files to split using Gemini with a custom prompt"""
    
    try:
        logger.info(f"Received Gemini prompt PDF split request - {len(pdf_files)} PDF file(s), prompt length: {len(custom_prompt)}")
        
        # Validate all files are PDFs
        for pdf_file in pdf_files:
            if not pdf_file.filename.endswith('.pdf'):
                raise HTTPException(status_code=400, detail=f"File {pdf_file.filename} must be a PDF")
        
        # Validate batch size
        if batch_size < 1 or batch_size > 50:
            raise HTTPException(status_code=400, detail="Batch size must be between 1 and 50")
        
        # Validate model
        valid_models = ["gemini-2.5-pro", "gemini-2.5-flash", "gemini-3-flash-preview", "gemini-3-pro-preview", "gemini-flash-latest"]
        if model not in valid_models:
            raise HTTPException(status_code=400, detail=f"Invalid model. Must be one of: {', '.join(valid_models)}")
        
        # Validate prompt
        if not custom_prompt or not custom_prompt.strip():
            raise HTTPException(status_code=400, detail="Custom prompt cannot be empty")
        
        # Configuration
        max_workers = 12
        
        logger.info(f"Using batch_size: {batch_size}, model: {model}, shift: {detection_shift}")
        
        # Generate job ID
        job_id = str(uuid.uuid4())
        job = ProcessingJob(job_id)
        job_status[job_id] = job
        
        logger.info(f"Created Gemini prompt split job {job_id}")
        
        # Save all uploaded PDFs
        pdf_paths = []
        for idx, pdf_file in enumerate(pdf_files):
            pdf_path = f"/tmp/{job_id}_input_{idx}.pdf"
            with open(pdf_path, "wb") as f:
                shutil.copyfileobj(pdf_file.file, f)
            pdf_paths.append(pdf_path)
            logger.info(f"PDF {idx+1}/{len(pdf_files)} saved - path: {pdf_path}")
        
        # Start background processing with Gemini prompt method
        background_tasks.add_task(split_pdf_gemini_prompt_background, job_id, pdf_paths, custom_prompt, batch_size, model, max_workers, detection_shift)
        
        logger.info(f"Background Gemini prompt split task started for job {job_id} with {len(pdf_files)} PDF(s)")
        
        return {"job_id": job_id, "message": f"{len(pdf_files)} PDF(s) uploaded and Gemini prompt splitting started"}
        
    except Exception as e:
        logger.error(f"Gemini prompt PDF split upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

@app.post("/manual-split-pdf/upload")
async def upload_pdf_for_manual_split(pdf_file: UploadFile = File(...)):
    """
    Upload a PDF file for manual splitting.
    Returns PDF info including page count.
    """
    try:
        if not pdf_file.filename.endswith('.pdf'):
            raise HTTPException(status_code=400, detail="File must be a PDF")
        
        # Save PDF temporarily
        import tempfile
        import fitz  # PyMuPDF
        from PyPDF2 import PdfReader
        
        temp_pdf = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        content = await pdf_file.read()
        temp_pdf.write(content)
        temp_pdf.close()
        
        # Get page count
        reader = PdfReader(temp_pdf.name)
        page_count = len(reader.pages)
        
        # Generate a unique ID for this PDF session
        pdf_session_id = str(uuid.uuid4())
        
        # Store PDF path in a temporary location (will be cleaned up later)
        session_pdf_path = f"/tmp/manual_split_{pdf_session_id}.pdf"
        shutil.copy2(temp_pdf.name, session_pdf_path)
        os.unlink(temp_pdf.name)
        
        # Store session info (in production, use Redis or database)
        if not hasattr(upload_pdf_for_manual_split, 'sessions'):
            upload_pdf_for_manual_split.sessions = {}
        upload_pdf_for_manual_split.sessions[pdf_session_id] = {
            'pdf_path': session_pdf_path,
            'filename': pdf_file.filename,
            'page_count': page_count,
            'created_at': time.time()
        }
        
        return {
            "pdf_session_id": pdf_session_id,
            "filename": pdf_file.filename,
            "page_count": page_count
        }
    
    except Exception as e:
        logger.error(f"Failed to upload PDF for manual split: {e}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

@app.get("/manual-split-pdf/{pdf_session_id}/page/{page_number}")
async def get_pdf_page_image(pdf_session_id: str, page_number: int):
    """
    Get a PDF page as an image for display.
    Returns page image as PNG.
    """
    try:
        import fitz  # PyMuPDF
        import base64
        from io import BytesIO
        
        # Get session info
        if not hasattr(upload_pdf_for_manual_split, 'sessions'):
            raise HTTPException(status_code=404, detail="PDF session not found")
        
        sessions = upload_pdf_for_manual_split.sessions
        if pdf_session_id not in sessions:
            raise HTTPException(status_code=404, detail="PDF session not found")
        
        session_info = sessions[pdf_session_id]
        pdf_path = session_info['pdf_path']
        page_count = session_info['page_count']
        
        # Validate page number (0-based)
        if page_number < 0 or page_number >= page_count:
            raise HTTPException(status_code=400, detail=f"Page number must be between 0 and {page_count - 1}")
        
        # Open PDF and render page
        doc = fitz.open(pdf_path)
        page = doc[page_number]
        
        # Render page as image (scale for thumbnail)
        mat = fitz.Matrix(2.0, 2.0)  # 2x zoom for better quality
        pix = page.get_pixmap(matrix=mat)
        
        # Convert to PNG bytes
        img_bytes = pix.tobytes("png")
        doc.close()
        
        # Return as base64 encoded image
        img_base64 = base64.b64encode(img_bytes).decode('utf-8')
        
        return {
            "page_number": page_number,
            "image_data": f"data:image/png;base64,{img_base64}",
            "width": pix.width,
            "height": pix.height
        }
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get PDF page image: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to render page: {str(e)}")

@app.post("/manual-split-pdf/{pdf_session_id}/split")
async def split_pdf_manual(
    pdf_session_id: str,
    request: dict = Body(...)
):
    """
    Split PDF based on manually selected page indexes.
    selected_pages should be a list of page numbers (0-based) where splits should occur.
    """
    try:
        from PyPDF2 import PdfReader, PdfWriter
        import tempfile
        
        # Get session info
        if not hasattr(upload_pdf_for_manual_split, 'sessions'):
            raise HTTPException(status_code=404, detail="PDF session not found")
        
        sessions = upload_pdf_for_manual_split.sessions
        if pdf_session_id not in sessions:
            raise HTTPException(status_code=404, detail="PDF session not found")
        
        session_info = sessions[pdf_session_id]
        pdf_path = session_info['pdf_path']
        filename = session_info['filename']
        page_count = session_info['page_count']
        
        # Get selected pages from request body
        selected_pages = request.get('selected_pages', [])
        
        # Validate selected pages
        if not selected_pages:
            raise HTTPException(status_code=400, detail="At least one page must be selected")
        
        # Sort and validate page numbers
        selected_pages = sorted([int(p) for p in selected_pages])
        if selected_pages[0] < 0 or selected_pages[-1] >= page_count:
            raise HTTPException(status_code=400, detail=f"Page numbers must be between 0 and {page_count - 1}")
        
        # Read PDF
        reader = PdfReader(pdf_path)
        base_name = os.path.splitext(filename)[0]
        
        # Create output directory
        output_dir = Path("/tmp") / f"manual_split_{pdf_session_id}"
        output_dir.mkdir(exist_ok=True)
        
        # Create PDF sections
        created_pdfs = []
        for i, start_page in enumerate(selected_pages):
            # Determine end page (exclusive)
            if i + 1 < len(selected_pages):
                end_page = selected_pages[i + 1]
            else:
                end_page = page_count  # Last section goes to end
            
            # Create PDF for this section
            writer = PdfWriter()
            
            # Add pages to this section
            for page_idx in range(start_page, end_page):
                writer.add_page(reader.pages[page_idx])
            
            # Save section PDF
            section_filename = f"{base_name}_section_{i + 1:02d}_pages_{start_page + 1}-{end_page}.pdf"
            section_path = output_dir / section_filename
            
            with open(section_path, 'wb') as output_file:
                writer.write(output_file)
            
            created_pdfs.append({
                "filename": section_filename,
                "pages": f"{start_page + 1}-{end_page}",
                "page_count": end_page - start_page
            })
        
        # Create ZIP file with all split PDFs
        zip_path = output_dir / f"{base_name}_split.zip"
        with zipfile.ZipFile(zip_path, 'w') as zipf:
            for pdf_info in created_pdfs:
                pdf_file_path = output_dir / pdf_info['filename']
                zipf.write(pdf_file_path, pdf_info['filename'])
        
        # Return ZIP file
        return FileResponse(
            path=str(zip_path),
            filename=f"{base_name}_split.zip",
            media_type="application/zip"
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to split PDF manually: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to split PDF: {str(e)}")

@app.get("/download/{job_id}")
async def download_result(job_id: str, format: str = "csv", file_type: str = "result_file"):
    """Download the processed file in CSV, XLSX, or JSON format
    
    Args:
        job_id: The job identifier
        format: File format - 'csv' or 'xlsx' (default: 'csv')
        file_type: Which file to download - 'result_file', 'result_file_xlsx', or 'result_file_json' (default: 'result_file')
    """
    if job_id not in job_status:
        raise HTTPException(status_code=404, detail="Job not found")
    
    job = job_status[job_id]
    
    if job.status != "completed":
        raise HTTPException(status_code=400, detail="Job not completed yet")
    
    # Determine which file to download based on file_type and format parameters
    if file_type == "result_file_json":
        # User wants JSON file
        if hasattr(job, 'result_file_json') and job.result_file_json and os.path.exists(job.result_file_json):
            result_file = job.result_file_json
            filename_suffix = "json"
            media_type = "application/json"
        else:
            raise HTTPException(status_code=404, detail="JSON result file not available")
    elif format.lower() == "xlsx" or file_type == "result_file_xlsx":
        # User wants XLSX format
        if not job.result_file_xlsx or not os.path.exists(job.result_file_xlsx):
            # XLSX not available, fallback to CSV
            logger.warning(f"XLSX file not available for job {job_id}, falling back to CSV")
            result_file = job.result_file
            filename_suffix = "csv"
            media_type = "text/csv"
        else:
            result_file = job.result_file_xlsx
            filename_suffix = "xlsx"
            media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        # Default to CSV
        result_file = job.result_file
        filename_suffix = "csv"
        media_type = "text/csv"
    
    if not result_file or not os.path.exists(result_file):
        raise HTTPException(status_code=404, detail="Result file not found")
    
    # Determine filename based on job type
    if result_file.endswith('.zip'):
        filename = f"split_pdfs_{job_id}.zip"
        media_type = "application/zip"
    elif result_file.endswith('.json'):
        filename = f"error_analysis_{job_id}.json"
        media_type = "application/json"
    elif result_file.endswith('.txt'):
        filename = f"error_analysis_{job_id}.txt"
        media_type = "text/plain"
    elif result_file.endswith('.xlsx'):
        filename = f"processed_data_{job_id}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    else:
        filename = f"processed_data_{job_id}.csv"
        media_type = "text/csv"
    
    # Schedule cleanup after download (in background)
    import asyncio
    asyncio.create_task(cleanup_after_download(job_id))
    
    return FileResponse(
        result_file,
        media_type=media_type,
        filename=filename
    )

async def cleanup_after_download(job_id: str):
    """Clean up job after download with a delay"""
    import asyncio
    await asyncio.sleep(30)  # Wait 30 seconds after download
    
    if job_id in job_status:
        job = job_status[job_id]
        
        # Clean up CSV result file
        if job.result_file and os.path.exists(job.result_file):
            try:
                os.unlink(job.result_file)
                logger.info(f"Auto-cleaned up CSV result file: {job.result_file}")
            except Exception as e:
                logger.error(f"Failed to auto-cleanup CSV result file: {e}")
        
        # Clean up XLSX result file
        if job.result_file_xlsx and os.path.exists(job.result_file_xlsx):
            try:
                os.unlink(job.result_file_xlsx)
                logger.info(f"Auto-cleaned up XLSX result file: {job.result_file_xlsx}")
            except Exception as e:
                logger.error(f"Failed to auto-cleanup XLSX result file: {e}")
        
        # Remove job from memory
        del job_status[job_id]
        logger.info(f"Auto-cleaned up job {job_id} from memory")

@app.get("/")
async def root():
    """Root endpoint for Railway health check"""
    try:
        return {
            "status": "healthy", 
            "message": "Medical Data Processor API is running", 
            "port": PORT,
            "environment": os.environ.get('RAILWAY_ENVIRONMENT', 'development')
        }
    except Exception as e:
        logger.error(f"Root endpoint error: {str(e)}")
        return {"status": "error", "message": str(e)}

@app.get("/debug/folders")
async def debug_folders():
    """Debug endpoint to check folder contents"""
    try:
        current_dir = Path(__file__).parent / "current"
        input_dir = current_dir / "input"
        output_dir = current_dir / "output"
        
        return {
            "current_dir": str(current_dir),
            "current_dir_exists": current_dir.exists(),
            "input_dir": str(input_dir),
            "input_dir_exists": input_dir.exists(),
            "input_files": [f.name for f in input_dir.glob("*")] if input_dir.exists() else [],
            "output_dir": str(output_dir),
            "output_dir_exists": output_dir.exists(),
            "output_files": [f.name for f in output_dir.glob("*")] if output_dir.exists() else []
        }
    except Exception as e:
        return {"error": str(e)}

@app.get("/health")
async def health_check():
    """Health check endpoint"""
    try:
        # Check if input and output folders exist
        current_dir = Path(__file__).parent / "current"
        input_dir = current_dir / "input"
        output_dir = current_dir / "output"
        
        input_files = list(input_dir.glob("*.pdf")) if input_dir.exists() else []
        output_files = list(output_dir.glob("*.pdf")) if output_dir.exists() else []
        
        return {
            "status": "healthy", 
            "timestamp": "2025-08-15",
            "port": PORT,
            "environment": os.environ.get('RAILWAY_ENVIRONMENT', 'development'),
            "input_folder": str(input_dir),
            "input_files": [f.name for f in input_files],
            "output_folder": str(output_dir),
            "output_files": [f.name for f in output_files]
        }
    except Exception as e:
        logger.error(f"Health check error: {str(e)}")
        return {"status": "error", "message": str(e)}

def convert_uni_background(job_id: str, csv_path: str):
    """Background task to convert UNI CSV using the conversion script"""
    job = job_status[job_id]
    
    try:
        job.status = "processing"
        job.message = "Starting UNI CSV conversion..."
        job.progress = 10
        
        # Import the conversion functions
        import sys
        sys.path.append(str(Path(__file__).parent / "uni-conversion"))
        from convert import convert_data
        
        job.message = "Converting UNI CSV data..."
        job.progress = 50
        
        # Create output file path
        result_base = Path(f"/tmp/results/{job_id}_converted")
        result_base.parent.mkdir(exist_ok=True)
        
        output_file_csv = result_base.with_suffix('.csv')
        
        # Run the conversion
        success = convert_data(csv_path, str(output_file_csv))
        
        if not success:
            raise Exception("UNI conversion failed")
        
        job.message = "UNI conversion complete, preparing results..."
        job.progress = 90
        
        # Verify output file exists
        if not os.path.exists(output_file_csv):
            raise Exception("Output file was not created")
        
        # Create XLSX version
        try:
            output_file_xlsx = convert_csv_to_xlsx(output_file_csv, result_base.with_suffix('.xlsx'))
            job.result_file_xlsx = output_file_xlsx
        except Exception as e:
            logger.warning(f"Could not create XLSX version: {e}")
        
        job.result_file = str(output_file_csv)
        job.status = "completed"
        job.progress = 100
        job.message = "UNI conversion completed successfully!"
        
        # Clean up input file
        os.unlink(csv_path)
        
        # Force garbage collection to free memory
        gc.collect()
        logger.info(f"Job {job_id}: Memory cleaned up after UNI conversion")
        
    except Exception as e:
        job.status = "failed"
        job.error = str(e)
        job.message = f"UNI conversion failed: {str(e)}"
        logger.error(f"UNI conversion job {job_id} failed: {str(e)}")
        
        # Clean up memory even on failure
        gc.collect()

def convert_instructions_background(job_id: str, excel_path: str):
    """Background task to convert Excel file using the instructions conversion script"""
    job = job_status[job_id]
    
    try:
        job.status = "processing"
        job.message = "Starting instructions conversion..."
        job.progress = 10
        
        # Import required modules
        import pandas as pd
        import sys
        sys.path.append(str(Path(__file__).parent / "instructions-conversion"))
        from convert_instructions import convert_data
        
        job.message = "Reading Excel file..."
        job.progress = 30
        
        # Convert Excel to CSV first
        temp_csv_path = f"/tmp/{job_id}_temp_input.csv"
        try:
            # Read Excel file
            df = pd.read_excel(excel_path)
            # Save as CSV
            df.to_csv(temp_csv_path, index=False)
        except Exception as e:
            raise Exception(f"Failed to read Excel file: {str(e)}")
        
        job.message = "Preparing additions file..."
        job.progress = 45
        
        # Copy the default additions.csv file
        import shutil
        additions_dest = Path(temp_csv_path).parent / "additions.csv"
        additions_source = Path(__file__).parent / "instructions-conversion" / "additions.csv"
        
        if additions_source.exists():
            shutil.copy2(additions_source, additions_dest)
            logger.info(f"Copied additions.csv from {additions_source} to {additions_dest}")
        else:
            logger.warning(f"additions.csv not found at {additions_source}")
        
        job.message = "Converting data using instructions script..."
        job.progress = 60
        
        # Create output file path
        result_base = Path(f"/tmp/results/{job_id}_converted")
        result_base.parent.mkdir(exist_ok=True)
        
        output_file_csv = result_base.with_suffix('.csv')
        
        # Run the conversion using the convert_instructions script
        success = convert_data(temp_csv_path, str(output_file_csv))
        
        if not success:
            raise Exception("Instructions conversion failed")
        
        job.message = "Conversion complete, preparing Excel output..."
        job.progress = 85
        
        # Convert to Excel format
        excel_output_file = result_base.with_suffix('.xlsx')
        try:
            converted_df = pd.read_csv(output_file_csv, dtype=str)
            converted_df.to_excel(excel_output_file, index=False, engine='openpyxl')
            # Store both file paths
            job.result_file = str(output_file_csv)
            job.result_file_xlsx = str(excel_output_file)
            logger.info(f"Created both CSV and XLSX outputs")
        except Exception as e:
            # If Excel conversion fails, use CSV only
            logger.warning(f"Excel conversion failed, using CSV only: {str(e)}")
            job.result_file = str(output_file_csv)
        
        job.status = "completed"
        job.progress = 100
        job.message = "Instructions conversion completed successfully!"
        
        # Clean up temporary files
        if os.path.exists(temp_csv_path):
            os.unlink(temp_csv_path)
        if os.path.exists(additions_dest):
            os.unlink(additions_dest)
        os.unlink(excel_path)
        
        # Force garbage collection to free memory
        gc.collect()
        logger.info(f"Job {job_id}: Memory cleaned up after instructions conversion")
        
    except Exception as e:
        job.status = "failed"
        job.error = str(e)
        job.message = f"Instructions conversion failed: {str(e)}"
        logger.error(f"Instructions conversion job {job_id} failed: {str(e)}")
        
        # Clean up on error
        try:
            if 'temp_csv_path' in locals() and os.path.exists(temp_csv_path):
                os.unlink(temp_csv_path)
            if 'additions_dest' in locals() and os.path.exists(additions_dest):
                os.unlink(additions_dest)
        except Exception as cleanup_error:
            logger.error(f"Cleanup error: {cleanup_error}")
        
        # Clean up memory even on failure
        gc.collect()

def provider_mapping_background(job_id: str, excel_path: str):
    """Background task to process provider mapping Excel file"""
    job = job_status[job_id]
    
    try:
        job.status = "processing"
        job.message = "Reading Excel file..."
        job.progress = 10
        
        import pandas as pd
        import openpyxl
        
        # Try to read Excel file - handle both .xlsx and .xls formats
        excel_path_obj = Path(excel_path)
        
        # If it's an old .xls file, convert to .xlsx first using pandas
        if excel_path_obj.suffix.lower() == '.xls':
            try:
                # Read .xls file with pandas
                df_temp = pd.read_excel(excel_path, engine='xlrd')
                # Save as .xlsx
                temp_xlsx_path = excel_path_obj.with_suffix('.xlsx')
                df_temp.to_excel(temp_xlsx_path, index=False, engine='openpyxl')
                excel_path = str(temp_xlsx_path)
                logger.info(f"Converted .xls to .xlsx: {excel_path}")
            except Exception as e:
                logger.error(f"Failed to convert .xls file: {e}")
                raise Exception(f"Could not read .xls file. Please save as .xlsx format: {str(e)}")
        
        # Read Excel file with openpyxl to access raw cell values
        try:
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
        except Exception as e:
            logger.error(f"Failed to read Excel file with openpyxl: {e}")
            raise Exception(f"Could not read Excel file. Please ensure it's a valid .xlsx file: {str(e)}")
        
        sheet = workbook.active
        
        job.message = "Searching for column headers..."
        job.progress = 30
        
        # Find the header row and column positions
        # Look for "CRNA", "Name", and "MedNet Code" headers
        header_row = None
        crna_col = None
        name_col = None
        mednet_code_col = None
        
        # Search for header row (typically contains "CRNA", "Name", "MedNet Code")
        for row_idx, row in enumerate(sheet.iter_rows(), start=1):
            for col_idx, cell in enumerate(row, start=1):
                cell_value = cell.value
                if cell_value:
                    cell_str = str(cell_value).strip().upper()
                    if cell_str == "CRNA":
                        crna_col = col_idx
                        if header_row is None:
                            header_row = row_idx
                    elif cell_str == "NAME":
                        name_col = col_idx
                        if header_row is None:
                            header_row = row_idx
                    elif cell_str in ["MEDNET CODE", "MEDNETCODE", "MED NET CODE"]:
                        mednet_code_col = col_idx
                        if header_row is None:
                            header_row = row_idx
            
            # If we found all columns, we can stop
            if header_row and crna_col and name_col and mednet_code_col:
                break
        
        if not header_row:
            raise Exception("Could not find header row in the Excel file")
        
        if not name_col:
            raise Exception("Could not find 'Name' column in the Excel file")
        
        if not mednet_code_col:
            raise Exception("Could not find 'MedNet Code' column in the Excel file")
        
        logger.info(f"Found header row {header_row}: CRNA col={crna_col}, Name col={name_col}, MedNet Code col={mednet_code_col}")
        
        job.message = "Extracting CRNA providers..."
        job.progress = 50
        
        # Extract CRNA providers with MedNet codes
        # Start from row after header
        crna_providers = []
        data_start_row = header_row + 1
        
        # Find where CRNA section ends (either MD header or empty rows)
        while data_start_row <= sheet.max_row:
            # Check if we've reached the MD section (check all columns for "MD")
            found_md = False
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=data_start_row, column=col_idx)
                if cell.value and str(cell.value).strip().upper() == "MD":
                    found_md = True
                    break
            
            if found_md:
                break
            
            # Get provider name from Name column
            name_cell = sheet.cell(row=data_start_row, column=name_col)
            provider_name = str(name_cell.value).strip() if name_cell.value else ""
            
            # Get MedNet code from MedNet Code column
            mednet_code_cell = sheet.cell(row=data_start_row, column=mednet_code_col)
            mednet_code = str(mednet_code_cell.value).strip() if mednet_code_cell.value else ""
            
            # Stop if name is empty (end of data)
            if not provider_name:
                if crna_providers:
                    break
                else:
                    data_start_row += 1
                    continue
            
            # Store provider with MedNet code
            crna_providers.append({
                'name': provider_name,
                'mednet_code': mednet_code
            })
            
            data_start_row += 1
        
        logger.info(f"Found {len(crna_providers)} CRNA providers")
        
        job.message = "Searching for MD section..."
        job.progress = 70
        
        # Find MD section (if exists)
        md_providers = []
        md_row = None
        
        # Search for MD header starting from where CRNA section ended
        for row_idx in range(data_start_row, sheet.max_row + 1):
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_idx, column=col_idx)
                if cell.value and str(cell.value).strip().upper() == "MD":
                    md_row = row_idx
                    break
            if md_row:
                break
        
        if md_row:
            logger.info(f"Found MD at row {md_row}")
            # Extract MD providers with MedNet codes
            current_row = md_row + 1
            
            while current_row <= sheet.max_row:
                # Get provider name from Name column
                name_cell = sheet.cell(row=current_row, column=name_col)
                provider_name = str(name_cell.value).strip() if name_cell.value else ""
                
                # Get MedNet code from MedNet Code column
                mednet_code_cell = sheet.cell(row=current_row, column=mednet_code_col)
                mednet_code = str(mednet_code_cell.value).strip() if mednet_code_cell.value else ""
                
                if not provider_name:
                    break
                
                md_providers.append({
                    'name': provider_name,
                    'mednet_code': mednet_code
                })
                
                current_row += 1
            
            logger.info(f"Found {len(md_providers)} MD providers")
        
        job.message = "Formatting output..."
        job.progress = 90
        
        # Format providers: {last name}, {first name} {middle name}, {title} (MedNet Code: {code})
        def format_provider(provider_dict, title):
            """Parse provider name and format it with MedNet code"""
            name_str = provider_dict['name']
            mednet_code = provider_dict.get('mednet_code', '')
            
            # Split by comma
            parts = [p.strip() for p in name_str.split(',')]
            
            if len(parts) >= 2:
                last_name = parts[0]
                first_name = parts[1]
                middle_name = parts[2] if len(parts) > 2 else ""
                
                # Format: {last name}, {first name} {middle name}, {title}
                if middle_name:
                    formatted_name = f"{last_name}, {first_name} {middle_name}, {title}"
                else:
                    formatted_name = f"{last_name}, {first_name}, {title}"
            else:
                # Fallback: if format is unexpected, just add title
                formatted_name = f"{name_str}, {title}"
            
            # Add MedNet code if available
            if mednet_code:
                return f"{formatted_name} (MedNet Code: {mednet_code})"
            else:
                return formatted_name
        
        # Format CRNA providers
        formatted_crna = [format_provider(p, "CRNA") for p in crna_providers]
        
        # Format MD providers
        formatted_md = [format_provider(p, "MD") for p in md_providers]
        
        # Build output string
        output_lines = []
        output_lines.append("Billable CRNA providers:")
        output_lines.extend(formatted_crna)
        output_lines.append("")  # Empty line between sections
        output_lines.append("Billable MD providers:")
        output_lines.extend(formatted_md)
        
        output_text = "\n".join(output_lines)
        
        job.status = "completed"
        job.progress = 100
        job.message = f"Provider mapping completed! Found {len(crna_providers)} CRNA and {len(md_providers)} MD providers."
        job.result = {"output": output_text}
        
        # Clean up Excel file
        if os.path.exists(excel_path):
            os.unlink(excel_path)
        
        # Force garbage collection to free memory
        gc.collect()
        logger.info(f"Job {job_id}: Memory cleaned up after provider mapping")
        
    except Exception as e:
        job.status = "failed"
        job.error = str(e)
        job.message = f"Provider mapping failed: {str(e)}"
        logger.error(f"Provider mapping job {job_id} failed: {str(e)}")
        
        # Clean up Excel file
        if os.path.exists(excel_path):
            os.unlink(excel_path)
        
        # Clean up memory even on failure
        gc.collect()

def surgeon_mapping_background(job_id: str, excel_path: str):
    """Background task to process surgeon mapping Excel file"""
    job = job_status[job_id]
    
    try:
        job.status = "processing"
        job.message = "Reading Excel file..."
        job.progress = 10
        
        import pandas as pd
        import openpyxl
        
        # Try to read Excel file - handle both .xlsx and .xls formats
        excel_path_obj = Path(excel_path)
        
        # If it's an old .xls file, convert to .xlsx first using pandas
        if excel_path_obj.suffix.lower() == '.xls':
            try:
                # Read .xls file with pandas
                df_temp = pd.read_excel(excel_path, engine='xlrd')
                # Save as .xlsx
                temp_xlsx_path = excel_path_obj.with_suffix('.xlsx')
                df_temp.to_excel(temp_xlsx_path, index=False, engine='openpyxl')
                excel_path = str(temp_xlsx_path)
                logger.info(f"Converted .xls to .xlsx: {excel_path}")
            except Exception as e:
                logger.error(f"Failed to convert .xls file: {e}")
                raise Exception(f"Could not read .xls file. Please save as .xlsx format: {str(e)}")
        
        # Read Excel file with openpyxl to access raw cell values
        try:
            workbook = openpyxl.load_workbook(excel_path, data_only=True)
        except Exception as e:
            logger.error(f"Failed to read Excel file with openpyxl: {e}")
            raise Exception(f"Could not read Excel file. Please ensure it's a valid .xlsx file: {str(e)}")
        
        sheet = workbook.active
        
        job.message = "Searching for Last Name column..."
        job.progress = 30
        
        # Find the "Last Name" header cell
        last_name_row = None
        last_name_col = None
        
        for row_idx, row in enumerate(sheet.iter_rows(), start=1):
            for col_idx, cell in enumerate(row, start=1):
                cell_value = cell.value
                if cell_value and str(cell_value).strip() == "Last Name":
                    last_name_row = row_idx
                    last_name_col = col_idx
                    break
            if last_name_row:
                break
        
        if not last_name_row:
            raise Exception("Could not find 'Last Name' column header in the Excel file")
        
        logger.info(f"Found Last Name at row {last_name_row}, col {last_name_col}")
        
        # Find Middle Name and First Name columns (should be adjacent)
        middle_name_col = None
        first_name_col = None
        
        # Check adjacent columns for Middle Name and First Name
        for col_offset in [-2, -1, 1, 2]:
            check_col = last_name_col + col_offset
            if check_col < 1:
                continue
            try:
                header_cell = sheet.cell(row=last_name_row, column=check_col)
                header_value = str(header_cell.value).strip() if header_cell.value else ""
                if header_value == "Middle Name":
                    middle_name_col = check_col
                elif header_value == "First Name":
                    first_name_col = check_col
            except:
                continue
        
        # If not found adjacent, search the entire header row
        if not middle_name_col or not first_name_col:
            for col_idx in range(1, sheet.max_column + 1):
                header_cell = sheet.cell(row=last_name_row, column=col_idx)
                header_value = str(header_cell.value).strip() if header_cell.value else ""
                if header_value == "Middle Name" and not middle_name_col:
                    middle_name_col = col_idx
                elif header_value == "First Name" and not first_name_col:
                    first_name_col = col_idx
        
        if not first_name_col:
            raise Exception("Could not find 'First Name' column header in the Excel file")
        
        logger.info(f"Found First Name at col {first_name_col}")
        if middle_name_col:
            logger.info(f"Found Middle Name at col {middle_name_col}")
        
        job.message = "Extracting surgeon data..."
        job.progress = 50
        
        # Extract surgeons (data starts from row after header)
        surgeons = []
        data_start_row = last_name_row + 1
        
        while data_start_row <= sheet.max_row:
            last_name_cell = sheet.cell(row=data_start_row, column=last_name_col)
            first_name_cell = sheet.cell(row=data_start_row, column=first_name_col)
            
            last_name = str(last_name_cell.value).strip() if last_name_cell.value else ""
            first_name = str(first_name_cell.value).strip() if first_name_cell.value else ""
            
            # Stop if both are empty (end of data)
            if not last_name and not first_name:
                break
            
            # Get middle name if column exists
            middle_name = ""
            if middle_name_col:
                middle_name_cell = sheet.cell(row=data_start_row, column=middle_name_col)
                middle_name = str(middle_name_cell.value).strip() if middle_name_cell.value else ""
            
            # Only add if we have at least last name and first name
            if last_name and first_name:
                surgeons.append({
                    "last_name": last_name,
                    "first_name": first_name,
                    "middle_name": middle_name
                })
            
            data_start_row += 1
        
        logger.info(f"Found {len(surgeons)} surgeons")
        
        job.message = "Formatting output..."
        job.progress = 90
        
        # Format surgeons: {last name}, {first name} {middle name}, MD
        formatted_surgeons = []
        for surgeon in surgeons:
            last_name = surgeon["last_name"]
            first_name = surgeon["first_name"]
            middle_name = surgeon["middle_name"]
            
            if middle_name:
                formatted_name = f"{last_name}, {first_name} {middle_name}, MD"
            else:
                formatted_name = f"{last_name}, {first_name}, MD"
            
            formatted_surgeons.append(formatted_name)
        
        # Build output string
        output_lines = []
        output_lines.append("Billable MD providers:")
        output_lines.extend(formatted_surgeons)
        
        output_text = "\n".join(output_lines)
        
        job.status = "completed"
        job.progress = 100
        job.message = f"Surgeon mapping completed! Found {len(surgeons)} surgeons."
        job.result = {"output": output_text}
        
        # Clean up Excel file
        if os.path.exists(excel_path):
            os.unlink(excel_path)
        
        # Force garbage collection to free memory
        gc.collect()
        logger.info(f"Job {job_id}: Memory cleaned up after surgeon mapping")
        
    except Exception as e:
        job.status = "failed"
        job.error = str(e)
        job.message = f"Surgeon mapping failed: {str(e)}"
        logger.error(f"Surgeon mapping job {job_id} failed: {str(e)}")
        
        # Clean up Excel file
        if os.path.exists(excel_path):
            os.unlink(excel_path)
        
        # Clean up memory even on failure
        gc.collect()

def merge_by_csn_background(job_id: str, csv_path_1: str, csv_path_2: str):
    """Background task to merge two CSV files by CSN column"""
    job = job_status[job_id]
    
    try:
        job.status = "processing"
        job.message = "Reading CSV files..."
        job.progress = 10
        
        # Read both CSV files
        import pandas as pd
        
        # Try multiple encodings for CSV files
        encodings_to_try = ['utf-8', 'utf-8-sig', 'latin-1', 'iso-8859-1', 'cp1252']
        df1 = None
        df2 = None
        
        for encoding in encodings_to_try:
            try:
                df1 = pd.read_csv(csv_path_1, dtype=str, encoding=encoding)
                logger.info(f"Successfully read first file with encoding: {encoding}")
                break
            except UnicodeDecodeError:
                continue
        
        if df1 is None:
            raise Exception("Could not read first CSV file with any standard encoding")
        
        for encoding in encodings_to_try:
            try:
                df2 = pd.read_csv(csv_path_2, dtype=str, encoding=encoding)
                logger.info(f"Successfully read second file with encoding: {encoding}")
                break
            except UnicodeDecodeError:
                continue
        
        if df2 is None:
            raise Exception("Could not read second CSV file with any standard encoding")
        
        job.message = "Checking for CSN column..."
        job.progress = 30
        
        # Find CSN column (case-insensitive)
        csn_col_1 = None
        csn_col_2 = None
        
        for col in df1.columns:
            if col.upper() == 'CSN':
                csn_col_1 = col
                break
        
        for col in df2.columns:
            if col.upper() == 'CSN':
                csn_col_2 = col
                break
        
        if csn_col_1 is None:
            raise Exception("First CSV file does not have a CSN column")
        if csn_col_2 is None:
            raise Exception("Second CSV file does not have a CSN column")
        
        job.message = "Merging files by CSN..."
        job.progress = 50
        
        # Merge the dataframes on CSN column (inner join)
        merged_df = pd.merge(
            df1, 
            df2, 
            left_on=csn_col_1, 
            right_on=csn_col_2, 
            how='inner',
            suffixes=('', '_y')
        )
        
        # Remove duplicate CSN column if both had the same name
        if csn_col_1 == csn_col_2:
            # Keep only one CSN column
            if f'{csn_col_1}_y' in merged_df.columns:
                merged_df = merged_df.drop(columns=[f'{csn_col_1}_y'])
        
        # Remove any other duplicate columns (ending with _y)
        cols_to_drop = [col for col in merged_df.columns if col.endswith('_y')]
        if cols_to_drop:
            merged_df = merged_df.drop(columns=cols_to_drop)
        
        job.message = "Preparing merged file..."
        job.progress = 80
        
        # Create output file path
        result_base = Path(f"/tmp/results/{job_id}_merged")
        result_base.parent.mkdir(exist_ok=True)
        
        output_file_csv = result_base.with_suffix('.csv')
        
        # Save merged CSV
        merged_df.to_csv(output_file_csv, index=False, encoding='utf-8')
        
        # Create XLSX version
        try:
            output_file_xlsx = convert_csv_to_xlsx(output_file_csv, result_base.with_suffix('.xlsx'))
            job.result_file_xlsx = output_file_xlsx
        except Exception as e:
            logger.warning(f"Could not create XLSX version: {e}")
        
        job.result_file = str(output_file_csv)
        job.status = "completed"
        job.progress = 100
        job.message = f"Merge completed! {len(merged_df)} rows matched by CSN."
        
        # Clean up input files
        os.unlink(csv_path_1)
        os.unlink(csv_path_2)
        
        logger.info(f"Merge by CSN completed for job {job_id}: {len(merged_df)} rows")
        
    except Exception as e:
        logger.error(f"Merge by CSN background error: {str(e)}")
        job.status = "failed"
        job.error = str(e)
        job.message = f"Merge failed: {str(e)}"

def generate_modifiers_background(job_id: str, csv_path: str, turn_off_medical_direction: bool = False, generate_qk_duplicate: bool = False, limit_anesthesia_time: bool = False, turn_off_bcbs_medicare_modifiers: bool = True, peripheral_blocks_mode: str = "other", add_pt_for_non_medicare: bool = False, change_responsible_provider_to_md_if_p_only: bool = False):
    """Background task to generate medical modifiers using the modifiers script
    
    Args:
        job_id: Unique job identifier
        csv_path: Path to input CSV file
        turn_off_medical_direction: If True, override all medical direction YES to NO
        generate_qk_duplicate: If True, generate duplicate line when QK modifier is applied with CRNA as Responsible Provider
        limit_anesthesia_time: If True, limit anesthesia time to maximum 480 minutes based on An Start and An Stop columns
        turn_off_bcbs_medicare_modifiers: If True, for MedNet Code 003 (BCBS), only generate P modifiers (no M1/GC/QS)
        peripheral_blocks_mode: Mode for peripheral block generation - "UNI" (only General) or "other" (not MAC)
        add_pt_for_non_medicare: If True, add PT modifier for non-Medicare insurances when polyps found and screening colonoscopy
        change_responsible_provider_to_md_if_p_only: If True, change Responsible Provider to MD when only P modifier is used and both MD and CRNA are present
    """
    job = job_status[job_id]
    
    try:
        job.status = "processing"
        qk_msg = ' + QK Duplicates' if generate_qk_duplicate else ''
        time_limit_msg = ' + Time Limited' if limit_anesthesia_time else ''
        bcbs_msg = ' + BCBS Modifiers OFF' if turn_off_bcbs_medicare_modifiers else ''
        blocks_mode_msg = f' (Blocks: {peripheral_blocks_mode})' if peripheral_blocks_mode else ''
        pt_non_medicare_msg = ' + PT for Non-Medicare' if add_pt_for_non_medicare else ''
        job.message = f"Starting modifiers generation{' (Medical Direction OFF)' if turn_off_medical_direction else ''}{qk_msg}{time_limit_msg}{bcbs_msg}{blocks_mode_msg}{pt_non_medicare_msg}..."
        job.progress = 10
        
        # Import the modifiers generation function
        import sys
        sys.path.append(str(Path(__file__).parent / "modifiers"))
        from generate_modifiers import generate_modifiers
        
        job.message = "Generating medical modifiers..."
        job.progress = 50
        
        # Create output file path
        result_base = Path(f"/tmp/results/{job_id}_with_modifiers")
        result_base.parent.mkdir(exist_ok=True)
        
        output_file_csv = result_base.with_suffix('.csv')
        
        # Run the modifiers generation with medical direction override parameter, QK duplicate parameter, time limiting parameter, BCBS modifiers parameter, peripheral blocks mode, PT for non-Medicare parameter, and change Responsible Provider to MD if P only parameter
        success = generate_modifiers(csv_path, str(output_file_csv), turn_off_medical_direction, generate_qk_duplicate, limit_anesthesia_time, turn_off_bcbs_medicare_modifiers, peripheral_blocks_mode, add_pt_for_non_medicare, change_responsible_provider_to_md_if_p_only)
        
        if not success:
            raise Exception("Modifiers generation failed")
        
        job.message = "Modifiers generation complete, preparing results..."
        job.progress = 90
        
        # Verify output file exists
        if not os.path.exists(output_file_csv):
            raise Exception("Output file was not created")
        
        # Create XLSX version
        try:
            output_file_xlsx = convert_csv_to_xlsx(output_file_csv, result_base.with_suffix('.xlsx'))
            job.result_file_xlsx = output_file_xlsx
        except Exception as e:
            logger.warning(f"Could not create XLSX version: {e}")
        
        job.result_file = str(output_file_csv)
        job.status = "completed"
        job.progress = 100
        job.message = "Modifiers generation completed successfully!"
        
        # Clean up input file
        os.unlink(csv_path)
        
        # Force garbage collection to free memory
        gc.collect()
        logger.info(f"Job {job_id}: Memory cleaned up after modifiers generation")
        
    except Exception as e:
        job.status = "failed"
        job.error = str(e)
        job.message = f"Modifiers generation failed: {str(e)}"
        logger.error(f"Modifiers generation job {job_id} failed: {str(e)}")
        
        # Clean up memory even on failure
        gc.collect()

def predict_insurance_codes_background(job_id: str, data_csv_path: str, special_cases_csv_path: str = None, enable_ai: bool = True):
    """Background task to predict MedNet codes for insurance companies"""
    job = job_status[job_id]
    
    try:
        job.status = "processing"
        job.message = "Starting insurance code prediction..."
        job.progress = 10
        
        # Import the prediction function
        import sys
        sys.path.append(str(Path(__file__).parent / "sorting"))
        from predict_mednet import process_insurance_predictions
        
        job.message = "Loading MedNet database..."
        job.progress = 20
        
        # Path to the mednet.csv file
        mednet_csv_path = Path(__file__).parent / "sorting" / "mednet.csv"
        
        if not mednet_csv_path.exists():
            raise Exception("MedNet database not found")
        
        job.message = "Predicting insurance codes..." + (" (AI enabled)" if enable_ai else " (special cases only)")
        job.progress = 30
        
        # Create output file path
        result_base = Path(f"/tmp/results/{job_id}_with_insurance_codes")
        result_base.parent.mkdir(exist_ok=True)
        
        output_file_csv = result_base.with_suffix('.csv')
        
        # Run the prediction
        success = process_insurance_predictions(
            data_csv_path, 
            str(mednet_csv_path), 
            str(output_file_csv), 
            special_cases_csv_path,
            max_workers=10,
            enable_ai=enable_ai
        )
        
        if not success:
            raise Exception("Insurance code prediction failed")
        
        job.message = "Prediction complete, preparing results..."
        job.progress = 90
        
        # Verify output file exists
        if not os.path.exists(output_file_csv):
            raise Exception("Output file was not created")
        
        # Create XLSX version
        try:
            output_file_xlsx = convert_csv_to_xlsx(output_file_csv, result_base.with_suffix('.xlsx'))
            job.result_file_xlsx = output_file_xlsx
        except Exception as e:
            logger.warning(f"Could not create XLSX version: {e}")
        
        job.result_file = str(output_file_csv)
        job.status = "completed"
        job.progress = 100
        job.message = "Insurance code prediction completed successfully!"
        
        # Clean up input files
        os.unlink(data_csv_path)
        if special_cases_csv_path and os.path.exists(special_cases_csv_path):
            os.unlink(special_cases_csv_path)
        
        # Force garbage collection to free memory
        gc.collect()
        logger.info(f"Job {job_id}: Memory cleaned up after insurance code prediction")
        
    except Exception as e:
        job.status = "failed"
        job.error = str(e)
        job.message = f"Insurance code prediction failed: {str(e)}"
        logger.error(f"Insurance code prediction job {job_id} failed: {str(e)}")
        
        # Clean up memory even on failure
        gc.collect()

@app.post("/convert-uni")
async def convert_uni(
    background_tasks: BackgroundTasks,
    csv_file: UploadFile = File(...)
):
    """Upload a UNI CSV or XLSX file to convert using the conversion script"""
    
    try:
        logger.info(f"Received UNI conversion request - file: {csv_file.filename}")
        
        # Validate file type
        if not csv_file.filename.endswith(('.csv', '.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="File must be a CSV or XLSX")
        
        # Generate job ID
        job_id = str(uuid.uuid4())
        job = ProcessingJob(job_id)
        job_status[job_id] = job
        
        logger.info(f"Created UNI conversion job {job_id}")
        
        # Save uploaded file
        input_path = f"/tmp/{job_id}_uni_input{Path(csv_file.filename).suffix}"
        
        with open(input_path, "wb") as f:
            shutil.copyfileobj(csv_file.file, f)
        
        logger.info(f"File saved - path: {input_path}")
        
        # Convert to CSV if needed
        csv_path = ensure_csv_file(input_path, f"/tmp/{job_id}_uni_input.csv")
        
        # Clean up original file if it was converted
        if csv_path != input_path and os.path.exists(input_path):
            os.unlink(input_path)
        
        logger.info(f"CSV ready - path: {csv_path}")
        
        # Start background processing
        background_tasks.add_task(convert_uni_background, job_id, csv_path)
        
        logger.info(f"Background UNI conversion task started for job {job_id}")
        
        return {"job_id": job_id, "message": "UNI file uploaded and conversion started"}
        
    except Exception as e:
        logger.error(f"UNI conversion upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

@app.post("/convert-instructions")
async def convert_instructions(
    background_tasks: BackgroundTasks,
    excel_file: UploadFile = File(...)
):
    """
    Upload an Excel file to convert using the instructions conversion script.
    Uses the default additions.csv file for field instructions.
    """
    
    try:
        logger.info(f"Received instructions conversion request - excel: {excel_file.filename}")
        
        # Validate file type
        if not excel_file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
        
        # Generate job ID
        job_id = str(uuid.uuid4())
        job = ProcessingJob(job_id)
        job_status[job_id] = job
        
        logger.info(f"Created instructions conversion job {job_id}")
        
        # Save uploaded Excel file
        excel_path = f"/tmp/{job_id}_instructions_input.xlsx"
        
        with open(excel_path, "wb") as f:
            shutil.copyfileobj(excel_file.file, f)
        
        logger.info(f"Instructions Excel saved - path: {excel_path}")
        
        # Start background processing
        background_tasks.add_task(convert_instructions_background, job_id, excel_path)
        
        logger.info(f"Background instructions conversion task started for job {job_id}")
        
        return {"job_id": job_id, "message": "Excel file uploaded and instructions conversion started"}
        
    except Exception as e:
        logger.error(f"Instructions conversion upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

@app.post("/provider-mapping")
async def provider_mapping(
    background_tasks: BackgroundTasks,
    excel_file: UploadFile = File(...)
):
    """
    Upload an Excel file containing provider list to extract CRNA and MD providers.
    The file should contain a "CRNA" cell followed by CRNA provider names,
    and optionally an "MD" cell followed by MD provider names.
    """
    
    try:
        logger.info(f"Received provider mapping request - excel: {excel_file.filename}")
        
        # Validate file type
        if not excel_file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
        
        # Generate job ID
        job_id = str(uuid.uuid4())
        job = ProcessingJob(job_id)
        job_status[job_id] = job
        
        logger.info(f"Created provider mapping job {job_id}")
        
        # Save uploaded Excel file (preserve original extension)
        file_ext = Path(excel_file.filename).suffix.lower()
        excel_path = f"/tmp/{job_id}_provider_mapping_input{file_ext}"
        
        with open(excel_path, "wb") as f:
            shutil.copyfileobj(excel_file.file, f)
        
        logger.info(f"Provider mapping Excel saved - path: {excel_path}")
        
        # Start background processing
        background_tasks.add_task(provider_mapping_background, job_id, excel_path)
        
        logger.info(f"Background provider mapping task started for job {job_id}")
        
        return {"job_id": job_id, "message": "Excel file uploaded and provider mapping started"}
        
    except Exception as e:
        logger.error(f"Provider mapping upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

@app.post("/surgeon-mapping")
async def surgeon_mapping(
    background_tasks: BackgroundTasks,
    excel_file: UploadFile = File(...)
):
    """
    Upload an Excel file containing surgeon list to extract surgeons.
    The file should contain columns: Last Name, First Name, and optionally Middle Name.
    All surgeons will be formatted with title "MD".
    """
    
    try:
        logger.info(f"Received surgeon mapping request - excel: {excel_file.filename}")
        
        # Validate file type
        if not excel_file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")
        
        # Generate job ID
        job_id = str(uuid.uuid4())
        job = ProcessingJob(job_id)
        job_status[job_id] = job
        
        logger.info(f"Created surgeon mapping job {job_id}")
        
        # Save uploaded Excel file (preserve original extension)
        file_ext = Path(excel_file.filename).suffix.lower()
        excel_path = f"/tmp/{job_id}_surgeon_mapping_input{file_ext}"
        
        with open(excel_path, "wb") as f:
            shutil.copyfileobj(excel_file.file, f)
        
        logger.info(f"Surgeon mapping Excel saved - path: {excel_path}")
        
        # Start background processing
        background_tasks.add_task(surgeon_mapping_background, job_id, excel_path)
        
        logger.info(f"Background surgeon mapping task started for job {job_id}")
        
        return {"job_id": job_id, "message": "Excel file uploaded and surgeon mapping started"}
        
    except Exception as e:
        logger.error(f"Surgeon mapping upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

@app.post("/merge-by-csn")
async def merge_by_csn(
    background_tasks: BackgroundTasks,
    csv_file_1: UploadFile = File(...),
    csv_file_2: UploadFile = File(...)
):
    """Upload two CSV or XLSX files to merge them by matching rows based on the CSN column"""
    
    try:
        logger.info(f"Received merge by CSN request - file1: {csv_file_1.filename}, file2: {csv_file_2.filename}")
        
        # Validate file types
        if not csv_file_1.filename.endswith(('.csv', '.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="First file must be a CSV or XLSX")
        if not csv_file_2.filename.endswith(('.csv', '.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Second file must be a CSV or XLSX")
        
        # Generate job ID
        job_id = str(uuid.uuid4())
        job = ProcessingJob(job_id)
        job_status[job_id] = job
        
        logger.info(f"Created merge by CSN job {job_id}")
        
        # Save uploaded files
        input_path_1 = f"/tmp/{job_id}_merge_input1{Path(csv_file_1.filename).suffix}"
        input_path_2 = f"/tmp/{job_id}_merge_input2{Path(csv_file_2.filename).suffix}"
        
        with open(input_path_1, "wb") as f:
            shutil.copyfileobj(csv_file_1.file, f)
        with open(input_path_2, "wb") as f:
            shutil.copyfileobj(csv_file_2.file, f)
        
        logger.info(f"Files saved - path1: {input_path_1}, path2: {input_path_2}")
        
        # Convert to CSV if needed
        csv_path_1 = ensure_csv_file(input_path_1, f"/tmp/{job_id}_merge_input1.csv")
        csv_path_2 = ensure_csv_file(input_path_2, f"/tmp/{job_id}_merge_input2.csv")
        
        # Clean up original files if they were converted
        if csv_path_1 != input_path_1 and os.path.exists(input_path_1):
            os.unlink(input_path_1)
        if csv_path_2 != input_path_2 and os.path.exists(input_path_2):
            os.unlink(input_path_2)
        
        logger.info(f"CSV files ready - path1: {csv_path_1}, path2: {csv_path_2}")
        
        # Start background processing
        background_tasks.add_task(merge_by_csn_background, job_id, csv_path_1, csv_path_2)
        
        logger.info(f"Background merge by CSN task started for job {job_id}")
        
        return {"job_id": job_id, "message": "Files uploaded and merge by CSN started"}
        
    except Exception as e:
        logger.error(f"Merge by CSN upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

def check_cpt_codes_with_pdfs_background(job_id: str, predictions_path: str, ground_truth_path: str, pdfs_zip_path: str, charge_detail_path: str = None):
    """Background task to check CPT codes and extract error PDFs grouped by error type"""
    job = job_status[job_id]
    
    try:
        import zipfile
        import tempfile
        
        job.status = "processing"
        job.message = "Extracting PDFs from ZIP..."
        job.progress = 5
        
        # Create temporary directory for extracted PDFs
        pdfs_temp_dir = tempfile.mkdtemp()
        
        # Extract PDFs from ZIP
        with zipfile.ZipFile(pdfs_zip_path, 'r') as zip_ref:
            zip_ref.extractall(pdfs_temp_dir)
        
        logger.info(f"Extracted PDFs to {pdfs_temp_dir}")
        
        # Create a mapping from source_file (PDF filename) to full path
        pdf_mapping = {}
        for root, dirs, files in os.walk(pdfs_temp_dir):
            for file in files:
                if file.lower().endswith('.pdf'):
                    full_path = os.path.join(root, file)
                    # Store with just the filename as key (without path)
                    pdf_mapping[file] = full_path
        
        logger.info(f"Found {len(pdf_mapping)} PDFs in ZIP")
        
        job.message = "Running CPT/ICD comparison..."
        job.progress = 10
        
        # Run the normal CPT codes check (reuse existing logic)
        # We'll call the existing function but capture the results differently
        
        # Load predictions file
        try:
            predictions_df = pd.read_excel(predictions_path) if predictions_path.endswith(('.xlsx', '.xls')) else pd.read_csv(predictions_path, encoding='utf-8', dtype=str)
        except UnicodeDecodeError:
            predictions_df = pd.read_csv(predictions_path, encoding='latin-1', dtype=str)
        
        # Load ground truth file
        try:
            ground_truth_df = pd.read_excel(ground_truth_path) if ground_truth_path.endswith(('.xlsx', '.xls')) else pd.read_csv(ground_truth_path, encoding='utf-8', dtype=str)
        except UnicodeDecodeError:
            ground_truth_df = pd.read_csv(ground_truth_path, encoding='latin-1', dtype=str)
        
        # Load charge detail file if provided
        charge_detail_df = None
        if charge_detail_path:
            try:
                charge_detail_df = pd.read_excel(charge_detail_path) if charge_detail_path.endswith(('.xlsx', '.xls')) else pd.read_csv(charge_detail_path, encoding='utf-8', dtype=str)
            except UnicodeDecodeError:
                charge_detail_df = pd.read_csv(charge_detail_path, encoding='latin-1', dtype=str)
        
        job.message = "Analyzing errors and grouping PDFs..."
        job.progress = 30
        
        # Find source_file column in predictions
        source_file_col = None
        for col in predictions_df.columns:
            if col.lower().strip() in ['source_file', 'source file', 'sourcefile', 'pdf_file', 'pdf file', 'filename', 'file_name']:
                source_file_col = col
                break
        
        if not source_file_col:
            raise Exception("Could not find 'source_file' column in predictions file")
        
        # Find AccountId column in predictions
        account_id_col_pred = None
        for col in predictions_df.columns:
            col_upper = col.upper().strip()
            if col_upper in ['ACCOUNTID', 'ACCOUNT ID', 'ACCOUNT_ID', 'ACCOUNT #', 'ACCOUNT#']:
                account_id_col_pred = col
                break
        
        if not account_id_col_pred:
            raise Exception("Could not find AccountId column in predictions file")
        
        # Find AccountId column in ground truth
        account_id_col_gt = None
        for col in ground_truth_df.columns:
            col_upper = col.upper().strip()
            if col_upper in ['ACCOUNT #', 'ACCOUNT#', 'ACCOUNTID', 'ACCOUNT ID', 'ACCOUNT_ID']:
                account_id_col_gt = col
                break
        
        if not account_id_col_gt:
            raise Exception("Could not find Account # column in ground truth file")
        
        # Find CPT columns
        cpt_col_pred = None
        for col in predictions_df.columns:
            col_upper = col.upper().strip()
            if col_upper in ['CPT', 'ASA CODE', 'ASACODE', 'ASA_CODE', 'PROCEDURE CODE']:
                cpt_col_pred = col
                break
        
        cpt_col_gt = None
        for col in ground_truth_df.columns:
            col_upper = col.upper().strip()
            if col_upper in ['CPT', 'ASA CODE', 'ASACODE', 'ASA_CODE']:
                cpt_col_gt = col
                break
        
        # Find ICD columns in predictions
        icd_cols_pred = {}
        for i in range(1, 5):
            for col in predictions_df.columns:
                col_upper = col.upper().strip()
                if col_upper == f'ICD{i}' or col_upper == f'ICD {i}' or col_upper == f'ICD_{i}':
                    icd_cols_pred[i] = col
                    break
        
        # Find ICD column in ground truth (comma-separated)
        icd_col_gt = None
        for col in ground_truth_df.columns:
            col_upper = col.upper().strip()
            if col_upper in ['ICD', 'ICD CODES', 'ICD_CODES', 'ICDCODES']:
                icd_col_gt = col
                break
        
        # Find Anesthesia Type columns
        anesthesia_type_col_pred = None
        for col in predictions_df.columns:
            col_upper = col.upper().strip()
            if col_upper in ['ANESTHESIA TYPE', 'ANESTHESIATYPE', 'ANESTHESIA_TYPE']:
                anesthesia_type_col_pred = col
                break
        
        anesthesia_type_col_gt = None
        for col in ground_truth_df.columns:
            col_upper = col.upper().strip()
            if col_upper in ['ANESTHESIA TYPE', 'ANESTHESIATYPE', 'ANESTHESIA_TYPE']:
                anesthesia_type_col_gt = col
                break
        
        # Find Provider columns
        responsible_provider_col_pred = None
        for col in predictions_df.columns:
            col_upper = col.upper().strip()
            if col_upper == 'RESPONSIBLE PROVIDER':
                responsible_provider_col_pred = col
                break
        
        provider_col_gt = None
        for col in ground_truth_df.columns:
            col_upper = col.upper().strip()
            if col_upper == 'PROVIDER':
                provider_col_gt = col
                break
        
        # Find Surgeon columns
        surgeon_col_pred = None
        for col in predictions_df.columns:
            col_upper = col.upper().strip()
            if col_upper == 'SURGEON':
                surgeon_col_pred = col
                break
        
        surgeon_col_gt = None
        for col in ground_truth_df.columns:
            col_upper = col.upper().strip()
            if col_upper == 'SURGEON':
                surgeon_col_gt = col
                break
        
        # Create ground truth lookup dictionary
        gt_dict = {}
        for idx, row in ground_truth_df.iterrows():
            account_id = str(row[account_id_col_gt]).strip()
            if account_id and account_id.upper() != 'NAN':
                gt_dict[account_id] = {
                    'cpt': str(row[cpt_col_gt]).strip() if cpt_col_gt and pd.notna(row[cpt_col_gt]) else '',
                    'icd': str(row[icd_col_gt]).strip() if icd_col_gt and pd.notna(row[icd_col_gt]) else '',
                    'anesthesia_type': str(row[anesthesia_type_col_gt]).strip() if anesthesia_type_col_gt and pd.notna(row[anesthesia_type_col_gt]) else '',
                    'provider': str(row[provider_col_gt]).strip() if provider_col_gt and pd.notna(row[provider_col_gt]) else '',
                    'surgeon': str(row[surgeon_col_gt]).strip() if surgeon_col_gt and pd.notna(row[surgeon_col_gt]) else ''
                }
        
        # Helper function to normalize provider names
        def normalize_provider_name(provider_string, is_ground_truth=False):
            if pd.isna(provider_string) or not str(provider_string).strip():
                return ''
            
            provider_str = str(provider_string).strip()
            
            if ',' in provider_str:
                parts = [p.strip() for p in provider_str.split(',')]
                if len(parts) >= 1:
                    last_name = parts[0].strip()
                    if len(parts) >= 2:
                        first_middle = parts[1].strip()
                        name_parts = first_middle.split()
                        titles = {'MD', 'DO', 'DDS', 'DMD', 'RN', 'NP', 'PA', 'CRNA', 'M.D.', 'D.O.', 'MD.', 'DO.'}
                        name_parts = [p for p in name_parts if p.upper().rstrip('.') not in [t.rstrip('.') for t in titles]]
                        if len(name_parts) >= 1:
                            first_name = name_parts[0]
                            normalized = f"{last_name} {first_name}".strip()
                            return normalized.upper()
                        else:
                            return last_name.upper()
                    else:
                        return last_name.upper()
                else:
                    return ''
            else:
                parts = provider_str.split()
                if len(parts) >= 2:
                    return f"{parts[0]} {parts[1]}".upper()
                elif len(parts) == 1:
                    return parts[0].upper()
                else:
                    return ''
        
        # Dictionaries to store error PDFs by error type
        error_pdfs = {
            'anesthesia_type_errors': set(),
            'cpt_errors': set(),
            'icd_errors': set(),
            'provider_errors': set(),
            'surgeon_errors': set()
        }
        
        # Track totals for accuracy calculation
        total_accounts = 0
        anesthesia_total = 0
        cpt_total = 0
        icd_total = 0
        provider_total = 0
        surgeon_total = 0
        
        job.message = "Comparing predictions with ground truth..."
        job.progress = 50
        
        # Process each prediction
        for idx, row in predictions_df.iterrows():
            account_id = str(row[account_id_col_pred]).strip()
            source_file = str(row[source_file_col]).strip() if pd.notna(row[source_file_col]) else ''
            
            # Skip if no source file or account not in ground truth
            if not source_file or account_id not in gt_dict:
                continue
            
            total_accounts += 1
            gt_data = gt_dict[account_id]
            
            # Check CPT
            predicted_cpt = str(row[cpt_col_pred]).strip() if cpt_col_pred and pd.notna(row[cpt_col_pred]) else ''
            gt_cpt = gt_data['cpt']
            cpt_total += 1
            if predicted_cpt != gt_cpt:
                error_pdfs['cpt_errors'].add(source_file)
            
            # Check ICD
            predicted_icd_list = []
            for i in range(1, 5):
                if i in icd_cols_pred:
                    icd_val = str(row[icd_cols_pred[i]]).strip() if pd.notna(row[icd_cols_pred[i]]) else ''
                    if icd_val and icd_val.upper() != 'NAN':
                        predicted_icd_list.append(icd_val)
            
            gt_icd_list = []
            if gt_data['icd']:
                gt_icd_list = [code.strip() for code in gt_data['icd'].split(',') if code.strip()]
            
            # Check if ICD1 matches
            icd_match = True
            if len(predicted_icd_list) > 0 and len(gt_icd_list) > 0:
                icd_total += 1
                if predicted_icd_list[0] != gt_icd_list[0]:
                    icd_match = False
            
            if not icd_match:
                error_pdfs['icd_errors'].add(source_file)
            
            # Check Anesthesia Type
            if anesthesia_type_col_pred and anesthesia_type_col_gt:
                predicted_anesthesia = str(row[anesthesia_type_col_pred]).strip() if pd.notna(row[anesthesia_type_col_pred]) else ''
                gt_anesthesia = gt_data['anesthesia_type']
                if predicted_anesthesia and gt_anesthesia:
                    anesthesia_total += 1
                    if predicted_anesthesia.upper() != gt_anesthesia.upper():
                        error_pdfs['anesthesia_type_errors'].add(source_file)
            
            # Check Provider
            if responsible_provider_col_pred and provider_col_gt:
                predicted_provider = str(row[responsible_provider_col_pred]).strip() if pd.notna(row[responsible_provider_col_pred]) else ''
                gt_provider = gt_data['provider']
                if predicted_provider and gt_provider:
                    provider_total += 1
                    normalized_predicted = normalize_provider_name(predicted_provider, is_ground_truth=False)
                    normalized_gt = normalize_provider_name(gt_provider, is_ground_truth=True)
                    if normalized_predicted != normalized_gt:
                        error_pdfs['provider_errors'].add(source_file)
            
            # Check Surgeon
            if surgeon_col_pred and surgeon_col_gt:
                predicted_surgeon = str(row[surgeon_col_pred]).strip() if pd.notna(row[surgeon_col_pred]) else ''
                gt_surgeon = gt_data['surgeon']
                if predicted_surgeon and gt_surgeon:
                    surgeon_total += 1
                    normalized_predicted = normalize_provider_name(predicted_surgeon, is_ground_truth=False)
                    normalized_gt = normalize_provider_name(gt_surgeon, is_ground_truth=True)
                    if normalized_predicted != normalized_gt:
                        error_pdfs['surgeon_errors'].add(source_file)
        
        job.message = "Creating error PDFs ZIP..."
        job.progress = 70
        
        # Create output ZIP with error PDFs grouped by error type
        output_zip_path = f"/tmp/results/{job_id}_error_pdfs.zip"
        os.makedirs("/tmp/results", exist_ok=True)
        
        with zipfile.ZipFile(output_zip_path, 'w', zipfile.ZIP_DEFLATED) as output_zip:
            for error_type, pdf_files in error_pdfs.items():
                logger.info(f"Found {len(pdf_files)} PDFs with {error_type}")
                for pdf_file in pdf_files:
                    if pdf_file in pdf_mapping:
                        source_pdf_path = pdf_mapping[pdf_file]
                        # Add to ZIP in appropriate folder
                        arcname = f"{error_type}/{pdf_file}"
                        output_zip.write(source_pdf_path, arcname)
                        logger.info(f"Added {pdf_file} to {error_type} folder")
        
        job.message = "Creating Excel comparison report..."
        job.progress = 80
        
        # Also run the full comparison to create Excel report
        # Create a temporary job ID for the comparison
        comparison_job_id = f"{job_id}_comparison"
        comparison_job = ProcessingJob(comparison_job_id)
        job_status[comparison_job_id] = comparison_job
        
        # Run the comparison in the current thread (not as a background task)
        check_cpt_codes_background(comparison_job_id, predictions_path, ground_truth_path, charge_detail_path)
        
        # Get the Excel file from the comparison job
        excel_report_path = None
        if comparison_job.status == "completed" and comparison_job.result_file_xlsx:
            excel_report_path = comparison_job.result_file_xlsx
        
        # Clean up the comparison job from status
        del job_status[comparison_job_id]
        
        job.message = "Cleaning up..."
        job.progress = 90
        
        # Clean up temporary files
        shutil.rmtree(pdfs_temp_dir, ignore_errors=True)
        if os.path.exists(predictions_path):
            os.unlink(predictions_path)
        if os.path.exists(ground_truth_path):
            os.unlink(ground_truth_path)
        if os.path.exists(pdfs_zip_path):
            os.unlink(pdfs_zip_path)
        if charge_detail_path and os.path.exists(charge_detail_path):
            os.unlink(charge_detail_path)
        
        # Store both files
        job.result_file = output_zip_path  # Error PDFs ZIP
        job.result_file_xlsx = excel_report_path  # Excel comparison report
        job.status = "completed"
        job.progress = 100
        
        # Create summary message with totals and accuracies
        summary_parts = []
        
        # Anesthesia Type
        if anesthesia_total > 0:
            anesthesia_errors = len(error_pdfs['anesthesia_type_errors'])
            anesthesia_accuracy = ((anesthesia_total - anesthesia_errors) / anesthesia_total * 100)
            summary_parts.append(f"Anesthesia: {anesthesia_errors}/{anesthesia_total} errors ({anesthesia_accuracy:.1f}% accurate)")
        
        # CPT
        if cpt_total > 0:
            cpt_errors = len(error_pdfs['cpt_errors'])
            cpt_accuracy = ((cpt_total - cpt_errors) / cpt_total * 100)
            summary_parts.append(f"CPT: {cpt_errors}/{cpt_total} errors ({cpt_accuracy:.1f}% accurate)")
        
        # ICD
        if icd_total > 0:
            icd_errors = len(error_pdfs['icd_errors'])
            icd_accuracy = ((icd_total - icd_errors) / icd_total * 100)
            summary_parts.append(f"ICD: {icd_errors}/{icd_total} errors ({icd_accuracy:.1f}% accurate)")
        
        # Provider
        if provider_total > 0:
            provider_errors = len(error_pdfs['provider_errors'])
            provider_accuracy = ((provider_total - provider_errors) / provider_total * 100)
            summary_parts.append(f"Provider: {provider_errors}/{provider_total} errors ({provider_accuracy:.1f}% accurate)")
        
        # Surgeon
        if surgeon_total > 0:
            surgeon_errors = len(error_pdfs['surgeon_errors'])
            surgeon_accuracy = ((surgeon_total - surgeon_errors) / surgeon_total * 100)
            summary_parts.append(f"Surgeon: {surgeon_errors}/{surgeon_total} errors ({surgeon_accuracy:.1f}% accurate)")
        
        job.message = f"Error PDFs extracted! Total: {total_accounts} | {' | '.join(summary_parts) if summary_parts else 'No errors found'}"
        
        logger.info(f"CPT codes check with PDFs completed for job {job_id}")
        
    except Exception as e:
        logger.error(f"CPT codes check with PDFs background error: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        job.status = "failed"
        job.error = str(e)

def check_cpt_codes_background(job_id: str, predictions_path: str, ground_truth_path: str, charge_detail_path: str = None):
    """Background task to compare predictions vs ground truth and calculate accuracy
    
    Args:
        job_id: Unique job identifier
        predictions_path: Path to predictions file (demo detail report)
        ground_truth_path: Path to ground truth file (demo detail report)
        charge_detail_path: Optional path to charge detail report (for time comparison)
    """
    job = job_status[job_id]
    
    try:
        job.status = "processing"
        job.message = "Reading Excel files..."
        job.progress = 10
        
        import pandas as pd
        
        # Read predictions file (handle both Excel and CSV)
        predictions_path_obj = Path(predictions_path)
        predictions_df = None
        
        if predictions_path_obj.suffix.lower() in ('.xlsx', '.xls'):
            try:
                # Use appropriate engine based on file extension
                if predictions_path_obj.suffix.lower() == '.xlsx':
                    predictions_df = pd.read_excel(predictions_path, dtype=str, engine='openpyxl')
                else:  # .xls file
                    # Try xlrd first, then fall back to auto-detect
                    try:
                        predictions_df = pd.read_excel(predictions_path, dtype=str, engine='xlrd')
                    except Exception:
                        # Auto-detect engine (will use xlrd if available, otherwise openpyxl)
                        predictions_df = pd.read_excel(predictions_path, dtype=str, engine=None)
                logger.info(f"Successfully read predictions file as Excel ({predictions_path_obj.suffix.lower()})")
            except Exception as e:
                logger.warning(f"Failed to read predictions as Excel: {e}, trying CSV with multiple encodings")
                # Try multiple encodings for CSV files
                encodings_to_try = ['utf-8', 'utf-8-sig', 'latin-1', 'iso-8859-1', 'cp1252', 'windows-1252']
                for encoding in encodings_to_try:
                    try:
                        predictions_df = pd.read_csv(predictions_path, dtype=str, encoding=encoding)
                        logger.info(f"Successfully read predictions file as CSV with encoding: {encoding}")
                        break
                    except (UnicodeDecodeError, Exception) as csv_error:
                        continue
        else:
            # Try multiple encodings for CSV files
            encodings_to_try = ['utf-8', 'utf-8-sig', 'latin-1', 'iso-8859-1', 'cp1252', 'windows-1252']
            for encoding in encodings_to_try:
                try:
                    predictions_df = pd.read_csv(predictions_path, dtype=str, encoding=encoding)
                    logger.info(f"Successfully read predictions file with encoding: {encoding}")
                    break
                except (UnicodeDecodeError, Exception) as csv_error:
                    continue
        
        if predictions_df is None:
            raise Exception("Could not read predictions file. Please ensure it's a valid Excel (.xlsx, .xls) or CSV file.")
        
        # Read ground truth file (handle both Excel and CSV)
        ground_truth_path_obj = Path(ground_truth_path)
        ground_truth_df = None
        
        if ground_truth_path_obj.suffix.lower() in ('.xlsx', '.xls'):
            try:
                # Use appropriate engine based on file extension
                if ground_truth_path_obj.suffix.lower() == '.xlsx':
                    ground_truth_df = pd.read_excel(ground_truth_path, dtype=str, engine='openpyxl')
                else:  # .xls file
                    # Try xlrd first, then fall back to auto-detect
                    try:
                        ground_truth_df = pd.read_excel(ground_truth_path, dtype=str, engine='xlrd')
                    except Exception:
                        # Auto-detect engine (will use xlrd if available, otherwise openpyxl)
                        ground_truth_df = pd.read_excel(ground_truth_path, dtype=str, engine=None)
                logger.info(f"Successfully read ground truth file as Excel ({ground_truth_path_obj.suffix.lower()})")
            except Exception as e:
                logger.warning(f"Failed to read ground truth as Excel: {e}, trying CSV with multiple encodings")
                # Try multiple encodings for CSV files
                encodings_to_try = ['utf-8', 'utf-8-sig', 'latin-1', 'iso-8859-1', 'cp1252', 'windows-1252']
                for encoding in encodings_to_try:
                    try:
                        ground_truth_df = pd.read_csv(ground_truth_path, dtype=str, encoding=encoding)
                        logger.info(f"Successfully read ground truth file as CSV with encoding: {encoding}")
                        break
                    except (UnicodeDecodeError, Exception) as csv_error:
                        continue
        else:
            # Try multiple encodings for CSV files
            encodings_to_try = ['utf-8', 'utf-8-sig', 'latin-1', 'iso-8859-1', 'cp1252', 'windows-1252']
            for encoding in encodings_to_try:
                try:
                    ground_truth_df = pd.read_csv(ground_truth_path, dtype=str, encoding=encoding)
                    logger.info(f"Successfully read ground truth file with encoding: {encoding}")
                    break
                except (UnicodeDecodeError, Exception) as csv_error:
                    continue
        
        if ground_truth_df is None:
            raise Exception("Could not read ground truth file. Please ensure it's a valid Excel (.xlsx, .xls) or CSV file.")
        
        # Read charge detail report if provided (for time comparison)
        charge_detail_df = None
        if charge_detail_path and os.path.exists(charge_detail_path):
            job.message = "Reading charge detail report..."
            job.progress = 15
            
            charge_detail_path_obj = Path(charge_detail_path)
            if charge_detail_path_obj.suffix.lower() in ('.xlsx', '.xls'):
                try:
                    if charge_detail_path_obj.suffix.lower() == '.xlsx':
                        charge_detail_df = pd.read_excel(charge_detail_path, dtype=str, engine='openpyxl')
                    else:
                        try:
                            charge_detail_df = pd.read_excel(charge_detail_path, dtype=str, engine='xlrd')
                        except Exception:
                            charge_detail_df = pd.read_excel(charge_detail_path, dtype=str, engine=None)
                    logger.info(f"Successfully read charge detail report as Excel ({charge_detail_path_obj.suffix.lower()})")
                except Exception as e:
                    logger.warning(f"Failed to read charge detail report as Excel: {e}, trying CSV")
                    encodings_to_try = ['utf-8', 'utf-8-sig', 'latin-1', 'iso-8859-1', 'cp1252', 'windows-1252']
                    for encoding in encodings_to_try:
                        try:
                            charge_detail_df = pd.read_csv(charge_detail_path, dtype=str, encoding=encoding)
                            logger.info(f"Successfully read charge detail report as CSV with encoding: {encoding}")
                            break
                        except (UnicodeDecodeError, Exception):
                            continue
            else:
                encodings_to_try = ['utf-8', 'utf-8-sig', 'latin-1', 'iso-8859-1', 'cp1252', 'windows-1252']
                for encoding in encodings_to_try:
                    try:
                        charge_detail_df = pd.read_csv(charge_detail_path, dtype=str, encoding=encoding)
                        logger.info(f"Successfully read charge detail report with encoding: {encoding}")
                        break
                    except (UnicodeDecodeError, Exception):
                        continue
        
        job.message = "Finding required columns..."
        job.progress = 20
        
        # Find AccountId column in predictions (case-insensitive) - prioritize "Account #" format
        account_id_col_pred = None
        for col in predictions_df.columns:
            col_upper = col.upper().strip()
            if col_upper in ['ACCOUNT #', 'ACCOUNTID', 'ACCOUNT ID', 'ACCOUNT', 'ID', 'ACC. #', 'ACC #']:
                account_id_col_pred = col
                break
        
        # Find Cpt column in predictions (case-insensitive) - accepts "CPT", "Cpt", or "ASA Code"
        cpt_col_pred = None
        for col in predictions_df.columns:
            col_upper = col.upper().strip()
            if col_upper in ['CPT', 'ASA CODE']:
                cpt_col_pred = col
                break
        
        # Find ICD columns in predictions (ICD1, ICD2, ICD3, ICD4)
        icd_cols_pred = {}
        for col in predictions_df.columns:
            col_upper = col.upper().strip()
            if col_upper in ['ICD1', 'ICD2', 'ICD3', 'ICD4']:
                icd_cols_pred[col_upper] = col
        
        # Find Anesthesia Type column in predictions (if exists)
        anesthesia_type_col_pred = None
        for col in predictions_df.columns:
            col_upper = col.upper().strip()
            if col_upper == 'ANESTHESIA TYPE':
                anesthesia_type_col_pred = col
                break
        
        # Find AccountId column in ground truth (case-insensitive) - prioritize "Account #" format
        account_id_col_gt = None
        for col in ground_truth_df.columns:
            col_upper = col.upper().strip()
            if col_upper in ['ACCOUNT #', 'ACCOUNTID', 'ACCOUNT ID', 'ACCOUNT', 'ID', 'ACC. #', 'ACC #']:
                account_id_col_gt = col
                break
        
        # Find Cpt column in ground truth (case-insensitive) - supports "CPT" or "Cpt"
        cpt_col_gt = None
        for col in ground_truth_df.columns:
            col_upper = col.upper().strip()
            if col_upper in ['CPT']:
                cpt_col_gt = col
                break
        
        # Find Icd column in ground truth (case-insensitive) - comma-separated, supports "ICD" or "Icd"
        icd_col_gt = None
        for col in ground_truth_df.columns:
            col_upper = col.upper().strip()
            if col_upper in ['ICD']:
                icd_col_gt = col
                break
        
        # Find Anesthesia Type column in ground truth (if exists)
        anesthesia_type_col_gt = None
        for col in ground_truth_df.columns:
            col_upper = col.upper().strip()
            if col_upper == 'ANESTHESIA TYPE':
                anesthesia_type_col_gt = col
                break
        
        # Find Responsible Provider column in predictions (if exists)
        responsible_provider_col_pred = None
        for col in predictions_df.columns:
            col_upper = col.upper().strip()
            if col_upper == 'RESPONSIBLE PROVIDER':
                responsible_provider_col_pred = col
                break
        
        # Find Surgeon column in predictions (if exists)
        surgeon_col_pred = None
        for col in predictions_df.columns:
            col_upper = col.upper().strip()
            if col_upper == 'SURGEON':
                surgeon_col_pred = col
                break
        
        # Find Provider column in ground truth (if exists)
        provider_col_gt = None
        for col in ground_truth_df.columns:
            col_upper = col.upper().strip()
            if col_upper == 'PROVIDER':
                provider_col_gt = col
                break
        
        # Find Surgeon column in ground truth (if exists)
        surgeon_col_gt = None
        for col in ground_truth_df.columns:
            col_upper = col.upper().strip()
            if col_upper == 'SURGEON':
                surgeon_col_gt = col
                break
        
        # Find An Start and An Stop columns in predictions (if exists)
        an_start_col_pred = None
        an_stop_col_pred = None
        for col in predictions_df.columns:
            col_upper = col.upper().strip()
            if col_upper == 'AN START':
                an_start_col_pred = col
            elif col_upper == 'AN STOP':
                an_stop_col_pred = col
        
        # Find Start Time and Stop Time columns in charge detail report (if exists)
        start_time_col_charge = None
        stop_time_col_charge = None
        if charge_detail_df is not None:
            for col in charge_detail_df.columns:
                col_upper = col.upper().strip()
                if col_upper == 'START TIME':
                    start_time_col_charge = col
                elif col_upper == 'STOP TIME':
                    stop_time_col_charge = col
        
        # Validate required columns
        if account_id_col_pred is None:
            raise Exception("Predictions file must have an 'AccountId' or 'Account ID' column")
        if cpt_col_pred is None:
            raise Exception("Predictions file must have a 'Cpt' or 'ASA Code' column")
        if account_id_col_gt is None:
            raise Exception("Ground truth file must have an 'AccountId', 'Account ID', or 'Account #' column")
        if cpt_col_gt is None:
            raise Exception("Ground truth file must have a 'Cpt' column")
        
        job.message = "Matching accounts and comparing codes..."
        job.progress = 40
        
        # Helper function to parse comma-separated ICD codes into a list (preserving order)
        def parse_icd_codes(icd_string):
            """Parse comma-separated ICD codes into a list of normalized codes"""
            if pd.isna(icd_string) or not str(icd_string).strip():
                return []
            codes = [code.strip().upper() for code in str(icd_string).split(',')]
            return [c for c in codes if c]  # Remove empty strings
        
        # Helper function to parse predicted time strings (from predictions file)
        def parse_predicted_time(time_string):
            """Parse time string from predictions file into a comparable format.
            
            Expected format: "9/30/2025  9:53:00 AM" (date + time with AM/PM, double spaces)
            Returns time object (ignoring date) for comparison.
            """
            if pd.isna(time_string) or not str(time_string).strip():
                return None
            try:
                time_str = str(time_string).strip()
                # Normalize multiple spaces to single space for parsing
                time_str = ' '.join(time_str.split())
                
                from datetime import datetime
                # Parse the exact format: "9/30/2025 9:53:00 AM"
                dt = datetime.strptime(time_str, '%m/%d/%Y %I:%M:%S %p')
                return dt.time()  # Return time object (ignoring date) for comparison
            except Exception as e:
                logger.warning(f"Error parsing predicted time '{time_string}': {e}")
                return None
        
        # Helper function to parse ground truth time strings (from charge detail report)
        def parse_ground_truth_time(time_string):
            """Parse time string from ground truth charge detail report.
            
            Expected format: "15:41" (HH:MM, 24-hour format)
            Returns time object for comparison.
            """
            if pd.isna(time_string) or not str(time_string).strip():
                return None
            try:
                time_str = str(time_string).strip()
                from datetime import datetime
                # Parse the exact format: "15:41" (HH:MM)
                dt = datetime.strptime(time_str, '%H:%M')
                return dt.time()  # Return time object for comparison
            except Exception as e:
                logger.warning(f"Error parsing ground truth time '{time_string}': {e}")
                return None
        
        # Helper function to normalize provider names for comparison
        def normalize_provider_name(provider_string, is_ground_truth=False):
            """
            Normalize provider names to compare them regardless of format differences.
            
            Predictions format: "{last name}, {first name} {middle name}, {title}"
            Ground truth format: "{last name} {first name}" (NO middle name ever)
            
            IMPORTANT: Ground truth NEVER has middle names, so we ignore middle names
            from predictions when comparing.
            
            Returns normalized string: "{last name} {first name}" (all uppercase, no title, no middle name)
            """
            if pd.isna(provider_string) or not str(provider_string).strip():
                return ''
            
            provider_str = str(provider_string).strip()
            
            # Check if it's predictions format (has comma) or ground truth format (no comma)
            if ',' in provider_str:
                # Handle predictions format: "Last, First Middle, Title"
                # Split by comma first
                parts = [p.strip() for p in provider_str.split(',')]
                
                if len(parts) >= 1:
                    # First part is "Last"
                    last_name = parts[0].strip()
                    
                    # Second part (if exists) is "First Middle" (may have title)
                    if len(parts) >= 2:
                        first_middle = parts[1].strip()
                        # Remove common title suffixes if present (MD, DO, etc.)
                        # Split by space and take only name parts (ignore titles)
                        name_parts = first_middle.split()
                        # Filter out common titles
                        titles = {'MD', 'DO', 'DDS', 'DMD', 'RN', 'NP', 'PA', 'CRNA', 'M.D.', 'D.O.', 'MD.', 'DO.'}
                        name_parts = [p for p in name_parts if p.upper().rstrip('.') not in [t.rstrip('.') for t in titles]]
                        
                        if len(name_parts) >= 1:
                            # IMPORTANT: Only take first name, ignore middle name(s)
                            first_name = name_parts[0]
                            # Combine: "Last First" (no middle name)
                            normalized = f"{last_name} {first_name}".strip()
                            return normalized.upper()
                        else:
                            # Only last name
                            return last_name.upper()
                    else:
                        # Only last name provided
                        return last_name.upper()
                else:
                    return ''
            else:
                # Handle ground truth format: "Last First" (no comma, no middle name)
                # Just normalize by removing extra spaces and converting to uppercase
                normalized = ' '.join(provider_str.split())
                return normalized.upper()
        
        # Helper function to get predicted ICD codes as a list
        def get_predicted_icd_list(row):
            """Get predicted ICD codes from ICD1, ICD2, ICD3, ICD4 columns as a list"""
            icd_list = []
            for icd_num in ['ICD1', 'ICD2', 'ICD3', 'ICD4']:
                if icd_num in icd_cols_pred:
                    col_name = icd_cols_pred[icd_num]
                    value = row[col_name]
                    if pd.notna(value) and str(value).strip():
                        icd_list.append(str(value).strip().upper())
                    else:
                        icd_list.append('')  # Keep empty strings to preserve position
                else:
                    icd_list.append('')  # Column doesn't exist, use empty
            return icd_list
        
        # Create a dictionary for ground truth lookup (store full row data)
        # Store FIRST occurrence and also all additional rows for BLOCK checking
        gt_dict = {}
        gt_row_dict = {}  # Store full row data for detailed reporting
        gt_all_rows = {}  # Store ALL rows by account ID for BLOCK checking
        
        for idx, row in ground_truth_df.iterrows():
            account_id = str(row[account_id_col_gt]).strip()
            
            # Store ALL rows for this account ID (for BLOCK checking)
            if account_id not in gt_all_rows:
                gt_all_rows[account_id] = []
            gt_all_rows[account_id].append(row)
            
            # Only add FIRST occurrence to gt_dict (for main comparison)
            if account_id not in gt_dict:
                gt_dict[account_id] = {
                    'cpt': str(row[cpt_col_gt]).strip() if pd.notna(row[cpt_col_gt]) else '',
                    'icd': str(row[icd_col_gt]).strip() if icd_col_gt and pd.notna(row[icd_col_gt]) else '',
                    'anesthesia_type': str(row[anesthesia_type_col_gt]).strip() if anesthesia_type_col_gt and pd.notna(row[anesthesia_type_col_gt]) else '',
                    'provider': str(row[provider_col_gt]).strip() if provider_col_gt and pd.notna(row[provider_col_gt]) else '',
                    'surgeon': str(row[surgeon_col_gt]).strip() if surgeon_col_gt and pd.notna(row[surgeon_col_gt]) else ''
                }
                gt_row_dict[account_id] = row.to_dict()  # Store full row for detailed report
        
        # Group predictions by account ID (store ALL rows for each account)
        pred_all_rows = {}
        for idx, row in predictions_df.iterrows():
            account_id = str(row[account_id_col_pred]).strip()
            if account_id not in pred_all_rows:
                pred_all_rows[account_id] = []
            pred_all_rows[account_id].append(row)
        
        # Create a dictionary for charge detail report lookup (for time comparison)
        charge_detail_dict = {}
        charge_detail_row_dict = {}
        if charge_detail_df is not None and start_time_col_charge and stop_time_col_charge:
            # Find account ID column in charge detail report
            account_id_col_charge = None
            for col in charge_detail_df.columns:
                col_upper = col.upper().strip()
                if col_upper in ['ACCOUNTID', 'ACCOUNT ID', 'ACCOUNT', 'ID', 'ACC. #', 'ACC #', 'ACCOUNT #']:
                    account_id_col_charge = col
                    break
            
            if account_id_col_charge:
                for idx, row in charge_detail_df.iterrows():
                    account_id = str(row[account_id_col_charge]).strip()
                    if account_id not in charge_detail_dict:
                        charge_detail_dict[account_id] = {
                            'start_time': str(row[start_time_col_charge]).strip() if pd.notna(row[start_time_col_charge]) else '',
                            'stop_time': str(row[stop_time_col_charge]).strip() if pd.notna(row[stop_time_col_charge]) else ''
                        }
                        charge_detail_row_dict[account_id] = row.to_dict()
        
        # Compare predictions with ground truth
        comparison_data = []
        detailed_report_data = []
        case_overview_data = []  # Overview of each case
        cpt_matches = 0
        cpt_mismatches = 0
        icd_matches = 0
        icd_mismatches = 0
        icd1_matches = 0
        icd2_matches = 0
        icd3_matches = 0
        icd4_matches = 0
        anesthesia_matches = 0
        anesthesia_mismatches = 0
        provider_matches = 0
        provider_mismatches = 0
        surgeon_matches = 0
        surgeon_mismatches = 0
        start_time_matches = 0
        start_time_mismatches = 0
        stop_time_matches = 0
        stop_time_mismatches = 0
        # ICD1 mismatch analysis
        icd1_mismatches_total = 0
        icd1_mismatch_but_found_in_other_slots = 0
        icd1_critical_errors = 0
        not_found = 0
        # BLOCK checking (for additional rows beyond first row)
        block_matches = 0
        block_mismatches = 0
        block_total = 0
        
        # Track which account IDs have been processed (to only process first row in main loop)
        processed_accounts = set()
        
        for idx, row in predictions_df.iterrows():
            account_id = str(row[account_id_col_pred]).strip()
            
            # Skip if we've already processed this account ID (only process first occurrence)
            if account_id in processed_accounts:
                continue
            processed_accounts.add(account_id)
            
            predicted_cpt = str(row[cpt_col_pred]).strip() if pd.notna(row[cpt_col_pred]) else ''
            
            # Get predicted ICD codes as a list (position-based)
            predicted_icd_list = get_predicted_icd_list(row)
            
            # Get predicted Anesthesia Type
            predicted_anesthesia = ''
            if anesthesia_type_col_pred:
                predicted_anesthesia = str(row[anesthesia_type_col_pred]).strip() if pd.notna(row[anesthesia_type_col_pred]) else ''
            
            # Get predicted Responsible Provider
            predicted_provider = ''
            if responsible_provider_col_pred:
                predicted_provider = str(row[responsible_provider_col_pred]).strip() if pd.notna(row[responsible_provider_col_pred]) else ''
            
            # Get predicted Surgeon
            predicted_surgeon = ''
            if surgeon_col_pred:
                predicted_surgeon = str(row[surgeon_col_pred]).strip() if pd.notna(row[surgeon_col_pred]) else ''
            
            # Get predicted An Start and An Stop times (store raw and parsed)
            predicted_start_time_raw = ''
            predicted_stop_time_raw = ''
            predicted_start_time = None
            predicted_stop_time = None
            if an_start_col_pred:
                predicted_start_time_raw = str(row[an_start_col_pred]).strip() if pd.notna(row[an_start_col_pred]) else ''
                if predicted_start_time_raw:
                    predicted_start_time = parse_predicted_time(predicted_start_time_raw)
            if an_stop_col_pred:
                predicted_stop_time_raw = str(row[an_stop_col_pred]).strip() if pd.notna(row[an_stop_col_pred]) else ''
                if predicted_stop_time_raw:
                    predicted_stop_time = parse_predicted_time(predicted_stop_time_raw)
            
            # Start building detailed report entry
            detailed_entry = {
                'Row Number': idx + 1,  # 1-indexed for readability
                'AccountId': account_id,
                'Match Status': '',
                'Predicted Cpt': predicted_cpt,
                'Ground Truth Cpt': '',
                'Cpt Match': '',
                'Predicted Icd1': predicted_icd_list[0] if len(predicted_icd_list) > 0 else '',
                'Predicted Icd2': predicted_icd_list[1] if len(predicted_icd_list) > 1 else '',
                'Predicted Icd3': predicted_icd_list[2] if len(predicted_icd_list) > 2 else '',
                'Predicted Icd4': predicted_icd_list[3] if len(predicted_icd_list) > 3 else '',
                'Ground Truth Icd1': '',
                'Ground Truth Icd2': '',
                'Ground Truth Icd3': '',
                'Ground Truth Icd4': '',
                'Icd1 Match': '',
                'Icd2 Match': '',
                'Icd3 Match': '',
                'Icd4 Match': '',
                'Icd Match': '',
                'Predicted Anesthesia Type': predicted_anesthesia,
                'Ground Truth Anesthesia Type': '',
                'Anesthesia Type Match': '',
                'Predicted Responsible Provider': predicted_provider,
                'Ground Truth Provider': '',
                'Provider Match': '',
                'Predicted Surgeon': predicted_surgeon,
                'Ground Truth Surgeon': '',
                'Surgeon Match': '',
                'Predicted An Start': predicted_start_time_raw if an_start_col_pred and pd.notna(row[an_start_col_pred]) else '',
                'Predicted An Stop': predicted_stop_time_raw if an_stop_col_pred and pd.notna(row[an_stop_col_pred]) else '',
                'Ground Truth Start Time': '',
                'Ground Truth Stop Time': '',
                'Start Time Match': '',
                'Stop Time Match': '',
                'Overall Status': '',
                'Notes': '',
            }
            
            # Start building case overview entry
            case_overview = {
                'Case #': idx + 1,
                'Account ID': account_id,
                'Status': '',
                'Predicted Cpt': predicted_cpt,
                'Ground Truth Cpt': '',
                'Cpt Match': '',
                'Predicted Icd1': predicted_icd_list[0] if len(predicted_icd_list) > 0 else '',
                'Predicted Icd2': predicted_icd_list[1] if len(predicted_icd_list) > 1 else '',
                'Predicted Icd3': predicted_icd_list[2] if len(predicted_icd_list) > 2 else '',
                'Predicted Icd4': predicted_icd_list[3] if len(predicted_icd_list) > 3 else '',
                'Ground Truth Icd1': '',
                'Ground Truth Icd2': '',
                'Ground Truth Icd3': '',
                'Ground Truth Icd4': '',
                'Icd1 Match': '',
                'Icd2 Match': '',
                'Icd3 Match': '',
                'Icd4 Match': '',
                'Icd Match': '',
                'Predicted Anesthesia Type': predicted_anesthesia,
                'Ground Truth Anesthesia Type': '',
                'Anesthesia Type Match': '',
                'Predicted Responsible Provider': predicted_provider,
                'Ground Truth Provider': '',
                'Provider Match': '',
                'Predicted Surgeon': predicted_surgeon,
                'Ground Truth Surgeon': '',
                'Surgeon Match': '',
                'Result': '',
                'Notes': '',
            }
            
            # Add ALL columns from predictions file to overview
            excluded_cols = {account_id_col_pred, cpt_col_pred}
            excluded_cols.update(icd_cols_pred.values())
            if anesthesia_type_col_pred:
                excluded_cols.add(anesthesia_type_col_pred)
            if responsible_provider_col_pred:
                excluded_cols.add(responsible_provider_col_pred)
            if surgeon_col_pred:
                excluded_cols.add(surgeon_col_pred)
            
            for col in predictions_df.columns:
                if col not in excluded_cols:
                    value = row[col]
                    detailed_entry[f'Predictions: {col}'] = str(value) if pd.notna(value) else ''
                    case_overview[f'Predictions: {col}'] = str(value) if pd.notna(value) else ''
            
            if account_id in gt_dict:
                gt_data = gt_dict[account_id]
                gt_row = gt_row_dict[account_id]
                
                # Compare CPT codes
                gt_cpt = gt_data['cpt']
                cpt_match = predicted_cpt == gt_cpt
                if cpt_match:
                    cpt_matches += 1
                else:
                    cpt_mismatches += 1
                
                # Compare ICD codes - only check ICD1 (slot 1) for accuracy
                gt_icd_list = parse_icd_codes(gt_data['icd']) if gt_data['icd'] else []
                
                # Compare each position (for display purposes only)
                icd1_match = False
                icd2_match = False
                icd3_match = False
                icd4_match = False
                
                if len(predicted_icd_list) > 0 and len(gt_icd_list) > 0:
                    icd1_match = predicted_icd_list[0] == gt_icd_list[0]
                    if icd1_match:
                        icd1_matches += 1
                
                if len(predicted_icd_list) > 1 and len(gt_icd_list) > 1:
                    icd2_match = predicted_icd_list[1] == gt_icd_list[1]
                    if icd2_match:
                        icd2_matches += 1
                
                if len(predicted_icd_list) > 2 and len(gt_icd_list) > 2:
                    icd3_match = predicted_icd_list[2] == gt_icd_list[2]
                    if icd3_match:
                        icd3_matches += 1
                
                if len(predicted_icd_list) > 3 and len(gt_icd_list) > 3:
                    icd4_match = predicted_icd_list[3] == gt_icd_list[3]
                    if icd4_match:
                        icd4_matches += 1
                
                # Overall ICD match: only check ICD1 (slot 1) for accuracy calculation
                # Only count cases where BOTH predicted and ground truth have ICD1 codes
                if len(predicted_icd_list) > 0 and len(gt_icd_list) > 0:
                    predicted_icd1 = predicted_icd_list[0] if predicted_icd_list[0] else ''
                    gt_icd1 = gt_icd_list[0] if gt_icd_list[0] else ''
                    icd_match = predicted_icd1 == gt_icd1
                    
                    if icd_match:
                        icd_matches += 1
                    else:
                        icd_mismatches += 1
                        # Track ICD1 mismatch analysis
                        icd1_mismatches_total += 1
                        
                        # Check if this is a critical error (will cause claim denial)
                        from accuracy_utils import is_critical_error
                        is_critical = is_critical_error(predicted_icd1, gt_icd1)
                        if is_critical:
                            icd1_critical_errors += 1
                        
                        # Check if the correct ICD1 code appears in ICD2, ICD3, or ICD4
                        if gt_icd1:
                            found_in_other_slots = False
                            for slot_idx in [1, 2, 3]:  # Check ICD2, ICD3, ICD4 (indices 1, 2, 3)
                                if slot_idx < len(predicted_icd_list) and predicted_icd_list[slot_idx] == gt_icd1:
                                    found_in_other_slots = True
                                    break
                            if found_in_other_slots:
                                icd1_mismatch_but_found_in_other_slots += 1
                else:
                    # If either side is missing ICD1, don't count it in accuracy
                    icd_match = False
                
                # Compare Anesthesia Type (if present in both)
                anesthesia_match = None
                if anesthesia_type_col_pred and anesthesia_type_col_gt:
                    gt_anesthesia = gt_data['anesthesia_type']
                    anesthesia_match = predicted_anesthesia.upper() == gt_anesthesia.upper() if predicted_anesthesia and gt_anesthesia else False
                    if anesthesia_match:
                        anesthesia_matches += 1
                    elif predicted_anesthesia and gt_anesthesia:
                        anesthesia_mismatches += 1
                
                # Compare Responsible Provider (if present in predictions)
                provider_match = None
                if responsible_provider_col_pred and provider_col_gt:
                    gt_provider = gt_data['provider']
                    if predicted_provider and gt_provider:
                        # Normalize both names for comparison
                        # Ground truth never has middle names, so ignore middle names from predictions
                        normalized_predicted = normalize_provider_name(predicted_provider, is_ground_truth=False)
                        normalized_gt = normalize_provider_name(gt_provider, is_ground_truth=True)
                        provider_match = normalized_predicted == normalized_gt
                        if provider_match:
                            provider_matches += 1
                        else:
                            provider_mismatches += 1
                    elif not predicted_provider and not gt_provider:
                        # Both empty - consider as match
                        provider_match = True
                        provider_matches += 1
                    else:
                        # One is empty, one is not - mismatch
                        provider_match = False
                        provider_mismatches += 1
                
                # Compare Surgeon (if present in both)
                surgeon_match = None
                if surgeon_col_pred and surgeon_col_gt:
                    gt_surgeon = gt_data['surgeon']
                    if predicted_surgeon and gt_surgeon:
                        # Normalize both names for comparison
                        # Ground truth never has middle names, so ignore middle names from predictions
                        normalized_predicted = normalize_provider_name(predicted_surgeon, is_ground_truth=False)
                        normalized_gt = normalize_provider_name(gt_surgeon, is_ground_truth=True)
                        surgeon_match = normalized_predicted == normalized_gt
                        if surgeon_match:
                            surgeon_matches += 1
                        else:
                            surgeon_mismatches += 1
                    elif not predicted_surgeon and not gt_surgeon:
                        # Both empty - consider as match
                        surgeon_match = True
                        surgeon_matches += 1
                    else:
                        # One is empty, one is not - mismatch
                        surgeon_match = False
                        surgeon_mismatches += 1
                
                # Compare Start Time and Stop Time (if charge detail report is provided)
                start_time_match = None
                stop_time_match = None
                gt_start_time = None
                gt_stop_time = None
                
                if charge_detail_dict and account_id in charge_detail_dict:
                    charge_data = charge_detail_dict[account_id]
                    gt_start_time_str = charge_data.get('start_time', '')
                    gt_stop_time_str = charge_data.get('stop_time', '')
                    
                    # Parse ground truth times using the ground truth parser (HH:MM format)
                    if gt_start_time_str:
                        gt_start_time = parse_ground_truth_time(gt_start_time_str)
                    if gt_stop_time_str:
                        gt_stop_time = parse_ground_truth_time(gt_stop_time_str)
                    
                    # Compare start times
                    if predicted_start_time is not None and gt_start_time is not None:
                        start_time_match = predicted_start_time == gt_start_time
                        if start_time_match:
                            start_time_matches += 1
                        else:
                            start_time_mismatches += 1
                    elif predicted_start_time is None and gt_start_time is None:
                        # Both empty - consider as match
                        start_time_match = True
                        start_time_matches += 1
                    elif predicted_start_time is not None or gt_start_time is not None:
                        # One is empty, one is not - mismatch
                        start_time_match = False
                        start_time_mismatches += 1
                    
                    # Compare stop times
                    if predicted_stop_time is not None and gt_stop_time is not None:
                        stop_time_match = predicted_stop_time == gt_stop_time
                        if stop_time_match:
                            stop_time_matches += 1
                        else:
                            stop_time_mismatches += 1
                    elif predicted_stop_time is None and gt_stop_time is None:
                        # Both empty - consider as match
                        stop_time_match = True
                        stop_time_matches += 1
                    elif predicted_stop_time is not None or gt_stop_time is not None:
                        # One is empty, one is not - mismatch
                        stop_time_match = False
                        stop_time_mismatches += 1
                
                # Determine overall status
                all_match = cpt_match and icd_match
                if anesthesia_match is not None:
                    all_match = all_match and anesthesia_match
                # Note: Provider match is tracked but doesn't affect overall match status
                
                # Build status strings
                status_parts = []
                if cpt_match:
                    status_parts.append('✅ CPT')
                else:
                    status_parts.append('❌ CPT')
                
                if icd_match:
                    status_parts.append('✅ ICD')
                else:
                    status_parts.append('❌ ICD')
                
                if anesthesia_match is not None:
                    if anesthesia_match:
                        status_parts.append('✅ ANESTHESIA')
                    else:
                        status_parts.append('❌ ANESTHESIA')
                
                if provider_match is not None:
                    if provider_match:
                        status_parts.append('✅ PROVIDER')
                    else:
                        status_parts.append('❌ PROVIDER')
                
                if surgeon_match is not None:
                    if surgeon_match:
                        status_parts.append('✅ SURGEON')
                    else:
                        status_parts.append('❌ SURGEON')
                
                if start_time_match is not None:
                    if start_time_match:
                        status_parts.append('✅ START TIME')
                    else:
                        status_parts.append('❌ START TIME')
                
                if stop_time_match is not None:
                    if stop_time_match:
                        status_parts.append('✅ STOP TIME')
                    else:
                        status_parts.append('❌ STOP TIME')
                
                overall_status = ' | '.join(status_parts)
                
                if all_match:
                    detailed_entry['Match Status'] = 'Match'
                    detailed_entry['Overall Status'] = '✅ ALL MATCH'
                    case_overview['Status'] = '✅ ALL MATCH'
                    case_overview['Result'] = 'CORRECT'
                    notes = 'All codes match'
                else:
                    detailed_entry['Match Status'] = 'Mismatch'
                    detailed_entry['Overall Status'] = overall_status
                    case_overview['Status'] = overall_status
                    case_overview['Result'] = 'INCORRECT'
                    notes_parts = []
                    if not cpt_match:
                        notes_parts.append(f'CPT: predicted {predicted_cpt} vs ground truth {gt_cpt}')
                    if not icd_match:
                        pred_icd_str = ', '.join(predicted_icd_list) if predicted_icd_list else '(none)'
                        gt_icd_str = ', '.join(gt_icd_list) if gt_icd_list else '(none)'
                        notes_parts.append(f'ICD: predicted [{pred_icd_str}] vs ground truth [{gt_icd_str}]')
                    if anesthesia_match is False:
                        notes_parts.append(f'Anesthesia Type: predicted {predicted_anesthesia} vs ground truth {gt_data["anesthesia_type"]}')
                    if provider_match is False:
                        notes_parts.append(f'Provider: predicted {predicted_provider} vs ground truth {gt_data["provider"]}')
                    if surgeon_match is False:
                        notes_parts.append(f'Surgeon: predicted {predicted_surgeon} vs ground truth {gt_data["surgeon"]}')
                    if start_time_match is False:
                        notes_parts.append(f'Start Time: predicted {predicted_start_time or "N/A"} vs ground truth {gt_start_time or "N/A"}')
                    if stop_time_match is False:
                        notes_parts.append(f'Stop Time: predicted {predicted_stop_time or "N/A"} vs ground truth {gt_stop_time or "N/A"}')
                    notes = '; '.join(notes_parts)
                
                # Fill in comparison fields
                detailed_entry['Ground Truth Cpt'] = gt_cpt
                detailed_entry['Cpt Match'] = 'Yes' if cpt_match else 'No'
                case_overview['Ground Truth Cpt'] = gt_cpt
                case_overview['Cpt Match'] = '✅' if cpt_match else '❌'
                
                # Fill in ICD comparison fields
                detailed_entry['Ground Truth Icd1'] = gt_icd_list[0] if len(gt_icd_list) > 0 else ''
                detailed_entry['Ground Truth Icd2'] = gt_icd_list[1] if len(gt_icd_list) > 1 else ''
                detailed_entry['Ground Truth Icd3'] = gt_icd_list[2] if len(gt_icd_list) > 2 else ''
                detailed_entry['Ground Truth Icd4'] = gt_icd_list[3] if len(gt_icd_list) > 3 else ''
                detailed_entry['Icd1 Match'] = 'Yes' if icd1_match else 'No'
                detailed_entry['Icd2 Match'] = 'Yes' if icd2_match else 'No'
                detailed_entry['Icd3 Match'] = 'Yes' if icd3_match else 'No'
                detailed_entry['Icd4 Match'] = 'Yes' if icd4_match else 'No'
                detailed_entry['Icd Match'] = 'Yes' if icd_match else 'No'
                
                case_overview['Ground Truth Icd1'] = gt_icd_list[0] if len(gt_icd_list) > 0 else ''
                case_overview['Ground Truth Icd2'] = gt_icd_list[1] if len(gt_icd_list) > 1 else ''
                case_overview['Ground Truth Icd3'] = gt_icd_list[2] if len(gt_icd_list) > 2 else ''
                case_overview['Ground Truth Icd4'] = gt_icd_list[3] if len(gt_icd_list) > 3 else ''
                case_overview['Icd1 Match'] = '✅' if icd1_match else '❌'
                case_overview['Icd2 Match'] = '✅' if icd2_match else '❌'
                case_overview['Icd3 Match'] = '✅' if icd3_match else '❌'
                case_overview['Icd4 Match'] = '✅' if icd4_match else '❌'
                case_overview['Icd Match'] = '✅' if icd_match else '❌'
                
                if anesthesia_type_col_pred and anesthesia_type_col_gt:
                    detailed_entry['Ground Truth Anesthesia Type'] = gt_data['anesthesia_type']
                    detailed_entry['Anesthesia Type Match'] = 'Yes' if anesthesia_match else 'No'
                    case_overview['Ground Truth Anesthesia Type'] = gt_data['anesthesia_type']
                    case_overview['Anesthesia Type Match'] = '✅' if anesthesia_match else '❌'
                
                if responsible_provider_col_pred and provider_col_gt:
                    detailed_entry['Ground Truth Provider'] = gt_data['provider']
                    detailed_entry['Provider Match'] = 'Yes' if provider_match else 'No'
                    case_overview['Ground Truth Provider'] = gt_data['provider']
                    case_overview['Provider Match'] = '✅' if provider_match else '❌'
                
                if surgeon_col_pred and surgeon_col_gt:
                    detailed_entry['Ground Truth Surgeon'] = gt_data['surgeon']
                    detailed_entry['Surgeon Match'] = 'Yes' if surgeon_match else 'No'
                    case_overview['Ground Truth Surgeon'] = gt_data['surgeon']
                    case_overview['Surgeon Match'] = '✅' if surgeon_match else '❌'
                
                detailed_entry['Notes'] = notes
                case_overview['Notes'] = notes
                
                # Add ALL columns from ground truth file (with prefix)
                excluded_gt_cols = {account_id_col_gt, cpt_col_gt}
                if icd_col_gt:
                    excluded_gt_cols.add(icd_col_gt)
                if anesthesia_type_col_gt:
                    excluded_gt_cols.add(anesthesia_type_col_gt)
                
                for col in ground_truth_df.columns:
                    if col not in excluded_gt_cols:
                        value = gt_row[col]
                        detailed_entry[f'Ground Truth: {col}'] = str(value) if pd.notna(value) else ''
                        case_overview[f'Ground Truth: {col}'] = str(value) if pd.notna(value) else ''
                
                # Simple comparison entry (for backward compatibility)
                comparison_data.append({
                    'AccountId': account_id,
                    'Predicted Cpt': predicted_cpt,
                    'Ground Truth Cpt': gt_cpt,
                    'Cpt Match': 'Yes' if cpt_match else 'No',
                    'Predicted Icd1': predicted_icd_list[0] if len(predicted_icd_list) > 0 else '',
                    'Predicted Icd2': predicted_icd_list[1] if len(predicted_icd_list) > 1 else '',
                    'Predicted Icd3': predicted_icd_list[2] if len(predicted_icd_list) > 2 else '',
                    'Predicted Icd4': predicted_icd_list[3] if len(predicted_icd_list) > 3 else '',
                    'Ground Truth Icd1': gt_icd_list[0] if len(gt_icd_list) > 0 else '',
                    'Ground Truth Icd2': gt_icd_list[1] if len(gt_icd_list) > 1 else '',
                    'Ground Truth Icd3': gt_icd_list[2] if len(gt_icd_list) > 2 else '',
                    'Ground Truth Icd4': gt_icd_list[3] if len(gt_icd_list) > 3 else '',
                    'Icd1 Match': 'Yes' if icd1_match else 'No',
                    'Icd2 Match': 'Yes' if icd2_match else 'No',
                    'Icd3 Match': 'Yes' if icd3_match else 'No',
                    'Icd4 Match': 'Yes' if icd4_match else 'No',
                    'Icd Match': 'Yes' if icd_match else 'No',
                    'Predicted Anesthesia Type': predicted_anesthesia,
                    'Ground Truth Anesthesia Type': gt_data['anesthesia_type'] if anesthesia_type_col_gt else '',
                    'Anesthesia Type Match': 'Yes' if anesthesia_match else ('N/A' if anesthesia_match is None else 'No'),
                    'Predicted Responsible Provider': predicted_provider,
                    'Ground Truth Provider': gt_data['provider'] if provider_col_gt else '',
                    'Provider Match': 'Yes' if provider_match else ('N/A' if provider_match is None else 'No'),
                    'Predicted Surgeon': predicted_surgeon,
                    'Ground Truth Surgeon': gt_data['surgeon'] if surgeon_col_gt else '',
                    'Surgeon Match': 'Yes' if surgeon_match else ('N/A' if surgeon_match is None else 'No'),
                    'Predicted An Start': predicted_start_time_raw if an_start_col_pred and pd.notna(row[an_start_col_pred]) else '',
                    'Predicted An Stop': predicted_stop_time_raw if an_stop_col_pred and pd.notna(row[an_stop_col_pred]) else '',
                    'Ground Truth Start Time': charge_detail_dict[account_id]['start_time'] if charge_detail_dict and account_id in charge_detail_dict else '',
                    'Ground Truth Stop Time': charge_detail_dict[account_id]['stop_time'] if charge_detail_dict and account_id in charge_detail_dict else '',
                    'Start Time Match': 'Yes' if start_time_match else ('N/A' if start_time_match is None else 'No'),
                    'Stop Time Match': 'Yes' if stop_time_match else ('N/A' if stop_time_match is None else 'No'),
                    'Overall Match': 'Yes' if all_match else 'No',
                    'Status': 'Match' if all_match else 'Mismatch'
                })
            else:
                not_found += 1
                detailed_entry['Match Status'] = 'Account Not Found'
                detailed_entry['Overall Status'] = '⚠️ NOT FOUND'
                detailed_entry['Ground Truth Cpt'] = 'NOT FOUND'
                detailed_entry['Ground Truth Icd'] = 'NOT FOUND'
                detailed_entry['Ground Truth Anesthesia Type'] = 'NOT FOUND'
                detailed_entry['Ground Truth Provider'] = 'NOT FOUND'
                detailed_entry['Ground Truth Location'] = 'NOT FOUND'
                detailed_entry['Notes'] = 'Account ID not found in ground truth file'
                case_overview['Status'] = '⚠️ NOT FOUND'
                case_overview['Result'] = 'NO COMPARISON'
                case_overview['Ground Truth Cpt'] = 'NOT FOUND'
                case_overview['Ground Truth Icd'] = 'NOT FOUND'
                case_overview['Ground Truth Anesthesia Type'] = 'NOT FOUND'
                case_overview['Ground Truth Provider'] = 'NOT FOUND'
                case_overview['Ground Truth Surgeon'] = 'NOT FOUND'
                case_overview['Ground Truth Location'] = 'NOT FOUND'
                case_overview['Ground Truth Start Time'] = 'NOT FOUND'
                case_overview['Ground Truth Stop Time'] = 'NOT FOUND'
                case_overview['Notes'] = f'Account {account_id} not found in ground truth file'
                
                # Add empty columns for ground truth (to maintain structure)
                excluded_gt_cols = {account_id_col_gt, cpt_col_gt}
                if icd_col_gt:
                    excluded_gt_cols.add(icd_col_gt)
                if anesthesia_type_col_gt:
                    excluded_gt_cols.add(anesthesia_type_col_gt)
                if provider_col_gt:
                    excluded_gt_cols.add(provider_col_gt)
                if surgeon_col_gt:
                    excluded_gt_cols.add(surgeon_col_gt)
                
                for col in ground_truth_df.columns:
                    if col not in excluded_gt_cols:
                        detailed_entry[f'Ground Truth: {col}'] = ''
                
                # Simple comparison entry (for backward compatibility)
                comparison_data.append({
                    'AccountId': account_id,
                    'Predicted Cpt': predicted_cpt,
                    'Ground Truth Cpt': 'NOT FOUND',
                    'Cpt Match': 'No',
                    'Predicted Icd1': predicted_icd_list[0] if len(predicted_icd_list) > 0 else '',
                    'Predicted Icd2': predicted_icd_list[1] if len(predicted_icd_list) > 1 else '',
                    'Predicted Icd3': predicted_icd_list[2] if len(predicted_icd_list) > 2 else '',
                    'Predicted Icd4': predicted_icd_list[3] if len(predicted_icd_list) > 3 else '',
                    'Ground Truth Icd1': 'NOT FOUND',
                    'Ground Truth Icd2': 'NOT FOUND',
                    'Ground Truth Icd3': 'NOT FOUND',
                    'Ground Truth Icd4': 'NOT FOUND',
                    'Icd1 Match': 'No',
                    'Icd2 Match': 'No',
                    'Icd3 Match': 'No',
                    'Icd4 Match': 'No',
                    'Icd Match': 'No',
                    'Predicted Anesthesia Type': predicted_anesthesia,
                    'Ground Truth Anesthesia Type': 'NOT FOUND',
                    'Anesthesia Type Match': 'No',
                    'Predicted Responsible Provider': predicted_provider,
                    'Ground Truth Provider': 'NOT FOUND',
                    'Provider Match': 'No',
                    'Predicted Surgeon': predicted_surgeon,
                    'Ground Truth Surgeon': 'NOT FOUND',
                    'Surgeon Match': 'No',
                    'Predicted An Start': predicted_start_time_raw if an_start_col_pred and pd.notna(row[an_start_col_pred]) else '',
                    'Predicted An Stop': predicted_stop_time_raw if an_stop_col_pred and pd.notna(row[an_stop_col_pred]) else '',
                    'Ground Truth Start Time': 'NOT FOUND',
                    'Ground Truth Stop Time': 'NOT FOUND',
                    'Start Time Match': 'No',
                    'Stop Time Match': 'No',
                    'Overall Match': 'No',
                    'Status': 'Account Not Found'
                })
            
            # ========== BLOCK CHECKING ==========
            # Check additional rows (beyond first row) for this account ID
            # Compare CPT codes from additional rows as sets
            if account_id in pred_all_rows and account_id in gt_all_rows:
                pred_rows = pred_all_rows[account_id]
                gt_rows = gt_all_rows[account_id]
                
                # Only check if there are additional rows (more than 1 row)
                if len(pred_rows) > 1 or len(gt_rows) > 1:
                    block_total += 1
                    
                    # Collect CPT codes from additional rows (skip first row, index 0)
                    pred_additional_cpts = []
                    for i in range(1, len(pred_rows)):
                        cpt_val = str(pred_rows[i][cpt_col_pred]).strip() if pd.notna(pred_rows[i][cpt_col_pred]) else ''
                        if cpt_val and cpt_val.upper() != 'NAN':
                            pred_additional_cpts.append(cpt_val.upper())
                    
                    gt_additional_cpts = []
                    for i in range(1, len(gt_rows)):
                        cpt_val = str(gt_rows[i][cpt_col_gt]).strip() if pd.notna(gt_rows[i][cpt_col_gt]) else ''
                        if cpt_val and cpt_val.upper() != 'NAN':
                            gt_additional_cpts.append(cpt_val.upper())
                    
                    # Convert to sets for comparison (ignores duplicates and order)
                    pred_cpt_set = set(pred_additional_cpts)
                    gt_cpt_set = set(gt_additional_cpts)
                    
                    # Check if sets match
                    block_match = pred_cpt_set == gt_cpt_set
                    if block_match:
                        block_matches += 1
                        logger.info(f"BLOCK Match for Account {account_id}: Predicted additional CPTs {pred_cpt_set} == GT additional CPTs {gt_cpt_set}")
                    else:
                        block_mismatches += 1
                        logger.info(f"BLOCK Mismatch for Account {account_id}: Predicted additional CPTs {pred_cpt_set} != GT additional CPTs {gt_cpt_set}")
                    
                    # Add BLOCK check result to case overview and detailed entry
                    case_overview['Block Check'] = '✅' if block_match else '❌'
                    case_overview['Block Predicted CPTs'] = ', '.join(sorted(pred_cpt_set)) if pred_cpt_set else '(none)'
                    case_overview['Block Ground Truth CPTs'] = ', '.join(sorted(gt_cpt_set)) if gt_cpt_set else '(none)'
                    detailed_entry['Block Check'] = 'Yes' if block_match else 'No'
                    detailed_entry['Block Predicted CPTs'] = ', '.join(sorted(pred_cpt_set)) if pred_cpt_set else '(none)'
                    detailed_entry['Block Ground Truth CPTs'] = ', '.join(sorted(gt_cpt_set)) if gt_cpt_set else '(none)'
                else:
                    # No additional rows, so no block check needed
                    case_overview['Block Check'] = 'N/A'
                    case_overview['Block Predicted CPTs'] = ''
                    case_overview['Block Ground Truth CPTs'] = ''
                    detailed_entry['Block Check'] = 'N/A'
                    detailed_entry['Block Predicted CPTs'] = ''
                    detailed_entry['Block Ground Truth CPTs'] = ''
            else:
                # Account not found in one of the files, so no block check
                case_overview['Block Check'] = 'N/A'
                case_overview['Block Predicted CPTs'] = ''
                case_overview['Block Ground Truth CPTs'] = ''
                detailed_entry['Block Check'] = 'N/A'
                detailed_entry['Block Predicted CPTs'] = ''
                detailed_entry['Block Ground Truth CPTs'] = ''
            # ========== END BLOCK CHECKING ==========
            
            detailed_report_data.append(detailed_entry)
            case_overview_data.append(case_overview)
        
        job.message = "Creating comparison report..."
        job.progress = 70
        
        # Create comparison DataFrame
        comparison_df = pd.DataFrame(comparison_data)
        
        # Create detailed report DataFrame
        detailed_report_df = pd.DataFrame(detailed_report_data)
        
        # Create case overview DataFrame
        case_overview_df = pd.DataFrame(case_overview_data)
        
        # Calculate accuracy metrics (use separate denominators for CPT and ICD like AI refinement)
        cpt_total = cpt_matches + cpt_mismatches
        cpt_accuracy = (cpt_matches / cpt_total * 100) if cpt_total > 0 else 0
        icd_total = icd_matches + icd_mismatches
        icd_accuracy = (icd_matches / icd_total * 100) if icd_total > 0 else 0
        
        # Calculate ICD1 mismatch analysis metrics
        icd1_mismatch_percentage_found_in_other_slots = 0
        if icd1_mismatches_total > 0:
            icd1_mismatch_percentage_found_in_other_slots = (icd1_mismatch_but_found_in_other_slots / icd1_mismatches_total * 100)
        
        # Calculate critical error rate (percentage of ICD1 errors that are critical)
        icd1_critical_error_rate = (icd1_critical_errors / icd1_mismatches_total * 100) if icd1_mismatches_total > 0 else 0
        
        # Calculate theoretical ICD accuracy if "found in other slots" were counted as correct
        theoretical_icd_matches = icd_matches + icd1_mismatch_but_found_in_other_slots
        theoretical_icd_accuracy = (theoretical_icd_matches / icd_total * 100) if icd_total > 0 else 0
        
        overall_matches = sum(1 for entry in comparison_data if entry.get('Overall Match') == 'Yes' and entry.get('Status') != 'Account Not Found')
        total_comparable = cpt_total  # Use CPT total for overall
        overall_accuracy = (overall_matches / total_comparable * 100) if total_comparable > 0 else 0
        
        # Create summary sheet with more detailed metrics
        summary_data = {
            'Metric': [
                'Total Predictions',
                'Accounts Found in Ground Truth',
                'Accounts Not Found',
                'CPT Matches',
                'CPT Mismatches',
                'CPT Accuracy (%)',
                'ICD Matches',
                'ICD Mismatches',
                'ICD Accuracy (%)',
                'Anesthesia Type Matches',
                'Anesthesia Type Mismatches',
                'Provider Matches',
                'Provider Mismatches',
                'Provider Accuracy (%)',
                'Surgeon Matches',
                'Surgeon Mismatches',
                'Surgeon Accuracy (%)',
                'Block Checks (Accounts with Additional Rows)',
                'Block Matches',
                'Block Mismatches',
                'Block Accuracy (%)',
                'Start Time Matches',
                'Start Time Mismatches',
                'Start Time Accuracy (%)',
                'Stop Time Matches',
                'Stop Time Mismatches',
                'Stop Time Accuracy (%)',
                'ICD1 Mismatches (Total)',
                'ICD1 Critical Errors (Will Cause Denial)',
                'ICD1 Critical Error Rate (%)',
                'ICD1 Mismatches Found in ICD2-4',
                'ICD1 Mismatch Recovery Rate (%)',
                'Theoretical ICD Accuracy (%)',
                'Overall Matches (All Codes)',
                'Overall Accuracy (%)',
                'Not Found Rate (%)'
            ],
            'Value': [
                len(predictions_df),
                total_comparable,
                not_found,
                cpt_matches,
                cpt_mismatches,
                f'{cpt_accuracy:.2f}%',
                icd_matches,
                icd_mismatches,
                f'{icd_accuracy:.2f}%',
                anesthesia_matches,
                anesthesia_mismatches,
                provider_matches,
                provider_mismatches,
                f'{(provider_matches / (provider_matches + provider_mismatches) * 100):.2f}%' if (provider_matches + provider_mismatches) > 0 else 'N/A',
                surgeon_matches,
                surgeon_mismatches,
                f'{(surgeon_matches / (surgeon_matches + surgeon_mismatches) * 100):.2f}%' if (surgeon_matches + surgeon_mismatches) > 0 else 'N/A',
                block_total,
                block_matches,
                block_mismatches,
                f'{(block_matches / block_total * 100):.2f}%' if block_total > 0 else 'N/A',
                start_time_matches,
                start_time_mismatches,
                f'{(start_time_matches / (start_time_matches + start_time_mismatches) * 100):.2f}%' if (start_time_matches + start_time_mismatches) > 0 else 'N/A',
                stop_time_matches,
                stop_time_mismatches,
                f'{(stop_time_matches / (stop_time_matches + stop_time_mismatches) * 100):.2f}%' if (stop_time_matches + stop_time_mismatches) > 0 else 'N/A',
                icd1_mismatches_total,
                icd1_critical_errors,
                f'{icd1_critical_error_rate:.2f}%',
                icd1_mismatch_but_found_in_other_slots,
                f'{icd1_mismatch_percentage_found_in_other_slots:.2f}%',
                f'{theoretical_icd_accuracy:.2f}%',
                overall_matches,
                f'{overall_accuracy:.2f}%',
                f'{(not_found / len(predictions_df) * 100):.2f}%' if len(predictions_df) > 0 else '0.00%'
            ]
        }
        summary_df = pd.DataFrame(summary_data)
        
        job.message = "Generating Excel report..."
        job.progress = 85
        
        # Create output file path
        result_base = Path(f"/tmp/results/{job_id}_cpt_comparison")
        result_base.parent.mkdir(exist_ok=True)
        
        output_file_xlsx = result_base.with_suffix('.xlsx')
        
        # Write to Excel with multiple sheets
        with pd.ExcelWriter(output_file_xlsx, engine='openpyxl') as writer:
            # Summary sheet
            summary_df.to_excel(writer, sheet_name='Summary', index=False)
            
            # Case Overview sheet - shows each case with key information (PRIMARY VIEW)
            case_overview_df.to_excel(writer, sheet_name='Case Overview', index=False)
            
            # Detailed report sheet (all cases with full details)
            detailed_report_df.to_excel(writer, sheet_name='Detailed Report', index=False)
            
            # Simple comparison sheet (for backward compatibility)
            comparison_df.to_excel(writer, sheet_name='Comparison', index=False)
            
            # Create a sheet with only mismatches (from case overview)
            mismatches_overview_df = case_overview_df[case_overview_df['Status'].str.contains('MISMATCH', na=False)]
            mismatches_overview_df.to_excel(writer, sheet_name='Mismatches Overview', index=False)
            
            # Create a sheet with only matches (from case overview)
            matches_overview_df = case_overview_df[case_overview_df['Status'].str.contains('MATCH', na=False)]
            matches_overview_df.to_excel(writer, sheet_name='Matches Overview', index=False)
            
            # Create a sheet with accounts not found (from case overview)
            not_found_overview_df = case_overview_df[case_overview_df['Status'].str.contains('NOT FOUND', na=False)]
            not_found_overview_df.to_excel(writer, sheet_name='Not Found Overview', index=False)
            
            # Create a sheet with only mismatches (from detailed report)
            mismatches_detailed_df = detailed_report_df[detailed_report_df['Match Status'] == 'Mismatch']
            mismatches_detailed_df.to_excel(writer, sheet_name='Mismatches (Detailed)', index=False)
            
            # Create a sheet with only matches (from detailed report)
            matches_detailed_df = detailed_report_df[detailed_report_df['Match Status'] == 'Match']
            matches_detailed_df.to_excel(writer, sheet_name='Matches (Detailed)', index=False)
            
            # Create a sheet with accounts not found
            not_found_df = detailed_report_df[detailed_report_df['Match Status'] == 'Account Not Found']
            not_found_df.to_excel(writer, sheet_name='Not Found (Detailed)', index=False)
        
        job.result_file_xlsx = str(output_file_xlsx)
        job.result_file = str(output_file_xlsx)  # Use XLSX as primary result
        job.status = "completed"
        job.progress = 100
        provider_accuracy_str = f'{(provider_matches / (provider_matches + provider_mismatches) * 100):.2f}%' if (provider_matches + provider_mismatches) > 0 else 'N/A'
        surgeon_accuracy_str = f'{(surgeon_matches / (surgeon_matches + surgeon_mismatches) * 100):.2f}%' if (surgeon_matches + surgeon_mismatches) > 0 else 'N/A'
        block_accuracy_str = f'{(block_matches / block_total * 100):.2f}%' if block_total > 0 else 'N/A'
        job.message = f"Comparison completed! CPT: {cpt_accuracy:.2f}% | ICD: {icd_accuracy:.2f}% | Block: {block_accuracy_str} | Provider: {provider_accuracy_str} | Surgeon: {surgeon_accuracy_str} | Overall: {overall_accuracy:.2f}%"
        
        # Clean up input files
        if os.path.exists(predictions_path):
            os.unlink(predictions_path)
        if os.path.exists(ground_truth_path):
            os.unlink(ground_truth_path)
        if charge_detail_path and os.path.exists(charge_detail_path):
            os.unlink(charge_detail_path)
        
        logger.info(f"CPT+ICD codes check completed for job {job_id}: CPT {cpt_accuracy:.2f}%, ICD {icd_accuracy:.2f}%, Block {block_accuracy_str}, Provider {provider_accuracy_str}, Surgeon {surgeon_accuracy_str}, Overall {overall_accuracy:.2f}%")
        logger.info(f"Block Checking: {block_total} accounts with additional rows, {block_matches} matches, {block_mismatches} mismatches")
        logger.info(f"ICD1 Mismatch Analysis: {icd1_mismatches_total} total mismatches, {icd1_critical_errors} critical errors ({icd1_critical_error_rate:.2f}%), {icd1_mismatch_but_found_in_other_slots} found in ICD2-4 ({icd1_mismatch_percentage_found_in_other_slots:.2f}%), Theoretical ICD accuracy: {theoretical_icd_accuracy:.2f}%")
        
    except Exception as e:
        logger.error(f"CPT codes check background error: {str(e)}")
        job.status = "failed"
        job.error = str(e)
        job.message = f"Comparison failed: {str(e)}"
        
        # Clean up memory even on failure
        gc.collect()

@app.post("/check-cpt-codes")
async def check_cpt_codes(
    background_tasks: BackgroundTasks,
    predictions_file: UploadFile = File(...),
    ground_truth_file: UploadFile = File(...),
    charge_detail_file: UploadFile = File(None)
):
    """Upload CSV/XLSX files to compare CPT codes, ICD codes, and other fields.
    
    Args:
        predictions_file: Predictions file (Demo Detail Report) - required
        ground_truth_file: Ground truth file (Demo Detail Report) - required
        charge_detail_file: Charge Detail Report with Start Time and Stop Time - optional
    """
    
    try:
        logger.info(f"Received CPT codes check request - predictions: {predictions_file.filename}, ground_truth: {ground_truth_file.filename}, charge_detail: {charge_detail_file.filename if charge_detail_file else 'None'}")
        
        # Validate file types
        if not predictions_file.filename.endswith(('.csv', '.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Predictions file must be a CSV or XLSX")
        if not ground_truth_file.filename.endswith(('.csv', '.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Ground truth file must be a CSV or XLSX")
        if charge_detail_file and not charge_detail_file.filename.endswith(('.csv', '.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Charge detail file must be a CSV or XLSX")
        
        # Generate job ID
        job_id = str(uuid.uuid4())
        job = ProcessingJob(job_id)
        job_status[job_id] = job
        
        logger.info(f"Created CPT codes check job {job_id}")
        
        # Save uploaded files
        predictions_path = f"/tmp/{job_id}_predictions{Path(predictions_file.filename).suffix}"
        ground_truth_path = f"/tmp/{job_id}_ground_truth{Path(ground_truth_file.filename).suffix}"
        
        with open(predictions_path, "wb") as f:
            shutil.copyfileobj(predictions_file.file, f)
        with open(ground_truth_path, "wb") as f:
            shutil.copyfileobj(ground_truth_file.file, f)
        
        charge_detail_path = None
        if charge_detail_file:
            charge_detail_path = f"/tmp/{job_id}_charge_detail{Path(charge_detail_file.filename).suffix}"
            with open(charge_detail_path, "wb") as f:
                shutil.copyfileobj(charge_detail_file.file, f)
            logger.info(f"Files saved - predictions: {predictions_path}, ground_truth: {ground_truth_path}, charge_detail: {charge_detail_path}")
        else:
            logger.info(f"Files saved - predictions: {predictions_path}, ground_truth: {ground_truth_path}")
        
        # Start background processing
        background_tasks.add_task(check_cpt_codes_background, job_id, predictions_path, ground_truth_path, charge_detail_path)
        
        logger.info(f"Background CPT codes check task started for job {job_id}")
        
        return {"job_id": job_id, "message": "Files uploaded and CPT codes comparison started"}
        
    except Exception as e:
        logger.error(f"CPT codes check upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

@app.post("/check-cpt-codes-with-pdfs")
async def check_cpt_codes_with_pdfs_route(
    background_tasks: BackgroundTasks,
    predictions_file: UploadFile = File(...),
    ground_truth_file: UploadFile = File(...),
    pdfs_zip: UploadFile = File(...),
    charge_detail_file: UploadFile = File(None)
):
    """Check CPT codes and return error PDFs grouped by error type
    
    Args:
        predictions_file: Predictions file (Demo Detail Report) - required
        ground_truth_file: Ground truth file (Demo Detail Report) - required
        pdfs_zip: ZIP file containing all PDFs - required
        charge_detail_file: Charge Detail Report with Start Time and Stop Time - optional
    
    Returns:
        A ZIP file containing error PDFs grouped by error type in separate folders
    """
    
    try:
        logger.info(f"Received CPT codes check with PDFs request - predictions: {predictions_file.filename}, ground_truth: {ground_truth_file.filename}, pdfs_zip: {pdfs_zip.filename}")
        
        # Validate file types
        if not predictions_file.filename.endswith(('.csv', '.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Predictions file must be a CSV or XLSX")
        if not ground_truth_file.filename.endswith(('.csv', '.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Ground truth file must be a CSV or XLSX")
        if not pdfs_zip.filename.endswith('.zip'):
            raise HTTPException(status_code=400, detail="PDFs file must be a ZIP")
        if charge_detail_file and not charge_detail_file.filename.endswith(('.csv', '.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Charge detail file must be a CSV or XLSX")
        
        # Generate job ID
        job_id = str(uuid.uuid4())
        job = ProcessingJob(job_id)
        job_status[job_id] = job
        
        logger.info(f"Created CPT codes check with PDFs job {job_id}")
        
        # Save uploaded files
        predictions_path = f"/tmp/{job_id}_predictions{Path(predictions_file.filename).suffix}"
        ground_truth_path = f"/tmp/{job_id}_ground_truth{Path(ground_truth_file.filename).suffix}"
        pdfs_zip_path = f"/tmp/{job_id}_pdfs.zip"
        
        with open(predictions_path, "wb") as f:
            shutil.copyfileobj(predictions_file.file, f)
        with open(ground_truth_path, "wb") as f:
            shutil.copyfileobj(ground_truth_file.file, f)
        with open(pdfs_zip_path, "wb") as f:
            shutil.copyfileobj(pdfs_zip.file, f)
        
        charge_detail_path = None
        if charge_detail_file:
            charge_detail_path = f"/tmp/{job_id}_charge_detail{Path(charge_detail_file.filename).suffix}"
            with open(charge_detail_path, "wb") as f:
                shutil.copyfileobj(charge_detail_file.file, f)
            logger.info(f"Files saved - predictions: {predictions_path}, ground_truth: {ground_truth_path}, pdfs_zip: {pdfs_zip_path}, charge_detail: {charge_detail_path}")
        else:
            logger.info(f"Files saved - predictions: {predictions_path}, ground_truth: {ground_truth_path}, pdfs_zip: {pdfs_zip_path}")
        
        # Start background processing
        background_tasks.add_task(check_cpt_codes_with_pdfs_background, job_id, predictions_path, ground_truth_path, pdfs_zip_path, charge_detail_path)
        
        logger.info(f"Background CPT codes check with PDFs task started for job {job_id}")
        
        return {"job_id": job_id, "message": "Files uploaded and CPT codes comparison with PDF extraction started"}
        
    except Exception as e:
        logger.error(f"CPT codes check with PDFs upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

def analyze_field_errors_background(job_id: str, predictions_path: str, ground_truth_path: str, pdfs_zip_path: str, selected_fields: List[str], model: str = "gemini-3-flash-preview"):
    """Background task to analyze errors for selected fields using Gemini Flash
    
    Args:
        job_id: Unique job identifier
        predictions_path: Path to predictions file
        ground_truth_path: Path to ground truth file
        pdfs_zip_path: Path to ZIP file containing PDFs
        selected_fields: List of field names to analyze (e.g., ['CPT', 'ICD', 'Anesthesia Type'])
        model: Gemini model to use for analysis
    """
    job = job_status[job_id]
    
    try:
        job.status = "processing"
        job.message = "Reading files..."
        job.progress = 5
        
        import pandas as pd
        from google import genai
        from google.genai import types
        
        # Initialize Gemini client
        api_key = os.getenv("GOOGLE_API_KEY")
        if not api_key:
            raise Exception("Google API key not found")
        
        client = genai.Client(api_key=api_key)
        
        # Read predictions and ground truth files
        predictions_df = None
        ground_truth_df = None
        
        # Read predictions file
        predictions_path_obj = Path(predictions_path)
        if predictions_path_obj.suffix.lower() in ('.xlsx', '.xls'):
            if predictions_path_obj.suffix.lower() == '.xlsx':
                predictions_df = pd.read_excel(predictions_path, dtype=str, engine='openpyxl')
            else:
                try:
                    predictions_df = pd.read_excel(predictions_path, dtype=str, engine='xlrd')
                except:
                    predictions_df = pd.read_excel(predictions_path, dtype=str, engine=None)
        else:
            encodings = ['utf-8', 'utf-8-sig', 'latin-1', 'iso-8859-1', 'cp1252']
            for enc in encodings:
                try:
                    predictions_df = pd.read_csv(predictions_path, dtype=str, encoding=enc)
                    break
                except:
                    continue
        
        if predictions_df is None:
            raise Exception("Could not read predictions file")
        
        # Read ground truth file
        ground_truth_path_obj = Path(ground_truth_path)
        if ground_truth_path_obj.suffix.lower() in ('.xlsx', '.xls'):
            if ground_truth_path_obj.suffix.lower() == '.xlsx':
                ground_truth_df = pd.read_excel(ground_truth_path, dtype=str, engine='openpyxl')
            else:
                try:
                    ground_truth_df = pd.read_excel(ground_truth_path, dtype=str, engine='xlrd')
                except:
                    ground_truth_df = pd.read_excel(ground_truth_path, dtype=str, engine=None)
        else:
            encodings = ['utf-8', 'utf-8-sig', 'latin-1', 'iso-8859-1', 'cp1252']
            for enc in encodings:
                try:
                    ground_truth_df = pd.read_csv(ground_truth_path, dtype=str, encoding=enc)
                    break
                except:
                    continue
        
        if ground_truth_df is None:
            raise Exception("Could not read ground truth file")
        
        job.message = "Extracting PDFs..."
        job.progress = 10
        
        # Extract PDFs to temporary directory
        pdfs_temp_dir = f"/tmp/{job_id}_pdfs_extracted"
        os.makedirs(pdfs_temp_dir, exist_ok=True)
        
        with zipfile.ZipFile(pdfs_zip_path, 'r') as zip_ref:
            zip_ref.extractall(pdfs_temp_dir)
        
        # Create PDF mapping (filename -> path)
        pdf_mapping = {}
        for root, dirs, files in os.walk(pdfs_temp_dir):
            for file in files:
                if file.lower().endswith('.pdf'):
                    pdf_mapping[file] = os.path.join(root, file)
        
        job.message = "Finding columns..."
        job.progress = 15
        
        # Find columns in both dataframes
        def find_column(df, possible_names):
            for col in df.columns:
                if col.upper().strip() in [name.upper() for name in possible_names]:
                    return col
            return None
        
        # Predictions columns
        account_id_col_pred = find_column(predictions_df, ['Account #', 'AccountId', 'Account ID', 'Account', 'ID'])
        source_file_col = find_column(predictions_df, ['Source File', 'SourceFile', 'Filename', 'File Name', 'PDF', 'PDF File', 'source_file'])
        cpt_col_pred = find_column(predictions_df, ['CPT', 'Cpt', 'ASA Code'])
        anesthesia_type_col_pred = find_column(predictions_df, ['Anesthesia Type'])
        provider_col_pred = find_column(predictions_df, ['Responsible Provider'])
        surgeon_col_pred = find_column(predictions_df, ['Surgeon'])
        an_start_col_pred = find_column(predictions_df, ['An Start'])
        an_stop_col_pred = find_column(predictions_df, ['An Stop'])
        
        icd_cols_pred = {}
        for i in range(1, 5):
            col = find_column(predictions_df, [f'ICD{i}'])
            if col:
                icd_cols_pred[i] = col
        
        # Ground truth columns
        account_id_col_gt = find_column(ground_truth_df, ['Account #', 'AccountId', 'Account ID', 'Account', 'ID'])
        cpt_col_gt = find_column(ground_truth_df, ['CPT', 'Cpt'])
        icd_col_gt = find_column(ground_truth_df, ['ICD', 'Icd'])
        anesthesia_type_col_gt = find_column(ground_truth_df, ['Anesthesia Type'])
        provider_col_gt = find_column(ground_truth_df, ['Provider'])
        surgeon_col_gt = find_column(ground_truth_df, ['Surgeon'])
        
        if not account_id_col_pred or not account_id_col_gt:
            raise Exception("Account ID columns not found in files")
        
        if not source_file_col:
            raise Exception("Source File column not found in predictions file")
        
        job.message = "Building ground truth lookup..."
        job.progress = 20
        
        # Build ground truth lookup
        gt_dict = {}
        for idx, row in ground_truth_df.iterrows():
            account_id = str(row[account_id_col_gt]).strip()
            if account_id not in gt_dict:
                gt_dict[account_id] = {
                    'cpt': str(row[cpt_col_gt]).strip() if cpt_col_gt and pd.notna(row[cpt_col_gt]) else '',
                    'icd': str(row[icd_col_gt]).strip() if icd_col_gt and pd.notna(row[icd_col_gt]) else '',
                    'anesthesia_type': str(row[anesthesia_type_col_gt]).strip() if anesthesia_type_col_gt and pd.notna(row[anesthesia_type_col_gt]) else '',
                    'provider': str(row[provider_col_gt]).strip() if provider_col_gt and pd.notna(row[provider_col_gt]) else '',
                    'surgeon': str(row[surgeon_col_gt]).strip() if surgeon_col_gt and pd.notna(row[surgeon_col_gt]) else ''
                }
        
        job.message = "Collecting errors for selected fields..."
        job.progress = 25
        
        # Collect errors for each selected field
        field_errors = {field: [] for field in selected_fields}
        
        for idx, row in predictions_df.iterrows():
            account_id = str(row[account_id_col_pred]).strip()
            source_file = str(row[source_file_col]).strip() if pd.notna(row[source_file_col]) else ''
            
            if not source_file or account_id not in gt_dict:
                continue
            
            gt_data = gt_dict[account_id]
            pdf_path = pdf_mapping.get(source_file)
            
            if not pdf_path or not os.path.exists(pdf_path):
                continue
            
            # Check each selected field for errors
            if 'CPT' in selected_fields and cpt_col_pred and cpt_col_gt:
                predicted_cpt = str(row[cpt_col_pred]).strip() if pd.notna(row[cpt_col_pred]) else ''
                gt_cpt = gt_data['cpt']
                if predicted_cpt != gt_cpt:
                    field_errors['CPT'].append({
                        'account_id': account_id,
                        'predicted': predicted_cpt,
                        'ground_truth': gt_cpt,
                        'pdf_path': pdf_path,
                        'source_file': source_file
                    })
            
            if 'ICD' in selected_fields and icd_col_gt:
                # Get ICD1 (first ICD code)
                predicted_icd1 = ''
                if 1 in icd_cols_pred:
                    predicted_icd1 = str(row[icd_cols_pred[1]]).strip() if pd.notna(row[icd_cols_pred[1]]) else ''
                
                gt_icd_codes = [code.strip() for code in gt_data['icd'].split(',') if code.strip()]
                gt_icd1 = gt_icd_codes[0] if gt_icd_codes else ''
                
                if predicted_icd1 and gt_icd1 and predicted_icd1 != gt_icd1:
                    field_errors['ICD'].append({
                        'account_id': account_id,
                        'predicted': predicted_icd1,
                        'ground_truth': gt_icd1,
                        'pdf_path': pdf_path,
                        'source_file': source_file
                    })
            
            if 'Anesthesia Type' in selected_fields and anesthesia_type_col_pred and anesthesia_type_col_gt:
                predicted_anes = str(row[anesthesia_type_col_pred]).strip() if pd.notna(row[anesthesia_type_col_pred]) else ''
                gt_anes = gt_data['anesthesia_type']
                if predicted_anes and gt_anes and predicted_anes.upper() != gt_anes.upper():
                    field_errors['Anesthesia Type'].append({
                        'account_id': account_id,
                        'predicted': predicted_anes,
                        'ground_truth': gt_anes,
                        'pdf_path': pdf_path,
                        'source_file': source_file
                    })
            
            if 'Provider' in selected_fields and provider_col_pred and provider_col_gt:
                predicted_prov = str(row[provider_col_pred]).strip() if pd.notna(row[provider_col_pred]) else ''
                gt_prov = gt_data['provider']
                if predicted_prov and gt_prov and predicted_prov.upper() != gt_prov.upper():
                    field_errors['Provider'].append({
                        'account_id': account_id,
                        'predicted': predicted_prov,
                        'ground_truth': gt_prov,
                        'pdf_path': pdf_path,
                        'source_file': source_file
                    })
            
            if 'Surgeon' in selected_fields and surgeon_col_pred and surgeon_col_gt:
                predicted_surg = str(row[surgeon_col_pred]).strip() if pd.notna(row[surgeon_col_pred]) else ''
                gt_surg = gt_data['surgeon']
                if predicted_surg and gt_surg and predicted_surg.upper() != gt_surg.upper():
                    field_errors['Surgeon'].append({
                        'account_id': account_id,
                        'predicted': predicted_surg,
                        'ground_truth': gt_surg,
                        'pdf_path': pdf_path,
                        'source_file': source_file
                    })
        
        job.message = "Analyzing errors with Gemini..."
        job.progress = 30
        
        # Analyze each field's errors with Gemini
        analysis_results = {}
        progress_per_field = 60 / max(len(selected_fields), 1)
        
        for field_idx, field_name in enumerate(selected_fields):
            errors = field_errors.get(field_name, [])
            
            if not errors:
                analysis_results[field_name] = {
                    'total_errors': 0,
                    'error_analyses': [],
                    'pattern_summary': 'No errors found for this field.'
                }
                continue
            
            job.message = f"Analyzing {field_name} errors ({len(errors)} errors)..."
            job.progress = 30 + int(field_idx * progress_per_field)
            
            error_analyses = []
            
            # Analyze each error (limit to first 20 to avoid excessive API calls)
            for error_idx, error in enumerate(errors[:20]):
                try:
                    # Convert PDF to images for Gemini using PyMuPDF
                    import fitz  # PyMuPDF
                    import base64
                    import io
                    from PIL import Image
                    
                    # Open PDF with PyMuPDF
                    pdf_doc = fitz.open(error['pdf_path'])
                    image_data_list = []
                    
                    # Convert first 5 pages to images
                    for page_num in range(min(5, len(pdf_doc))):
                        page = pdf_doc[page_num]
                        pix = page.get_pixmap(dpi=150)
                        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                        
                        buffer = io.BytesIO()
                        img.save(buffer, format='PNG')
                        img_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
                        image_data_list.append(img_base64)
                    
                    pdf_doc.close()
                    
                    # Build prompt for Gemini
                    prompt = f"""You are a medical coding expert analyzing prediction errors.

Field: {field_name}
Account ID: {error['account_id']}

AI Predicted: {error['predicted']}
Ground Truth: {error['ground_truth']}

Please analyze the PDF images provided and explain:
1. Why did the AI make this specific error?
2. What information in the PDF led to the incorrect prediction?
3. What information should have led to the correct answer?

Provide a concise explanation (2-3 sentences max)."""

                    # Create Gemini request
                    parts = [types.Part.from_text(text=prompt)]
                    for img_data in image_data_list:
                        parts.append(types.Part.from_bytes(
                            data=base64.b64decode(img_data),
                            mime_type="image/png"
                        ))
                    
                    response = client.models.generate_content(
                        model=model,
                        contents=types.Content(
                            role="user",
                            parts=parts
                        )
                    )
                    
                    explanation = response.text.strip() if response.text else "No explanation generated"
                    
                    error_analyses.append({
                        'account_id': error['account_id'],
                        'predicted': error['predicted'],
                        'ground_truth': error['ground_truth'],
                        'explanation': explanation
                    })
                    
                except Exception as e:
                    logger.error(f"Error analyzing {field_name} for account {error['account_id']}: {str(e)}")
                    error_analyses.append({
                        'account_id': error['account_id'],
                        'predicted': error['predicted'],
                        'ground_truth': error['ground_truth'],
                        'explanation': f"Analysis failed: {str(e)}"
                    })
            
            # Generate pattern summary for this field
            if len(error_analyses) > 0:
                job.message = f"Generating pattern summary for {field_name}..."
                
                # Build summary prompt
                error_summary = "\n\n".join([
                    f"Account {err['account_id']}: Predicted '{err['predicted']}', Expected '{err['ground_truth']}'\nExplanation: {err['explanation']}"
                    for err in error_analyses
                ])
                
                pattern_prompt = f"""Based on the following error analyses for the {field_name} field, provide a high-level summary of error patterns:

{error_summary}

Please identify:
1. Common patterns in the errors (e.g., specific types of misclassifications)
2. Root causes (e.g., OCR issues, ambiguous documentation, specific medical terms causing confusion)
3. Recommendations for improvement

Provide a concise summary (3-5 sentences)."""

                try:
                    pattern_response = client.models.generate_content(
                        model=model,
                        contents=pattern_prompt
                    )
                    pattern_summary = pattern_response.text.strip() if pattern_response.text else "No pattern summary generated"
                except Exception as e:
                    logger.error(f"Error generating pattern summary for {field_name}: {str(e)}")
                    pattern_summary = f"Pattern analysis failed: {str(e)}"
            else:
                pattern_summary = "No errors were successfully analyzed."
            
            analysis_results[field_name] = {
                'total_errors': len(errors),
                'errors_analyzed': len(error_analyses),
                'error_analyses': error_analyses,
                'pattern_summary': pattern_summary
            }
        
        job.message = "Creating analysis report..."
        job.progress = 95
        
        # Create output report (JSON and human-readable text)
        output_json_path = f"/tmp/results/{job_id}_error_analysis.json"
        output_txt_path = f"/tmp/results/{job_id}_error_analysis.txt"
        os.makedirs("/tmp/results", exist_ok=True)
        
        # Save JSON report
        with open(output_json_path, 'w') as f:
            json.dump(analysis_results, f, indent=2)
        
        # Create human-readable text report
        with open(output_txt_path, 'w') as f:
            f.write("=" * 80 + "\n")
            f.write("FIELD ERROR ANALYSIS REPORT\n")
            f.write("=" * 80 + "\n\n")
            
            for field_name, results in analysis_results.items():
                f.write(f"\n{'='*80}\n")
                f.write(f"FIELD: {field_name}\n")
                f.write(f"{'='*80}\n")
                f.write(f"Total Errors: {results['total_errors']}\n")
                f.write(f"Errors Analyzed: {results.get('errors_analyzed', 0)}\n\n")
                
                f.write("PATTERN SUMMARY:\n")
                f.write("-" * 80 + "\n")
                f.write(results['pattern_summary'] + "\n\n")
                
                f.write("INDIVIDUAL ERROR ANALYSES:\n")
                f.write("-" * 80 + "\n")
                for err in results.get('error_analyses', []):
                    f.write(f"\nAccount ID: {err['account_id']}\n")
                    f.write(f"Predicted: {err['predicted']}\n")
                    f.write(f"Ground Truth: {err['ground_truth']}\n")
                    f.write(f"Explanation: {err['explanation']}\n")
                    f.write("-" * 80 + "\n")
        
        # Clean up
        shutil.rmtree(pdfs_temp_dir, ignore_errors=True)
        if os.path.exists(predictions_path):
            os.unlink(predictions_path)
        if os.path.exists(ground_truth_path):
            os.unlink(ground_truth_path)
        if os.path.exists(pdfs_zip_path):
            os.unlink(pdfs_zip_path)
        
        job.status = "completed"
        job.result_file = output_txt_path
        job.result_file_json = output_json_path
        job.progress = 100
        job.message = f"Analysis complete! Analyzed {sum(r['total_errors'] for r in analysis_results.values())} total errors across {len(selected_fields)} field(s)"
        
        logger.info(f"Field error analysis completed for job {job_id}")
        
    except Exception as e:
        logger.error(f"Field error analysis error: {str(e)}")
        import traceback
        logger.error(traceback.format_exc())
        job.status = "failed"
        job.error = str(e)


@app.post("/analyze-field-errors")
async def analyze_field_errors(
    background_tasks: BackgroundTasks,
    predictions_file: UploadFile = File(...),
    ground_truth_file: UploadFile = File(...),
    pdfs_zip: UploadFile = File(...),
    selected_fields: str = Form(...)  # JSON string array
):
    """Analyze errors for selected fields using Gemini Flash
    
    Args:
        predictions_file: Predictions file (Demo Detail Report)
        ground_truth_file: Ground truth file (Demo Detail Report)
        pdfs_zip: ZIP file containing all PDFs
        selected_fields: JSON string array of field names (e.g., '["CPT", "ICD", "Anesthesia Type"]')
    """
    
    try:
        logger.info(f"Received field error analysis request")
        
        # Parse selected fields
        fields_list = json.loads(selected_fields)
        if not isinstance(fields_list, list) or len(fields_list) == 0:
            raise HTTPException(status_code=400, detail="selected_fields must be a non-empty JSON array")
        
        logger.info(f"Selected fields: {fields_list}")
        
        # Validate file types
        if not predictions_file.filename.endswith(('.csv', '.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Predictions file must be CSV or XLSX")
        if not ground_truth_file.filename.endswith(('.csv', '.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Ground truth file must be CSV or XLSX")
        if not pdfs_zip.filename.endswith('.zip'):
            raise HTTPException(status_code=400, detail="PDFs file must be a ZIP")
        
        # Generate job ID
        job_id = str(uuid.uuid4())
        job = ProcessingJob(job_id)
        job_status[job_id] = job
        
        # Save uploaded files
        predictions_path = f"/tmp/{job_id}_predictions{Path(predictions_file.filename).suffix}"
        ground_truth_path = f"/tmp/{job_id}_ground_truth{Path(ground_truth_file.filename).suffix}"
        pdfs_zip_path = f"/tmp/{job_id}_pdfs.zip"
        
        with open(predictions_path, "wb") as f:
            shutil.copyfileobj(predictions_file.file, f)
        with open(ground_truth_path, "wb") as f:
            shutil.copyfileobj(ground_truth_file.file, f)
        with open(pdfs_zip_path, "wb") as f:
            shutil.copyfileobj(pdfs_zip.file, f)
        
        # Start background analysis
        background_tasks.add_task(
            analyze_field_errors_background,
            job_id,
            predictions_path,
            ground_truth_path,
            pdfs_zip_path,
            fields_list
        )
        
        logger.info(f"Background field error analysis task started for job {job_id}")
        
        return {"job_id": job_id, "message": "Field error analysis started"}
        
    except json.JSONDecodeError:
        raise HTTPException(status_code=400, detail="selected_fields must be valid JSON")
    except Exception as e:
        logger.error(f"Field error analysis upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

@app.post("/generate-modifiers")
async def generate_modifiers_route(
    background_tasks: BackgroundTasks,
    csv_file: UploadFile = File(...),
    turn_off_medical_direction: bool = Form(False),
    generate_qk_duplicate: bool = Form(False),
    limit_anesthesia_time: bool = Form(False),
    turn_off_bcbs_medicare_modifiers: bool = Form(True),
    peripheral_blocks_mode: str = Form("other"),
    add_pt_for_non_medicare: bool = Form(False),
    change_responsible_provider_to_md_if_p_only: bool = Form(False)
):
    """Upload a CSV or XLSX file to generate medical modifiers
    
    Args:
        csv_file: CSV or XLSX file with billing data
        turn_off_medical_direction: If True, override all medical direction YES to NO
        generate_qk_duplicate: If True, generate duplicate line when QK modifier is applied with CRNA as Responsible Provider
        limit_anesthesia_time: If True, limit anesthesia time to maximum 480 minutes based on An Start and An Stop columns
        turn_off_bcbs_medicare_modifiers: If True, for MedNet Code 003 (BCBS), only generate P modifiers (no M1/GC/QS)
        peripheral_blocks_mode: Mode for peripheral block generation - "UNI" (only General) or "other" (not MAC)
        add_pt_for_non_medicare: If True, add PT modifier for non-Medicare insurances when polyps found and screening colonoscopy
        change_responsible_provider_to_md_if_p_only: If True, change Responsible Provider to MD when only P modifier is used and both MD and CRNA are present
    """
    
    try:
        logger.info(f"Received modifiers generation request - file: {csv_file.filename}, turn_off_medical_direction: {turn_off_medical_direction}, generate_qk_duplicate: {generate_qk_duplicate}, limit_anesthesia_time: {limit_anesthesia_time}, turn_off_bcbs_medicare_modifiers: {turn_off_bcbs_medicare_modifiers}, peripheral_blocks_mode: {peripheral_blocks_mode}, add_pt_for_non_medicare: {add_pt_for_non_medicare}, change_responsible_provider_to_md_if_p_only: {change_responsible_provider_to_md_if_p_only}")
        
        # Validate file type
        if not csv_file.filename.endswith(('.csv', '.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="File must be a CSV or XLSX")
        
        # Generate job ID
        job_id = str(uuid.uuid4())
        job = ProcessingJob(job_id)
        job_status[job_id] = job
        
        logger.info(f"Created modifiers generation job {job_id}")
        
        # Save uploaded file
        input_path = f"/tmp/{job_id}_modifiers_input{Path(csv_file.filename).suffix}"
        
        with open(input_path, "wb") as f:
            shutil.copyfileobj(csv_file.file, f)
        
        logger.info(f"File saved - path: {input_path}")
        
        # Convert to CSV if needed
        csv_path = ensure_csv_file(input_path, f"/tmp/{job_id}_modifiers_input.csv")
        
        # Clean up original file if it was converted
        if csv_path != input_path and os.path.exists(input_path):
            os.unlink(input_path)
        
        logger.info(f"CSV ready - path: {csv_path}")
        
        # Start background processing
        background_tasks.add_task(generate_modifiers_background, job_id, csv_path, turn_off_medical_direction, generate_qk_duplicate, limit_anesthesia_time, turn_off_bcbs_medicare_modifiers, peripheral_blocks_mode, add_pt_for_non_medicare, change_responsible_provider_to_md_if_p_only)
        
        logger.info(f"Background modifiers generation task started for job {job_id}")
        
        qk_msg = ' + QK Duplicates' if generate_qk_duplicate else ''
        time_limit_msg = ' + Time Limited' if limit_anesthesia_time else ''
        bcbs_msg = ' + BCBS Modifiers OFF' if turn_off_bcbs_medicare_modifiers else ''
        blocks_mode_msg = f' (Blocks: {peripheral_blocks_mode})' if peripheral_blocks_mode else ''
        pt_non_medicare_msg = ' + PT for Non-Medicare' if add_pt_for_non_medicare else ''
        pt_non_medicare_msg = ' + PT for Non-Medicare' if add_pt_for_non_medicare else ''
        return {"job_id": job_id, "message": f"File uploaded and modifiers generation started{' (Medical Direction OFF)' if turn_off_medical_direction else ''}{qk_msg}{time_limit_msg}{bcbs_msg}{blocks_mode_msg}{pt_non_medicare_msg}"}
        
    except Exception as e:
        logger.error(f"Modifiers generation upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

@app.post("/predict-insurance-codes")
async def predict_insurance_codes(
    background_tasks: BackgroundTasks,
    data_csv: UploadFile = File(...),
    special_cases_csv: UploadFile = File(None),
    special_cases_template_id: int = Form(None),
    enable_ai: bool = Form(True)
):
    """
    Upload data CSV/XLSX and provide special cases either via:
    - Upload CSV file (special_cases_csv)
    - Select saved template (special_cases_template_id)
    """
    
    try:
        logger.info(f"Received insurance code prediction request - data: {data_csv.filename}, AI enabled: {enable_ai}")
        
        # Validate file type
        if not data_csv.filename.endswith(('.csv', '.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Data file must be a CSV or XLSX")
        
        if special_cases_csv and not special_cases_csv.filename.endswith(('.csv', '.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Special cases file must be a CSV or XLSX")
        
        # Generate job ID
        job_id = str(uuid.uuid4())
        job = ProcessingJob(job_id)
        job_status[job_id] = job
        
        logger.info(f"Created insurance prediction job {job_id}")
        
        # Save uploaded data file
        data_input_path = f"/tmp/{job_id}_data_input{Path(data_csv.filename).suffix}"
        
        with open(data_input_path, "wb") as f:
            shutil.copyfileobj(data_csv.file, f)
        
        logger.info(f"Data file saved - path: {data_input_path}")
        
        # Convert to CSV if needed
        data_csv_path = ensure_csv_file(data_input_path, f"/tmp/{job_id}_data_input.csv")
        
        # Clean up original file if it was converted
        if data_csv_path != data_input_path and os.path.exists(data_input_path):
            os.unlink(data_input_path)
        
        logger.info(f"Data CSV ready - path: {data_csv_path}")
        
        # Handle special cases - either from upload or from template
        special_cases_csv_path = None
        if special_cases_template_id:
            # Load from database template
            from db_utils import get_special_cases_template
            template = get_special_cases_template(template_id=special_cases_template_id)
            
            if not template:
                raise HTTPException(status_code=404, detail=f"Special cases template {special_cases_template_id} not found")
            
            # Create CSV file from template mappings
            import pandas as pd
            mappings_df = pd.DataFrame(template['mappings'])
            mappings_df.columns = ['Company name', 'Mednet code']
            
            special_cases_csv_path = f"/tmp/{job_id}_special_cases.csv"
            mappings_df.to_csv(special_cases_csv_path, index=False)
            logger.info(f"Special cases loaded from template {special_cases_template_id} - {len(template['mappings'])} mappings")
            
        elif special_cases_csv:
            # Load from uploaded file
            special_cases_input_path = f"/tmp/{job_id}_special_cases{Path(special_cases_csv.filename).suffix}"
            with open(special_cases_input_path, "wb") as f:
                shutil.copyfileobj(special_cases_csv.file, f)
            logger.info(f"Special cases file saved - path: {special_cases_input_path}")
            
            # Convert to CSV if needed
            special_cases_csv_path = ensure_csv_file(special_cases_input_path, f"/tmp/{job_id}_special_cases.csv")
            
            # Clean up original file if it was converted
            if special_cases_csv_path != special_cases_input_path and os.path.exists(special_cases_input_path):
                os.unlink(special_cases_input_path)
            
            logger.info(f"Special cases CSV ready - path: {special_cases_csv_path}")
        
        # Start background processing
        background_tasks.add_task(predict_insurance_codes_background, job_id, data_csv_path, special_cases_csv_path, enable_ai)
        
        logger.info(f"Background insurance prediction task started for job {job_id} with AI: {enable_ai}")
        
        return {"job_id": job_id, "message": "File uploaded and insurance code prediction started"}
        
    except Exception as e:
        logger.error(f"Insurance prediction upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

@app.get("/memory")
async def memory_status():
    """Memory usage and job status endpoint"""
    import psutil
    
    # Get memory info
    process = psutil.Process()
    memory_info = process.memory_info()
    
    # Count jobs by status
    job_counts = {}
    for job in job_status.values():
        status = job.status
        job_counts[status] = job_counts.get(status, 0) + 1
    
    # Count result files
    results_dir = Path("/tmp/results")
    result_files = list(results_dir.glob("*")) if results_dir.exists() else []
    total_result_size = sum(f.stat().st_size for f in result_files if f.is_file())
    
    return {
        "memory_mb": round(memory_info.rss / 1024 / 1024, 2),
        "memory_percent": round(process.memory_percent(), 2),
        "total_jobs": len(job_status),
        "job_counts": job_counts,
        "result_files_count": len(result_files),
        "result_files_size_mb": round(total_result_size / 1024 / 1024, 2),
        "gc_counts": gc.get_count()
    }

@app.post("/force-gc")
async def force_garbage_collection():
    """Force garbage collection to free memory"""
    import psutil
    
    # Get memory before
    process = psutil.Process()
    memory_before = process.memory_info().rss / 1024 / 1024
    
    # Force garbage collection
    collected = gc.collect()
    
    # Get memory after
    memory_after = process.memory_info().rss / 1024 / 1024
    memory_freed = memory_before - memory_after
    
    logger.info(f"Forced garbage collection: collected {collected} objects, freed {memory_freed:.2f} MB")
    
    return {
        "message": "Garbage collection completed",
        "objects_collected": collected,
        "memory_before_mb": round(memory_before, 2),
        "memory_after_mb": round(memory_after, 2),
        "memory_freed_mb": round(memory_freed, 2)
    }

# ============================================================================
# Modifiers Configuration API Endpoints
# ============================================================================

@app.get("/api/modifiers")
async def get_modifiers(page: int = 1, page_size: int = 50, search: str = None):
    """Get modifier configurations from database with pagination"""
    try:
        from db_utils import get_all_modifiers
        result = get_all_modifiers(page=page, page_size=page_size, search=search)
        return result
    except Exception as e:
        logger.error(f"Failed to get modifiers: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to retrieve modifiers: {str(e)}")

@app.get("/api/modifiers/{mednet_code}")
async def get_modifier(mednet_code: str):
    """Get a specific modifier configuration by MedNet code"""
    try:
        from db_utils import get_modifier as get_modifier_by_code
        modifier = get_modifier_by_code(mednet_code)
        if modifier:
            return modifier
        else:
            raise HTTPException(status_code=404, detail=f"Modifier {mednet_code} not found")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get modifier {mednet_code}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to retrieve modifier: {str(e)}")

@app.post("/api/modifiers")
async def create_or_update_modifier(
    mednet_code: str = Form(...),
    medicare_modifiers: bool = Form(...),
    bill_medical_direction: bool = Form(...),
    enable_qs: bool = Form(True)
):
    """Create or update a modifier configuration"""
    try:
        from db_utils import upsert_modifier
        success = upsert_modifier(mednet_code, medicare_modifiers, bill_medical_direction, enable_qs)
        if success:
            return {
                "message": "Modifier saved successfully",
                "mednet_code": mednet_code,
                "medicare_modifiers": medicare_modifiers,
                "bill_medical_direction": bill_medical_direction,
                "enable_qs": enable_qs
            }
        else:
            raise HTTPException(status_code=500, detail="Failed to save modifier")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to save modifier {mednet_code}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to save modifier: {str(e)}")

@app.put("/api/modifiers/{mednet_code}")
async def update_modifier(
    mednet_code: str,
    medicare_modifiers: bool = Form(...),
    bill_medical_direction: bool = Form(...),
    enable_qs: bool = Form(True)
):
    """Update an existing modifier configuration"""
    try:
        from db_utils import upsert_modifier
        success = upsert_modifier(mednet_code, medicare_modifiers, bill_medical_direction, enable_qs)
        if success:
            return {
                "message": "Modifier updated successfully",
                "mednet_code": mednet_code,
                "medicare_modifiers": medicare_modifiers,
                "bill_medical_direction": bill_medical_direction,
                "enable_qs": enable_qs
            }
        else:
            raise HTTPException(status_code=500, detail="Failed to update modifier")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update modifier {mednet_code}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to update modifier: {str(e)}")

@app.delete("/api/modifiers/{mednet_code}")
async def delete_modifier(mednet_code: str):
    """Delete a modifier configuration"""
    try:
        from db_utils import delete_modifier as delete_modifier_by_code
        success = delete_modifier_by_code(mednet_code)
        if success:
            return {"message": f"Modifier {mednet_code} deleted successfully"}
        else:
            raise HTTPException(status_code=500, detail="Failed to delete modifier")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete modifier {mednet_code}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to delete modifier: {str(e)}")


# ============================================================================
# Instruction Templates API Endpoints
# ============================================================================

@app.get("/api/templates")
async def get_templates(page: int = 1, page_size: int = 50, search: str = None):
    """Get instruction templates from database with pagination"""
    try:
        from db_utils import get_all_templates
        result = get_all_templates(page=page, page_size=page_size, search=search)
        return result
    except Exception as e:
        logger.error(f"Failed to get templates: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to retrieve templates: {str(e)}")


@app.get("/api/templates/{template_id}")
async def get_template(template_id: int):
    """Get a specific instruction template by ID"""
    try:
        from db_utils import get_template as get_template_by_id
        template = get_template_by_id(template_id=template_id)
        if template:
            return template
        else:
            raise HTTPException(status_code=404, detail=f"Template {template_id} not found")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get template {template_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to retrieve template: {str(e)}")


@app.get("/api/templates/by-name/{template_name}")
async def get_template_by_name(template_name: str):
    """Get a specific instruction template by name"""
    try:
        from db_utils import get_template
        template = get_template(template_name=template_name)
        if template:
            return template
        else:
            raise HTTPException(status_code=404, detail=f"Template '{template_name}' not found")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get template '{template_name}': {e}")
        raise HTTPException(status_code=500, detail=f"Failed to retrieve template: {str(e)}")


@app.post("/api/templates/upload")
async def upload_template(
    name: str = Form(...),
    description: str = Form(""),
    excel_file: UploadFile = File(...)
):
    """Upload an Excel file and save it as an instruction template"""
    import pandas as pd
    
    try:
        # Validate file type
        if not excel_file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")
        
        # Save uploaded file temporarily
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        try:
            content = await excel_file.read()
            temp_file.write(content)
            temp_file.close()
            
            # Parse the Excel file to extract field definitions
            sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'current'))
            from field_definitions import load_field_definitions_from_excel
            
            field_definitions = load_field_definitions_from_excel(temp_file.name)
            
            if not field_definitions:
                raise HTTPException(status_code=400, detail="Failed to parse Excel file or no fields found")
            
            # Store template in database
            from db_utils import create_template
            template_id = create_template(
                name=name,
                description=description,
                template_data={'fields': field_definitions}
            )
            
            if template_id:
                return {
                    "message": "Template uploaded successfully",
                    "template_id": template_id,
                    "name": name,
                    "fields_count": len(field_definitions)
                }
            else:
                raise HTTPException(status_code=500, detail="Failed to save template to database")
        
        finally:
            # Clean up temporary file
            try:
                os.unlink(temp_file.name)
            except:
                pass
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to upload template: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to upload template: {str(e)}")


@app.put("/api/templates/{template_id}")
async def update_template(
    template_id: int,
    name: str = Form(None),
    description: str = Form(None),
    excel_file: UploadFile = File(None)
):
    """Update an existing instruction template"""
    import pandas as pd
    
    try:
        # Check if template exists
        from db_utils import get_template, update_template as update_template_in_db
        existing_template = get_template(template_id=template_id)
        if not existing_template:
            raise HTTPException(status_code=404, detail=f"Template {template_id} not found")
        
        # Prepare update parameters
        template_data = None
        
        # If new Excel file is provided, parse it
        if excel_file and excel_file.filename:
            if not excel_file.filename.endswith(('.xlsx', '.xls')):
                raise HTTPException(status_code=400, detail="Only Excel files (.xlsx, .xls) are supported")
            
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            try:
                content = await excel_file.read()
                temp_file.write(content)
                temp_file.close()
                
                # Parse the Excel file
                sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'current'))
                from field_definitions import load_field_definitions_from_excel
                
                field_definitions = load_field_definitions_from_excel(temp_file.name)
                
                if not field_definitions:
                    raise HTTPException(status_code=400, detail="Failed to parse Excel file or no fields found")
                
                template_data = {'fields': field_definitions}
            
            finally:
                try:
                    os.unlink(temp_file.name)
                except:
                    pass
        
        # Update template in database
        success = update_template_in_db(
            template_id=template_id,
            name=name,
            description=description,
            template_data=template_data
        )
        
        if success:
            # Get updated template
            updated_template = get_template(template_id=template_id)
            return {
                "message": "Template updated successfully",
                "template": updated_template
            }
        else:
            raise HTTPException(status_code=500, detail="Failed to update template")
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update template {template_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to update template: {str(e)}")


@app.put("/api/templates/{template_id}/fields")
async def update_template_fields(template_id: int, request: dict):
    """Update template fields directly via JSON"""
    try:
        from db_utils import get_template, update_template as update_template_in_db
        
        # Check if template exists
        existing_template = get_template(template_id=template_id)
        if not existing_template:
            raise HTTPException(status_code=404, detail=f"Template {template_id} not found")
        
        # Get template_data from request body
        template_data = request.get('template_data')
        if not template_data:
            raise HTTPException(status_code=400, detail="template_data is required")
        
        # Validate that fields exist
        fields = template_data.get('fields', [])
        if not fields:
            raise HTTPException(status_code=400, detail="At least one field is required")
        
        # Validate that all fields have names
        for field in fields:
            if not field.get('name'):
                raise HTTPException(status_code=400, detail="All fields must have a name")
        
        # Update template in database
        success = update_template_in_db(
            template_id=template_id,
            template_data=template_data
        )
        
        if success:
            # Get updated template
            updated_template = get_template(template_id=template_id)
            return {
                "message": "Template fields updated successfully",
                "template": updated_template
            }
        else:
            raise HTTPException(status_code=500, detail="Failed to update template")
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update template fields {template_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to update template fields: {str(e)}")


@app.delete("/api/templates/{template_id}")
async def delete_template(template_id: int):
    """Delete an instruction template"""
    try:
        from db_utils import delete_template as delete_template_by_id
        success = delete_template_by_id(template_id)
        if success:
            return {"message": f"Template {template_id} deleted successfully"}
        else:
            raise HTTPException(status_code=500, detail="Failed to delete template")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete template {template_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to delete template: {str(e)}")


@app.post("/api/templates/{template_id}/add-charge-fields")
async def add_charge_fields_to_template(
    template_id: int,
    location_value: str = Form(...),
    surgeon_list_text: str = Form(...),
    provider_list_text: str = Form(...)
):
    """
    Add charge fields to an existing template in bulk.
    
    Args:
        template_id: ID of the template to update
        location_value: Value for the Location field (e.g., "GAP-UMCS")
        surgeon_list_text: Formatted surgeon list text from surgeon mapping
        provider_list_text: Formatted provider list text from provider mapping
    """
    try:
        from db_utils import get_template, update_template as update_template_in_db
        
        # Check if template exists
        existing_template = get_template(template_id=template_id)
        if not existing_template:
            raise HTTPException(status_code=404, detail=f"Template {template_id} not found")
        
        # Parse surgeon list (simple list of surgeons, one per line)
        surgeon_lines = surgeon_list_text.strip().split('\n')
        surgeon_list = [line.strip() for line in surgeon_lines if line.strip()]
        surgeon_list_str = "\n".join(surgeon_list) if surgeon_list else "No surgeons"
        
        # Parse provider list to separate CRNA and MD
        lines = provider_list_text.strip().split('\n')
        crna_providers = []
        md_providers = []
        all_providers = []
        
        current_section = None
        for line in lines:
            line = line.strip()
            if not line:
                continue
            if "Billable CRNA providers:" in line:
                current_section = "crna"
                continue
            elif "Billable MD providers:" in line:
                current_section = "md"
                continue
            
            if current_section == "crna":
                crna_providers.append(line)
                all_providers.append(line)
            elif current_section == "md":
                md_providers.append(line)
                all_providers.append(line)
        
        # Format provider lists for instructions
        crna_list_str = "\n".join(crna_providers) if crna_providers else "No CRNA providers"
        md_list_str = "\n".join(md_providers) if md_providers else "No MD providers"
        all_providers_str = "\n".join(all_providers) if all_providers else "No providers"
        
        # Define all charge fields
        charge_fields = [
            {
                'name': 'Location',
                'description': f'In this field always output value "{location_value}", exactly that never anything else',
                'location': 'N/A - Fixed value',
                'output_format': 'String'
            },
            {
                'name': 'PatientClass',
                'description': '''PatientClass Extraction Rule

determine if the patient is an inpatient or an outpatient (likely listed somewhere in the record)

Only these two values are allowed: INPATIENT or OUTPATIENT

Do not output any other value!''',
                'location': 'Patient record - patient class field',
                'output_format': 'INPATIENT or OUTPATIENT'
            },
            {
                'name': 'Surgeon',
                'description': f'''AI INSTRUCTION — SURGEON (MD) PROVIDER EXTRACTION RULE

SECTION TO READ

ANESTHESIA RECORD → SURGEON

The AI must extract only ONE surgeon (MD) using the rules below.

1️⃣ PRIMARY RULE — Identify and Extract the Surgeon

If a surgeon is listed anywhere in the record, the AI must:

✔ Identify the Surgeon / MD
✔ Extract the name
✔ Standardize it using the reference list

If only one surgeon MD is listed → Extract that surgeon MD.

2️⃣ MULTIPLE SURGEONS (MD) LISTED

If more than one MD appears:

✔ Extract ONLY ONE
✔ Select the surgeon whose name appears as the primary surgeon, "Attending Surgeon," or the first listed surgeon.

3️⃣ NAME STANDARDIZATION REQUIREMENTS

The raw MD name in the record may contain:

Spelling errors
Missing middle initial
Extra or missing spaces
Abbreviations
Only ID format (ex: GOSSAGE-13379)
Incorrect punctuation
Missing credential

👉 The AI MUST replace the raw name with the EXACT standardized name from the MD reference list.

4️⃣ OUTPUT FORMAT (MANDATORY)

The extracted MD must ALWAYS follow this structure:

{{LAST NAME}}, {{FIRST NAME}} {{MIDDLE INITIAL}}, MD

5️⃣ CREDENTIAL RULE

All extracted surgeon providers must be labeled as:

MD

Do NOT change the credential to DO, MBBS, etc., even if raw text contains these.
Only use MD, as per the standardized provider list.

6️⃣ REFERENCE LIST (STANDARDIZED PROVIDER NAMES)

The AI must match the extracted surgeon to the following approved list:

MD Providers (Standardized)
{surgeon_list_str}

7️⃣ IF THE MD NAME IS NOT FOUND IN THE STANDARDIZED LIST

If the extracted surgeon does not match any name in the reference list:

👉 Leave the cell BLANK.

✔ FINAL OUTPUT RULE

Output format: LAST NAME, FIRST NAME MIDDLE INITIAL, MD''',
                'location': 'Anesthesia record - surgeon field',
                'output_format': 'LAST NAME, FIRST NAME MIDDLE INITIAL, MD'
            },
            {
                'name': 'Referring',
                'description': 'COPY VALUE FROM surgeon field and put here also, exactly same output value (needed for our integration system to work)',
                'location': 'Copy from Surgeon field',
                'output_format': 'Same as Surgeon field'
            },
            {
                'name': 'Responsible Provider',
                'description': f'''✅ AI INSTRUCTION — PROVIDER (MD or CRNA) EXTRACTION RULE

SECTION TO READ

1️⃣ ANESTHESIA RECORD MULTIPLE CRNAs or multiple MD LISTED — SELECT THE CRNA or MD WITH LONGEST DURATION

MD takes priority for being in Responsible Provider field over CRNA though, but ONLY if he was actually properly on the case participating, if the MD only SIGNED the record and nothing else but the CRNA did the work, then CRNA is responsible provider

When the anesthesia record contains more than one CRNA or MD, the AI must:
Step A — Identify each CRNA or MD by ID

Step B — Calculate total anesthesia time

Duration = End Time – Start Time

Step C — Select the CRNA or MD with longest time

This CRNA or MD becomes the performing provider.

2️⃣ANESTHESIA RECORD → CRNA Entries

if there is no multiple CRNA Then the AI must extract only ONE CRNA provider using the rules below.
PRIMARY RULE — Extract the CRNA Listed in the Anesthesia Record

The AI must always map ID or raw text to the standardized provider list (reference list given below).

3️⃣ NAME STANDARDIZATION — MUST MATCH REFERENCE LIST EXACTLY

The raw provider name in the record may have:

Spelling errors
Abbreviations
Extra text
Missing credentials
Only an ID
Partial names

👉 The AI must always replace the raw name with the exact standardized version from the reference list.

4️⃣ OUTPUT FORMAT (MANDATORY)

The provider must ALWAYS be output in this format:

{{LAST NAME}}, {{FIRST NAME}} {{MIDDLE INITIAL}}, CRNA
or
{{LAST NAME}}, {{FIRST NAME}} {{MIDDLE INITIAL}}, MD

if there is an MD or CRNA that is NOT on our reference list of our providers it CANNOT go in this or any other field, because they are not providers that we can bill for

REFERENCE LIST (STANDARDIZED PROVIDER NAMES):
{all_providers_str}''',
                'location': 'Anesthesia record - provider with longest duration',
                'output_format': 'LAST NAME, FIRST NAME MIDDLE INITIAL, CRNA or MD'
            },
            {
                'name': 'MD',
                'description': f'''if there is an MD from our standardized list of providers on the case, then put it in this field, if there are multiple pick the one with highest time on case, the list of billable providers is:

{md_list_str}''',
                'location': 'Anesthesia record - MD provider',
                'output_format': 'LAST NAME, FIRST NAME MIDDLE INITIAL, MD'
            },
            {
                'name': 'CRNA',
                'description': f'''if there is a CRNA from our standardized list of providers on the case, then put it in this field, if there are multiple pick the one with highest time on case, the list of billable providers is:

{crna_list_str}''',
                'location': 'Anesthesia record - CRNA provider',
                'output_format': 'LAST NAME, FIRST NAME MIDDLE INITIAL, CRNA'
            },
            {
                'name': 'SRNA',
                'description': "if there is a SRNA present on the case output in this field just text 'SRNA'",
                'location': 'Anesthesia record - SRNA field',
                'output_format': 'SRNA or empty'
            },
            {
                'name': 'Resident',
                'description': "if there is a RESIDENT present on the case output in this field just text 'RES'",
                'location': 'Anesthesia record - Resident field',
                'output_format': 'RES or empty'
            },
            {
                'name': 'Revenue',
                'description': 'if the case is a labor case then put OB otherwise always OR',
                'location': 'Determine from case type',
                'output_format': 'OR or OB'
            },
            {
                'name': 'PlaceOfService',
                'description': 'always leave empty string',
                'location': 'N/A',
                'output_format': 'Empty string'
            },
            {
                'name': 'TypeOfService',
                'description': 'always leave empty string',
                'location': 'N/A',
                'output_format': 'Empty string'
            },
            {
                'name': 'AnStart',
                'description': '''AI Instruction: Extract Anesthesia Start Time (With DOS + Required Format)

1️⃣ Primary Capture Rule — Extract Anesthesia Start Time
Look for fields labeled exactly or similarly as:
Anes Start or an start or anything indicating the start of anesthesia

If found Anes Start, extract this value EXACTLY as recorded.

2️⃣ Secondary Capture Rule — Use Anesthesia Ready Time ONLY if Start Time is Missing
If Anesthesia Start Time is completely missing, then search for:
Anes Ready
Extract Ready Time ONLY when Start Time is not available.

3️⃣ Required Output Format (MANDATORY)
The output must ALWAYS include:

Date of Service (DOS)
Time
Corrected, unified datetime format

Format to output:
Example:
9/30/2025  9:53:00 AM

4️⃣ Automatically Correct Inconsistent or Incomplete Formats

The AI must auto-correct formatting issues such as:

Missing seconds → add :00
Missing AM/PM → infer using recorded format (if ambiguous, use document standard)
Wrong separator (e.g., "." or "-") → convert to /
24-hour time → convert to AM/PM format
Missing leading zeros → correct to proper format

Example Corrections
Raw Value Found	Correct Output
09:22	9/30/2025 9:22:00 AM
14:05	9/30/2025 2:05:00 PM
9.22 AM	9/30/2025 9:22:00 AM

5️⃣ Final Output Rule

Always output:

<DOS + Time>

Example (Start Time Found)
9/30/2025  9:53:00 AM

Example (Start Time Missing → Ready Time Used)
9/30/2025  9:53:00 AM (Ready Time Used)''',
                'location': 'Anesthesia record - Anes Start or Anes Ready',
                'output_format': 'MM/DD/YYYY HH:MM:SS AM/PM'
            },
            {
                'name': 'AnStop',
                'description': '''AI Instruction: Extract Anesthesia End Time (With DOS + Required Format)

1️⃣ Primary Capture Rule — Extract Anesthesia End Time

Look for fields labeled exactly or similarly as:
Anes End, anesthesia end, anything indicating when the anesthesia stopped
👉 If found, extract the time as recorded.

3️⃣ Required Output Format (MANDATORY)

The output must always include:

Date of Service (DOS)
Time in correct unified format

Final required format:
MM/DD/YYYY HH:MM:SS AM/PM

Example:
9/30/2025 9:22:00 AM

4️⃣ Automatically Correct Time Formats

The AI must auto-correct inconsistencies such as:

Missing seconds → add :00
Missing AM/PM → infer from 24-hour or document standard
24-hour format → convert to AM/PM
Missing leading zeros → correct (9:2 → 9:02)
Wrong separators → convert to / or : appropriately
If only time is present, append the DOS

Examples of Corrections
Raw Value Found	Output
15:10	9/30/2025 3:10:00 PM
3.10 PM	9/30/2025 3:10:00 PM
03:10	9/30/2025 3:10:00 AM (based on context)
3:10	9/30/2025 3:10:00 AM

5️⃣ Final Output Rule

Always output:

<DOS + Time>

Example (End Time Found)
9/30/2025  9:53:00 AM''',
                'location': 'Anesthesia record - Anes End',
                'output_format': 'MM/DD/YYYY HH:MM:SS AM/PM'
            },
            {
                'name': 'Anesthesia Type',
                'description': 'output values can be GENERAL, MAC, REGIONAL, TIVA nothing else dont output any other value ever, whatever is listed in the record you should think like a medical coder, think like what a medical coder would put of one of these 4 standardized values in order to get the case PAID',
                'location': 'Anesthesia record - anesthesia type',
                'output_format': 'GENERAL, MAC, REGIONAL, or TIVA'
            },
            {
                'name': 'PhysicalStatus',
                'description': 'allowed output values are : 1, 2, 3, 4, 5, 6 , find the physical status, p status of the patient (medical billing term) and extract it as a number EXACTLY one of these numbers NO other value',
                'location': 'Anesthesia record - physical status',
                'output_format': '1, 2, 3, 4, 5, or 6'
            }
        ]
        
        # Get existing fields from template
        existing_fields = existing_template['template_data'].get('fields', [])
        
        # Get list of charge field names to check for duplicates
        charge_field_names = {field['name'] for field in charge_fields}
        
        # Find existing fields that will be overridden
        existing_field_names = {field.get('name') for field in existing_fields if field.get('name')}
        overridden_fields = existing_field_names & charge_field_names
        
        # Remove existing fields that have the same name as charge fields (override them)
        filtered_existing_fields = [
            field for field in existing_fields 
            if field.get('name') not in charge_field_names
        ]
        
        # Add charge fields to filtered existing fields (this will override any duplicates)
        updated_fields = filtered_existing_fields + charge_fields
        
        # Update template with new fields
        template_data = {'fields': updated_fields}
        success = update_template_in_db(
            template_id=template_id,
            template_data=template_data
        )
        
        if success:
            updated_template = get_template(template_id=template_id)
            override_count = len(overridden_fields)
            new_count = len(charge_fields) - override_count
            
            if override_count > 0:
                message = f"Successfully processed {len(charge_fields)} charge fields: {new_count} added, {override_count} overridden"
            else:
                message = f"Successfully added {len(charge_fields)} charge fields to template"
            
            return {
                "message": message,
                "template": updated_template,
                "fields_added": len(charge_fields),
                "fields_overridden": override_count,
                "fields_new": new_count
            }
        else:
            raise HTTPException(status_code=500, detail="Failed to update template with charge fields")
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to add charge fields to template {template_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to add charge fields: {str(e)}")


@app.post("/api/templates/{template_id}/add-peripheral-blocks-field")
async def add_peripheral_blocks_field_to_template(template_id: int):
    """
    Add peripheral_blocks field to an existing template.
    Reads instructions from peripheral_blocks.txt file in the backend directory.
    If a field with name 'peripheral_blocks' already exists, it will be overridden.
    """
    try:
        from db_utils import get_template, update_template as update_template_in_db
        from pathlib import Path
        
        # Check if template exists
        existing_template = get_template(template_id=template_id)
        if not existing_template:
            raise HTTPException(status_code=404, detail=f"Template {template_id} not found")
        
        # Read peripheral_blocks.txt file
        backend_dir = Path(__file__).parent
        peripheral_blocks_file = backend_dir / "peripheral_blocks.txt"
        
        if not peripheral_blocks_file.exists():
            raise HTTPException(
                status_code=404, 
                detail=f"peripheral_blocks.txt not found at {peripheral_blocks_file}. Please create this file with the field description."
            )
        
        # Read the file content
        try:
            with open(peripheral_blocks_file, 'r', encoding='utf-8') as f:
                peripheral_blocks_description = f.read().strip()
        except Exception as e:
            raise HTTPException(
                status_code=500,
                detail=f"Failed to read peripheral_blocks.txt: {str(e)}"
            )
        
        if not peripheral_blocks_description:
            raise HTTPException(
                status_code=400,
                detail="peripheral_blocks.txt is empty. Please add the field description."
            )
        
        # Define the peripheral_blocks field
        peripheral_blocks_field = {
            'name': 'peripheral_blocks',
            'description': peripheral_blocks_description,
            'location': 'N/A - See description',
            'output_format': 'See description'
        }
        
        # Get existing fields from template
        existing_fields = existing_template['template_data'].get('fields', [])
        
        # Check if peripheral_blocks field already exists
        field_exists = any(
            field.get('name') == 'peripheral_blocks' 
            for field in existing_fields
        )
        
        # Remove existing peripheral_blocks field if it exists (override)
        filtered_existing_fields = [
            field for field in existing_fields 
            if field.get('name') != 'peripheral_blocks'
        ]
        
        # Add the peripheral_blocks field
        updated_fields = filtered_existing_fields + [peripheral_blocks_field]
        
        # Update template with new field
        template_data = {'fields': updated_fields}
        success = update_template_in_db(
            template_id=template_id,
            template_data=template_data
        )
        
        if success:
            updated_template = get_template(template_id=template_id)
            
            if field_exists:
                message = "Successfully updated peripheral_blocks field in template (overridden existing field)"
            else:
                message = "Successfully added peripheral_blocks field to template"
            
            return {
                "message": message,
                "template": updated_template,
                "field_overridden": field_exists
            }
        else:
            raise HTTPException(status_code=500, detail="Failed to update template with peripheral_blocks field")
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to add peripheral_blocks field to template {template_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to add peripheral_blocks field: {str(e)}")


@app.post("/api/templates/{template_id}/add-colonoscopy-fields")
async def add_colonoscopy_fields_to_template(template_id: int):
    """
    Add colonoscopy-related fields to an existing template.
    Adds/overrides three fields: is_colonoscopy, colonoscopy_is_screening, and Polyps found.
    """
    try:
        from db_utils import get_template, update_template as update_template_in_db
        
        # Check if template exists
        existing_template = get_template(template_id=template_id)
        if not existing_template:
            raise HTTPException(status_code=404, detail=f"Template {template_id} not found")
        
        # Define the colonoscopy fields
        colonoscopy_fields = [
            {
                "name": "is_colonoscopy",
                "location": "",
                "priority": False,
                "description": "in this field please indicate if the procedure was a colonoscopy (TRUE) or anything else (FALSE)",
                "output_format": ""
            },
            {
                "name": "colonoscopy_is_screening",
                "location": "",
                "priority": False,
                "description": "here indicate if the colonoscopy description includes the word screening, it it was a screening colonoscopy put here TRUE otherwise put FALSE (if it was just standard diagnosit colonoscopy OR it was any other procedure), maybe it doesnt say in the pdf it was screening explicitly but if it says it is the patients first time also then indicate screening with TRUE, SOMETIMES they mention that it was screening in lower pages of the pdf also... check thouroughly",
                "output_format": ""
            },
            {
                "name": "Polyps found",
                "location": "",
                "priority": False,
                "description": "IF the procedure was *colonoscopy* AND somewhere in the pdf they indicated that they FOUND polyps, then put \"FOUND\" here, otherwise leave this field EMPTY just \"\"",
                "output_format": ""
            }
        ]
        
        # Get existing fields from template
        existing_fields = existing_template['template_data'].get('fields', [])
        
        # Get list of colonoscopy field names to check for duplicates
        colonoscopy_field_names = {field['name'] for field in colonoscopy_fields}
        
        # Find existing fields that will be overridden
        existing_field_names = {field.get('name') for field in existing_fields if field.get('name')}
        overridden_fields = existing_field_names & colonoscopy_field_names
        
        # Remove existing fields that have the same name as colonoscopy fields (override them)
        filtered_existing_fields = [
            field for field in existing_fields 
            if field.get('name') not in colonoscopy_field_names
        ]
        
        # Add colonoscopy fields to filtered existing fields (this will override any duplicates)
        updated_fields = filtered_existing_fields + colonoscopy_fields
        
        # Update template with new fields
        template_data = {'fields': updated_fields}
        success = update_template_in_db(
            template_id=template_id,
            template_data=template_data
        )
        
        if success:
            updated_template = get_template(template_id=template_id)
            override_count = len(overridden_fields)
            new_count = len(colonoscopy_fields) - override_count
            
            if override_count > 0:
                message = f"Successfully processed {len(colonoscopy_fields)} colonoscopy fields: {new_count} added, {override_count} overridden"
            else:
                message = f"Successfully added {len(colonoscopy_fields)} colonoscopy fields to template"
            
            return {
                "message": message,
                "template": updated_template,
                "fields_added": len(colonoscopy_fields),
                "fields_overridden": override_count,
                "fields_new": new_count
            }
        else:
            raise HTTPException(status_code=500, detail="Failed to update template with colonoscopy fields")
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to add colonoscopy fields to template {template_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to add colonoscopy fields: {str(e)}")


@app.post("/api/templates/{template_id}/export")
async def export_template_as_excel(template_id: int):
    """Export a template back to Excel format for download"""
    import pandas as pd
    
    try:
        from db_utils import get_template
        template = get_template(template_id=template_id)
        
        if not template:
            raise HTTPException(status_code=404, detail=f"Template {template_id} not found")
        
        # Extract field definitions from template
        fields = template['template_data'].get('fields', [])
        
        if not fields:
            raise HTTPException(status_code=400, detail="Template has no field definitions")
        
        # Create DataFrame in the expected format
        # Columns are field names, rows are description, location, output_format, priority
        data = {}
        for field in fields:
            data[field['name']] = [
                field.get('description', ''),
                field.get('location', ''),
                field.get('output_format', ''),
                'YES' if field.get('priority', False) else 'NO'
            ]
        
        df = pd.DataFrame(data)
        
        # Create temporary Excel file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        df.to_excel(temp_file.name, index=False, engine='openpyxl')
        temp_file.close()
        
        # Return file
        filename = f"{template['name'].replace(' ', '_')}.xlsx"
        return FileResponse(
            path=temp_file.name,
            filename=filename,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            background=BackgroundTask(lambda: os.unlink(temp_file.name))
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to export template {template_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to export template: {str(e)}")


@app.get("/api/templates/{template_id}/download-json")
async def download_template_json(template_id: int):
    """Download a template as a JSON file"""
    import json
    
    try:
        from db_utils import get_template
        template = get_template(template_id=template_id)
        
        if not template:
            raise HTTPException(status_code=404, detail=f"Template {template_id} not found")
        
        # Prepare the template data for download
        # Convert datetime objects to ISO format strings
        download_data = {
            'id': template['id'],
            'name': template['name'],
            'description': template.get('description', ''),
            'template_data': template['template_data'],
            'created_at': template['created_at'].isoformat() if template.get('created_at') else None,
            'updated_at': template['updated_at'].isoformat() if template.get('updated_at') else None
        }
        
        # Create temporary JSON file
        temp_file = tempfile.NamedTemporaryFile(mode='w', delete=False, suffix='.json', encoding='utf-8')
        json.dump(download_data, temp_file, indent=2, ensure_ascii=False)
        temp_file.close()
        
        # Return file
        filename = f"{template['name'].replace(' ', '_')}_template.json"
        return FileResponse(
            path=temp_file.name,
            filename=filename,
            media_type="application/json",
            background=BackgroundTask(lambda: os.unlink(temp_file.name))
        )
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to download template {template_id} as JSON: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to download template: {str(e)}")


@app.post("/api/templates/{template_id}/upload-json")
async def upload_template_json(
    template_id: int,
    json_file: UploadFile = File(...)
):
    """Upload a JSON file to update a template"""
    import json
    
    try:
        from db_utils import get_template, update_template as update_template_in_db
        
        # Check if template exists
        existing_template = get_template(template_id=template_id)
        if not existing_template:
            raise HTTPException(status_code=404, detail=f"Template {template_id} not found")
        
        # Validate file type
        if not json_file.filename.endswith('.json'):
            raise HTTPException(status_code=400, detail="Only JSON files (.json) are supported")
        
        # Read and parse JSON file
        content = await json_file.read()
        try:
            uploaded_data = json.loads(content.decode('utf-8'))
        except json.JSONDecodeError as e:
            raise HTTPException(status_code=400, detail=f"Invalid JSON file: {str(e)}")
        
        # Extract template_data
        template_data = uploaded_data.get('template_data')
        if not template_data:
            raise HTTPException(status_code=400, detail="JSON file must contain 'template_data' field")
        
        # Validate that fields exist
        fields = template_data.get('fields', [])
        if not fields:
            raise HTTPException(status_code=400, detail="template_data must contain at least one field")
        
        # Validate that all fields have names
        for idx, field in enumerate(fields):
            if not field.get('name'):
                raise HTTPException(status_code=400, detail=f"Field at index {idx} is missing 'name'")
        
        # Optional: Also update name and description if provided in JSON
        name = uploaded_data.get('name')
        description = uploaded_data.get('description')
        
        # Update template in database
        success = update_template_in_db(
            template_id=template_id,
            name=name if name and name != existing_template['name'] else None,
            description=description if description is not None else None,
            template_data=template_data
        )
        
        if success:
            # Get updated template
            updated_template = get_template(template_id=template_id)
            return {
                "message": "Template updated successfully from JSON",
                "template": updated_template,
                "fields_count": len(fields)
            }
        else:
            raise HTTPException(status_code=500, detail="Failed to update template")
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to upload JSON for template {template_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to upload JSON: {str(e)}")


# ============================================================================
# Prediction Instructions API Endpoints (CPT/ICD)
# ============================================================================

@app.get("/api/prediction-instructions")
async def get_prediction_instructions(
    instruction_type: str = None,
    page: int = 1,
    page_size: int = 50,
    search: str = None
):
    """Get prediction instruction templates (CPT/ICD) with pagination"""
    try:
        from db_utils import get_all_prediction_instructions
        result = get_all_prediction_instructions(
            instruction_type=instruction_type,
            page=page,
            page_size=page_size,
            search=search
        )
        return result
    except Exception as e:
        logger.error(f"Failed to get prediction instructions: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to retrieve instructions: {str(e)}")


@app.get("/api/prediction-instructions/{instruction_id}")
async def get_prediction_instruction(instruction_id: int):
    """Get a specific prediction instruction by ID"""
    try:
        from db_utils import get_prediction_instruction as get_instruction_by_id
        instruction = get_instruction_by_id(instruction_id=instruction_id)
        if instruction:
            return instruction
        else:
            raise HTTPException(status_code=404, detail=f"Instruction {instruction_id} not found")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get prediction instruction {instruction_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to retrieve instruction: {str(e)}")


@app.post("/api/prediction-instructions")
async def create_prediction_instruction(request: dict):
    """Create a new prediction instruction template"""
    try:
        name = request.get("name")
        instruction_type = request.get("instruction_type")
        instructions_text = request.get("instructions_text")
        description = request.get("description", "")
        
        if not name or not instruction_type or not instructions_text:
            raise HTTPException(
                status_code=400,
                detail="name, instruction_type, and instructions_text are required"
            )
        
        if instruction_type not in ['cpt', 'icd']:
            raise HTTPException(
                status_code=400,
                detail="instruction_type must be 'cpt' or 'icd'"
            )
        
        from db_utils import create_prediction_instruction
        instruction_id = create_prediction_instruction(
            name=name,
            instruction_type=instruction_type,
            instructions_text=instructions_text,
            description=description
        )
        
        if instruction_id:
            return {
                "message": "Instruction template created successfully",
                "instruction_id": instruction_id,
                "name": name,
                "instruction_type": instruction_type
            }
        else:
            raise HTTPException(status_code=500, detail="Failed to create instruction template")
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to create prediction instruction: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to create instruction: {str(e)}")


@app.put("/api/prediction-instructions/{instruction_id}")
async def update_prediction_instruction(instruction_id: int, request: dict):
    """Update an existing prediction instruction template"""
    try:
        from db_utils import get_prediction_instruction, update_prediction_instruction as update_instruction_in_db
        
        # Check if instruction exists
        existing = get_prediction_instruction(instruction_id=instruction_id)
        if not existing:
            raise HTTPException(status_code=404, detail=f"Instruction {instruction_id} not found")
        
        name = request.get("name")
        description = request.get("description")
        instructions_text = request.get("instructions_text")
        
        success = update_instruction_in_db(
            instruction_id=instruction_id,
            name=name,
            description=description,
            instructions_text=instructions_text
        )
        
        if success:
            updated = get_prediction_instruction(instruction_id=instruction_id)
            return {
                "message": "Instruction template updated successfully",
                "instruction": updated
            }
        else:
            raise HTTPException(status_code=500, detail="Failed to update instruction template")
    
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update prediction instruction {instruction_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to update instruction: {str(e)}")


@app.delete("/api/prediction-instructions/{instruction_id}")
async def delete_prediction_instruction(instruction_id: int):
    """Delete a prediction instruction template"""
    try:
        from db_utils import delete_prediction_instruction as delete_instruction_by_id
        success = delete_instruction_by_id(instruction_id)
        if success:
            return {"message": f"Instruction {instruction_id} deleted successfully"}
        else:
            raise HTTPException(status_code=500, detail="Failed to delete instruction template")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete prediction instruction {instruction_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to delete instruction: {str(e)}")


# ============================================================================
# Insurance Mappings API Endpoints
# ============================================================================

@app.get("/api/insurance-mappings")
async def get_insurance_mappings(page: int = 1, page_size: int = 50, search: str = None):
    """Get insurance mappings from database with pagination"""
    try:
        from db_utils import get_all_insurance_mappings
        result = get_all_insurance_mappings(page=page, page_size=page_size, search=search)
        return result
    except Exception as e:
        logger.error(f"Failed to get insurance mappings: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to retrieve mappings: {str(e)}")


@app.get("/api/insurance-mappings/{mapping_id}")
async def get_insurance_mapping(mapping_id: int):
    """Get a specific insurance mapping by ID"""
    try:
        from db_utils import get_insurance_mapping as get_mapping_by_id
        mapping = get_mapping_by_id(mapping_id=mapping_id)
        if mapping:
            return mapping
        else:
            raise HTTPException(status_code=404, detail=f"Mapping {mapping_id} not found")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get insurance mapping {mapping_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to retrieve mapping: {str(e)}")


@app.post("/api/insurance-mappings")
async def create_or_update_insurance_mapping(
    input_code: str = Form(...),
    output_code: str = Form(...),
    description: str = Form(default="")
):
    """Create or update an insurance mapping"""
    try:
        from db_utils import upsert_insurance_mapping
        success = upsert_insurance_mapping(input_code, output_code, description)
        if success:
            return {
                "message": "Insurance mapping saved successfully",
                "input_code": input_code,
                "output_code": output_code,
                "description": description
            }
        else:
            raise HTTPException(status_code=500, detail="Failed to save insurance mapping")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to save insurance mapping {input_code}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to save mapping: {str(e)}")


@app.put("/api/insurance-mappings/{mapping_id}")
async def update_insurance_mapping(
    mapping_id: int,
    input_code: str = Form(...),
    output_code: str = Form(...),
    description: str = Form(default="")
):
    """Update an existing insurance mapping"""
    try:
        from db_utils import upsert_insurance_mapping
        success = upsert_insurance_mapping(input_code, output_code, description)
        if success:
            return {
                "message": "Insurance mapping updated successfully",
                "input_code": input_code,
                "output_code": output_code,
                "description": description
            }
        else:
            raise HTTPException(status_code=500, detail="Failed to update insurance mapping")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update insurance mapping {mapping_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to update mapping: {str(e)}")


@app.delete("/api/insurance-mappings/{mapping_id}")
async def delete_insurance_mapping(mapping_id: int):
    """Delete an insurance mapping"""
    try:
        from db_utils import delete_insurance_mapping as delete_mapping_by_id
        success = delete_mapping_by_id(mapping_id)
        if success:
            return {"message": f"Insurance mapping {mapping_id} deleted successfully"}
        else:
            raise HTTPException(status_code=500, detail="Failed to delete insurance mapping")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete insurance mapping {mapping_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to delete mapping: {str(e)}")


@app.get("/api/special-cases-templates")
async def get_special_cases_templates(page: int = 1, page_size: int = 50, search: str = None):
    """Get all special cases templates with pagination and search"""
    try:
        from db_utils import get_all_special_cases_templates
        result = get_all_special_cases_templates(page=page, page_size=page_size, search=search)
        return result
    except Exception as e:
        logger.error(f"Failed to get special cases templates: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to get templates: {str(e)}")


@app.get("/api/special-cases-templates/{template_id}")
async def get_special_cases_template(template_id: int):
    """Get a specific special cases template by ID"""
    try:
        from db_utils import get_special_cases_template as get_template
        template = get_template(template_id=template_id)
        if template:
            return template
        else:
            raise HTTPException(status_code=404, detail="Template not found")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get special cases template {template_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to get template: {str(e)}")


@app.post("/api/special-cases-templates/upload")
async def upload_special_cases_template(
    csv_file: UploadFile = File(...),
    name: str = Form(...),
    description: str = Form(default="")
):
    """
    Upload a CSV file and save as a special cases template.
    CSV format: Company name,Mednet code
    """
    try:
        from db_utils import create_special_cases_template
        import pandas as pd
        import io
        
        # Read CSV file
        contents = await csv_file.read()
        
        # Try different encodings
        encodings = ['utf-8', 'utf-8-sig', 'latin-1', 'iso-8859-1', 'cp1252']
        df = None
        for encoding in encodings:
            try:
                df = pd.read_csv(io.BytesIO(contents), dtype=str, encoding=encoding)
                break
            except UnicodeDecodeError:
                continue
        
        if df is None:
            raise HTTPException(status_code=400, detail="Could not read CSV file with any standard encoding")
        
        # Validate columns
        if 'Company name' not in df.columns or 'Mednet code' not in df.columns:
            raise HTTPException(
                status_code=400,
                detail="CSV must have 'Company name' and 'Mednet code' columns"
            )
        
        # Parse mappings
        mappings = []
        for _, row in df.iterrows():
            company_name = str(row['Company name']).strip() if pd.notna(row['Company name']) else ''
            mednet_code = str(row['Mednet code']).strip() if pd.notna(row['Mednet code']) else ''
            
            if company_name and mednet_code:
                mappings.append({
                    'company_name': company_name,
                    'mednet_code': mednet_code
                })
        
        if not mappings:
            raise HTTPException(status_code=400, detail="No valid mappings found in CSV")
        
        # Create template
        template_id = create_special_cases_template(name, mappings, description)
        
        if template_id:
            return {
                "message": "Special cases template created successfully",
                "template_id": template_id,
                "name": name,
                "mappings_count": len(mappings)
            }
        else:
            raise HTTPException(status_code=500, detail="Failed to create template")
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to upload special cases template: {e}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


@app.put("/api/special-cases-templates/{template_id}")
async def update_special_cases_template_endpoint(
    template_id: int,
    name: str = Form(None),
    description: str = Form(None)
):
    """Update a special cases template's metadata"""
    try:
        from db_utils import update_special_cases_template
        success = update_special_cases_template(template_id, name=name, description=description)
        if success:
            return {"message": f"Template {template_id} updated successfully"}
        else:
            raise HTTPException(status_code=500, detail="Failed to update template")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update special cases template {template_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to update template: {str(e)}")


@app.put("/api/special-cases-templates/{template_id}/mappings")
async def update_special_cases_mappings(template_id: int, mappings: list = Body(...)):
    """Update the mappings in a special cases template"""
    try:
        from db_utils import update_special_cases_template
        success = update_special_cases_template(template_id, mappings=mappings)
        if success:
            return {"message": f"Mappings updated successfully", "count": len(mappings)}
        else:
            raise HTTPException(status_code=500, detail="Failed to update mappings")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to update mappings for template {template_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to update mappings: {str(e)}")


@app.delete("/api/special-cases-templates/{template_id}")
async def delete_special_cases_template_endpoint(template_id: int):
    """Delete a special cases template"""
    try:
        from db_utils import delete_special_cases_template
        success = delete_special_cases_template(template_id)
        if success:
            return {"message": f"Template {template_id} deleted successfully"}
        else:
            raise HTTPException(status_code=500, detail="Failed to delete template")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to delete special cases template {template_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to delete template: {str(e)}")




@app.post("/api/insurance-mappings/bulk-import")
async def bulk_import_insurance_mappings(
    csv_file: UploadFile = File(...),
    clear_existing: bool = Form(default=False)
):
    """
    Bulk import insurance mappings from CSV file.
    CSV format: InputValue,OutputValue
    
    Args:
        csv_file: CSV file with InputValue,OutputValue columns
        clear_existing: If true, delete all existing mappings before importing
    """
    try:
        from db_utils import bulk_import_insurance_mappings as bulk_import
        import pandas as pd
        import io
        
        # Read CSV file
        contents = await csv_file.read()
        
        # Try different encodings
        encodings = ['utf-8', 'utf-8-sig', 'latin-1', 'iso-8859-1', 'cp1252']
        df = None
        for encoding in encodings:
            try:
                df = pd.read_csv(io.BytesIO(contents), dtype=str, encoding=encoding)
                break
            except UnicodeDecodeError:
                continue
        
        if df is None:
            raise HTTPException(status_code=400, detail="Could not read CSV file with any standard encoding")
        
        # Validate columns
        if 'InputValue' not in df.columns or 'OutputValue' not in df.columns:
            raise HTTPException(
                status_code=400, 
                detail="CSV must have 'InputValue' and 'OutputValue' columns"
            )
        
        # Prepare mappings data
        mappings_data = []
        for _, row in df.iterrows():
            input_val = str(row['InputValue']).strip() if pd.notna(row['InputValue']) else ''
            output_val = str(row['OutputValue']).strip() if pd.notna(row['OutputValue']) else ''
            
            if input_val and output_val:
                mappings_data.append({
                    'input_code': input_val,
                    'output_code': output_val,
                    'description': f'Imported from CSV'
                })
        
        if not mappings_data:
            raise HTTPException(status_code=400, detail="No valid mappings found in CSV")
        
        # Bulk import
        result = bulk_import(mappings_data, clear_existing=clear_existing)
        
        if result['success']:
            return {
                "message": "Bulk import completed",
                "imported": result['imported'],
                "updated": result['updated'],
                "skipped": result['skipped'],
                "total": result['total'],
                "errors": result['errors']
            }
        else:
            raise HTTPException(
                status_code=500, 
                detail=f"Import failed: {'; '.join(result['errors'])}"
            )
            
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to bulk import insurance mappings: {e}")
        raise HTTPException(status_code=500, detail=f"Import failed: {str(e)}")


def process_unified_background(
    job_id: str,
    zip_path: str,
    excel_path: str,
    excel_filename: str,
    # Extraction params
    enable_extraction: bool,
    extraction_n_pages: int,
    extraction_model: str,
    extraction_max_workers: int,
    worktracker_group: str,
    worktracker_batch: str,
    extract_csn: bool,
    # CPT params
    enable_cpt: bool,
    cpt_vision_mode: bool,
    cpt_client: str,
    cpt_vision_pages: int,
    cpt_vision_model: str,
    cpt_include_code_list: bool,
    cpt_max_workers: int,
    cpt_custom_instructions: str,
    cpt_instruction_template_id: Optional[int],
    # ICD params
    enable_icd: bool,
    icd_n_pages: int,
    icd_vision_model: str,
    icd_max_workers: int,
    icd_custom_instructions: str,
    icd_instruction_template_id: Optional[int],
    output_filename: Optional[str] = None
):
    """Unified background task to run extraction + CPT + ICD prediction"""
    import os
    
    job = job_status[job_id]
    
    try:
        job.status = "processing"
        job.message = "Starting unified processing..."
        job.progress = 5
        
        # Create temporary directory for processing
        temp_dir = Path(f"/tmp/unified_{job_id}")
        temp_dir.mkdir(exist_ok=True)
        
        extraction_csv_path = None
        cpt_csv_path = None
        icd_csv_path = None
        
        # ==================== PARALLEL EXECUTION STRATEGY ====================
        # Determine which operations can run in parallel:
        # 1. All three (Extraction + CPT Vision + ICD) can run in parallel - MAXIMUM SPEED!
        #    - Extraction, CPT Vision, and ICD Vision can ALL run simultaneously since they're independent
        # 2. Extraction + ICD can run in parallel, then CPT non-vision after extraction
        #    - CPT non-vision needs extraction results, so it must wait
        # 3. Fall back to sequential processing for other cases
        
        # ALWAYS run all three in parallel when extraction is enabled and both CPT/ICD use vision mode
        # This is the fastest possible configuration - all operations are independent!
        run_all_three_parallel = (
            enable_extraction and 
            enable_cpt and cpt_vision_mode and 
            enable_icd
        )
        
        run_extraction_icd_parallel = (
            enable_extraction and 
            enable_cpt and not cpt_vision_mode and 
            enable_icd
        )
        
        if run_all_three_parallel:
            # ==================== MAXIMUM PARALLELIZATION ====================
            # Run Extraction + CPT Vision + ICD Vision all at the same time!
            job.message = "Running all three operations in parallel..."
            job.progress = 10
            logger.info(f"[Unified {job_id}] 🚀 MAXIMUM SPEED: Running Extraction + CPT Vision + ICD Vision in parallel!")
            
            # Unzip files once
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(temp_dir / "input")
            
            # Prepare extraction
            excel_dest = temp_dir / "instructions" / excel_filename
            excel_dest.parent.mkdir(exist_ok=True)
            shutil.copy2(excel_path, excel_dest)
            
            # Import prediction functions
            general_coding_path = Path(__file__).parent / "general-coding"
            sys.path.insert(0, str(general_coding_path))
            from predict_general import predict_codes_from_pdfs_api, predict_icd_codes_from_pdfs_api, pdf_pages_to_base64_images
            
            # OPTIMIZATION: Pre-extract and cache PDF images to share between CPT and ICD
            # This avoids extracting the same PDFs twice when both use vision mode
            pdf_image_cache = {}  # filename -> list of base64 images
            if enable_cpt and cpt_vision_mode and enable_icd:
                # Determine max pages needed (use the larger of the two)
                max_pages_needed = max(cpt_vision_pages, icd_n_pages)
                logger.info(f"[Unified {job_id}] ⚡ SPEED OPTIMIZATION: Pre-extracting PDF images (max {max_pages_needed} pages) to share between CPT and ICD...")
                
                # Match extraction script's pattern: search root + recursive, both cases, then deduplicate
                import glob as glob_module
                input_path_str = str(temp_dir / "input")
                pdf_files_list = glob_module.glob(os.path.join(input_path_str, "*.pdf"))
                pdf_files_list.extend(glob_module.glob(os.path.join(input_path_str, "*.PDF")))
                pdf_files_list.extend(glob_module.glob(os.path.join(input_path_str, "**", "*.pdf"), recursive=True))
                pdf_files_list.extend(glob_module.glob(os.path.join(input_path_str, "**", "*.PDF"), recursive=True))
                pdf_files = [Path(f) for f in set(pdf_files_list)]  # Deduplicate (matches extraction script logic)
                pdf_files = sorted(pdf_files, key=lambda x: x.name)
                
                # Pre-extract images in parallel
                from concurrent.futures import ThreadPoolExecutor, as_completed
                if pdf_files:
                    with ThreadPoolExecutor(max_workers=min(20, len(pdf_files))) as executor:
                        future_to_pdf = {executor.submit(pdf_pages_to_base64_images, str(pdf_path), max_pages_needed): pdf_path for pdf_path in pdf_files}
                        for future in as_completed(future_to_pdf):
                            pdf_path = future_to_pdf[future]
                            try:
                                images = future.result()
                                if images:
                                    pdf_image_cache[pdf_path.name] = images
                            except Exception as e:
                                logger.warning(f"[Unified {job_id}] Failed to pre-extract images from {pdf_path.name}: {e}")
                
                logger.info(f"[Unified {job_id}] ✅ Pre-extracted images for {len(pdf_image_cache)} PDFs (shared cache)")
            
            # Thread-safe progress tracking for all three operations
            import threading
            extraction_completed = [0]
            cpt_completed = [0]
            icd_completed = [0]
            extraction_total = [0]
            cpt_total = [0]
            icd_total = [0]
            lock = threading.Lock()
            
            def update_progress():
                with lock:
                    total_completed = extraction_completed[0] + cpt_completed[0] + icd_completed[0]
                    total_tasks = extraction_total[0] + cpt_total[0] + icd_total[0]
                    if total_tasks > 0:
                        progress_pct = 10 + int((total_completed / total_tasks) * 75)
                        job.progress = min(progress_pct, 85)
                        job.message = f"Extraction: {extraction_completed[0]}/{extraction_total[0]}, CPT: {cpt_completed[0]}/{cpt_total[0]}, ICD: {icd_completed[0]}/{icd_total[0]}"
            
            # Results storage
            extraction_result = [None]
            cpt_result = [None]
            icd_result = [None]
            extraction_error = [None]
            cpt_error = [None]
            icd_error = [None]
            
            # Thread 1: Extraction
            def run_extraction():
                try:
                    import os
                    logger.info(f"[Unified {job_id}] Starting extraction thread...")
                    env = os.environ.copy()
                    env['PYTHONPATH'] = str(Path(__file__).parent / "current")
                    env['OPENBLAS_NUM_THREADS'] = '12'
                    env['OMP_NUM_THREADS'] = '12'
                    env['MKL_NUM_THREADS'] = '12'
                    
                    script_path = Path(__file__).parent / "current" / "2-extract_info.py"
                    
                    if not script_path.exists():
                        raise Exception(f"Extraction script not found: {script_path}")
                    
                    cmd = [
                        sys.executable,
                        str(script_path),
                        str(temp_dir / "input"),
                        str(excel_dest),
                        str(extraction_n_pages),
                        str(extraction_max_workers),
                        extraction_model
                    ]
                    
                    if worktracker_group:
                        cmd.append(worktracker_group)
                    else:
                        cmd.append("")
                        
                    if worktracker_batch:
                        cmd.append(worktracker_batch)
                    else:
                        cmd.append("")
                    
                    if extract_csn:
                        cmd.append("true")
                    else:
                        cmd.append("false")
                    
                    # Create progress file for extraction script to write to
                    progress_file_path = temp_dir / "extraction_progress.txt"
                    cmd.append(str(progress_file_path))
                    
                    logger.info(f"[Unified {job_id}] Extraction command: {' '.join(cmd)}")
                    
                    # Estimate total (rough guess based on number of PDFs)
                    # Match extraction script's pattern: search root + recursive, both cases, then deduplicate
                    import glob as glob_module
                    input_path_str = str(temp_dir / "input")
                    pdf_files_list = glob_module.glob(os.path.join(input_path_str, "*.pdf"))
                    pdf_files_list.extend(glob_module.glob(os.path.join(input_path_str, "*.PDF")))
                    pdf_files_list.extend(glob_module.glob(os.path.join(input_path_str, "**", "*.pdf"), recursive=True))
                    pdf_files_list.extend(glob_module.glob(os.path.join(input_path_str, "**", "*.PDF"), recursive=True))
                    pdf_count = len(set(pdf_files_list))  # Deduplicate (matches extraction script logic)
                    logger.info(f"[Unified {job_id}] Found {pdf_count} PDFs to extract")
                    with lock:
                        extraction_total[0] = pdf_count
                    
                    process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True, cwd=temp_dir, env=env, bufsize=1)
                    logger.info(f"[Unified {job_id}] Extraction subprocess started with PID {process.pid}")
                    
                    # Start a thread to stream stdout/stderr in real-time
                    def stream_output():
                        for line in iter(process.stdout.readline, ''):
                            if line:
                                logger.info(f"[Unified {job_id}] Extraction: {line.strip()}")
                    
                    output_thread = threading.Thread(target=stream_output, daemon=True)
                    output_thread.start()
                    
                    # Monitor process and update progress from progress file
                    last_progress = -1
                    while process.poll() is None:
                        time.sleep(2)
                        # Read progress from file written by extraction script
                        if progress_file_path.exists():
                            try:
                                with open(progress_file_path, 'r') as f:
                                    lines = f.readlines()
                                    if len(lines) >= 2:
                                        completed = int(lines[0].strip())
                                        total = int(lines[1].strip())
                                        if completed != last_progress:
                                            logger.info(f"[Unified {job_id}] Extraction progress: {completed}/{total}")
                                            last_progress = completed
                                        with lock:
                                            extraction_completed[0] = completed
                                            if total > 0:
                                                extraction_total[0] = total
                            except (ValueError, IOError) as e:
                                logger.debug(f"[Unified {job_id}] Could not read progress file: {e}")
                        else:
                            logger.debug(f"[Unified {job_id}] Progress file does not exist yet: {progress_file_path}")
                        update_progress()
                    
                    # Wait for output thread to finish collecting stdout
                    output_thread.join(timeout=5)
                    
                    # Final progress update
                    if progress_file_path.exists():
                        try:
                            with open(progress_file_path, 'r') as f:
                                lines = f.readlines()
                                if len(lines) >= 2:
                                    completed = int(lines[0].strip())
                                    total = int(lines[1].strip())
                                    logger.info(f"[Unified {job_id}] Final extraction progress: {completed}/{total}")
                                    with lock:
                                        extraction_completed[0] = completed
                                        if total > 0:
                                            extraction_total[0] = total
                        except (ValueError, IOError) as e:
                            logger.debug(f"[Unified {job_id}] Could not read final progress file: {e}")
                    
                    logger.info(f"[Unified {job_id}] Extraction process completed with return code: {process.returncode}")
                    
                    if process.returncode != 0:
                        raise Exception(f"Extraction failed with return code {process.returncode}. Check logs for details.")
                    
                    # Find the output CSV
                    extracted_files = list(temp_dir.glob("extracted/combined_patient_data_*.csv"))
                    if extracted_files:
                        extraction_result[0] = str(extracted_files[0])
                        with lock:
                            extraction_completed[0] = extraction_total[0]
                        update_progress()
                        logger.info(f"[Unified {job_id}] Extraction completed: {extraction_result[0]}")
                    else:
                        raise Exception("Extraction completed but no output CSV found")
                        
                except Exception as e:
                    extraction_error[0] = str(e)
                    logger.error(f"[Unified {job_id}] Extraction error: {e}")
            
            # Thread 2: CPT Vision
            def run_cpt():
                try:
                    import os
                    cpt_csv_path_local = str(temp_dir / "cpt_predictions.csv")
                    Path(cpt_csv_path_local).parent.mkdir(exist_ok=True)
                    
                    # Check if using Gemini model
                    general_coding_path = Path(__file__).parent / "general-coding"
                    sys.path.insert(0, str(general_coding_path))
                    from predict_general import is_gemini_model
                    using_gemini = is_gemini_model(cpt_vision_model)
                    
                    # Use GOOGLE_API_KEY for Gemini models, OPENROUTER_API_KEY for others
                    if using_gemini:
                        api_key = os.environ.get('GOOGLE_API_KEY')
                    else:
                        api_key = os.environ.get('OPENROUTER_API_KEY') or os.environ.get('OPENAI_API_KEY')
                    
                    def cpt_progress(completed, total, message):
                        with lock:
                            cpt_completed[0] = completed
                            cpt_total[0] = total
                        update_progress()
                    
                    # Use shared image cache if available (optimization)
                    result = predict_codes_from_pdfs_api(
                        pdf_folder=str(temp_dir / "input"),
                        output_file=cpt_csv_path_local,
                        n_pages=cpt_vision_pages,
                        model=cpt_vision_model,
                        api_key=api_key,
                        max_workers=cpt_max_workers,
                        progress_callback=cpt_progress,
                        custom_instructions=cpt_custom_instructions,
                        include_code_list=cpt_include_code_list,
                        image_cache=pdf_image_cache if pdf_image_cache else None
                    )
                    cpt_result[0] = cpt_csv_path_local if result else None
                    if not result:
                        raise Exception("CPT prediction returned False")
                except Exception as e:
                    cpt_error[0] = str(e)
                    logger.error(f"[Unified {job_id}] CPT error: {e}")
            
            # Thread 3: ICD Vision
            def run_icd():
                try:
                    import os
                    icd_csv_path_local = str(temp_dir / "icd_predictions.csv")
                    Path(icd_csv_path_local).parent.mkdir(exist_ok=True)
                    
                    # Check if using Gemini model
                    general_coding_path = Path(__file__).parent / "general-coding"
                    sys.path.insert(0, str(general_coding_path))
                    from predict_general import is_gemini_model
                    using_gemini = is_gemini_model(icd_vision_model)
                    
                    # Use GOOGLE_API_KEY for Gemini models, OPENROUTER_API_KEY for others
                    if using_gemini:
                        api_key = os.environ.get('GOOGLE_API_KEY')
                    else:
                        api_key = os.environ.get('OPENROUTER_API_KEY') or os.environ.get('OPENAI_API_KEY')
                    
                    def icd_progress(completed, total, message):
                        with lock:
                            icd_completed[0] = completed
                            icd_total[0] = total
                        update_progress()
                    
                    # Note: No image cache here since extraction+ICD parallel doesn't share with CPT
                    result = predict_icd_codes_from_pdfs_api(
                        pdf_folder=str(temp_dir / "input"),
                        output_file=icd_csv_path_local,
                        n_pages=icd_n_pages,
                        model=icd_vision_model,
                        api_key=api_key,
                        max_workers=icd_max_workers,
                        progress_callback=icd_progress,
                        custom_instructions=icd_custom_instructions,
                        image_cache=None
                    )
                    icd_result[0] = icd_csv_path_local if result else None
                    if not result:
                        raise Exception("ICD prediction returned False")
                except Exception as e:
                    icd_error[0] = str(e)
                    logger.error(f"[Unified {job_id}] ICD error: {e}")
            
            # Start all three threads
            extraction_thread = threading.Thread(target=run_extraction)
            cpt_thread = threading.Thread(target=run_cpt)
            icd_thread = threading.Thread(target=run_icd)
            
            logger.info(f"[Unified {job_id}] Starting extraction, CPT, and ICD threads...")
            extraction_thread.start()
            cpt_thread.start()
            icd_thread.start()
            logger.info(f"[Unified {job_id}] All threads started")
            
            # Wait for all three to complete
            extraction_thread.join()
            cpt_thread.join()
            icd_thread.join()
            
            # Check results
            if extraction_error[0]:
                raise Exception(f"Extraction failed: {extraction_error[0]}")
            if cpt_error[0]:
                raise Exception(f"CPT prediction failed: {cpt_error[0]}")
            if icd_error[0]:
                raise Exception(f"ICD prediction failed: {icd_error[0]}")
            
            # Set paths for merging
            extraction_csv_path = extraction_result[0]
            cpt_csv_path = cpt_result[0]
            icd_csv_path = icd_result[0]
            
            # Verify files were created
            if extraction_csv_path and not os.path.exists(extraction_csv_path):
                raise Exception(f"Extraction reported success but file not found: {extraction_csv_path}")
            if cpt_csv_path and not os.path.exists(cpt_csv_path):
                raise Exception(f"CPT prediction reported success but file not found: {cpt_csv_path}")
            if icd_csv_path and not os.path.exists(icd_csv_path):
                raise Exception(f"ICD prediction reported success but file not found: {icd_csv_path}")
            
            logger.info(f"[Unified {job_id}] ✨ All three operations completed in parallel!")
            job.message = "All operations completed in parallel"
            job.progress = 85
            gc.collect()
        
        elif run_extraction_icd_parallel:
            # ==================== EXTRACTION + ICD IN PARALLEL ====================
            # Run Extraction + ICD in parallel, then CPT non-vision after
            job.message = "Running extraction and ICD in parallel..."
            job.progress = 10
            logger.info(f"[Unified {job_id}] Running Extraction + ICD in parallel, then CPT")
            
            # Unzip files once
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(temp_dir / "input")
            
            # Prepare extraction
            excel_dest = temp_dir / "instructions" / excel_filename
            excel_dest.parent.mkdir(exist_ok=True)
            shutil.copy2(excel_path, excel_dest)
            
            # Import ICD prediction function
            general_coding_path = Path(__file__).parent / "general-coding"
            sys.path.insert(0, str(general_coding_path))
            from predict_general import predict_icd_codes_from_pdfs_api
            
            # Thread-safe progress tracking
            import threading
            extraction_completed = [0]
            icd_completed = [0]
            extraction_total = [0]
            icd_total = [0]
            lock = threading.Lock()
            
            def update_progress():
                with lock:
                    total_completed = extraction_completed[0] + icd_completed[0]
                    total_tasks = extraction_total[0] + icd_total[0]
                    if total_tasks > 0:
                        progress_pct = 10 + int((total_completed / total_tasks) * 50)
                        job.progress = min(progress_pct, 60)
                        job.message = f"Extraction: {extraction_completed[0]}/{extraction_total[0]}, ICD: {icd_completed[0]}/{icd_total[0]}"
            
            # Results storage
            extraction_result = [None]
            icd_result = [None]
            extraction_error = [None]
            icd_error = [None]
            
            # Thread 1: Extraction
            def run_extraction():
                try:
                    import os
                    env = os.environ.copy()
                    env['PYTHONPATH'] = str(Path(__file__).parent / "current")
                    env['OPENBLAS_NUM_THREADS'] = '12'
                    env['OMP_NUM_THREADS'] = '12'
                    env['MKL_NUM_THREADS'] = '12'
                    
                    script_path = Path(__file__).parent / "current" / "2-extract_info.py"
                    
                    cmd = [
                        sys.executable,
                        "-u",  # Unbuffered output for real-time logging
                        str(script_path),
                        str(temp_dir / "input"),
                        str(excel_dest),
                        str(extraction_n_pages),
                        str(extraction_max_workers),
                        extraction_model
                    ]
                    
                    if worktracker_group:
                        cmd.append(worktracker_group)
                    else:
                        cmd.append("")
                        
                    if worktracker_batch:
                        cmd.append(worktracker_batch)
                    else:
                        cmd.append("")
                    
                    if extract_csn:
                        cmd.append("true")
                    else:
                        cmd.append("false")
                    
                    # Create progress file for extraction script to write to
                    progress_file_path = temp_dir / "extraction_progress.txt"
                    cmd.append(str(progress_file_path))
                    
                    # Estimate total
                    # Match extraction script's pattern: search root + recursive, both cases, then deduplicate
                    import glob as glob_module
                    input_path_str = str(temp_dir / "input")
                    pdf_files_list = glob_module.glob(os.path.join(input_path_str, "*.pdf"))
                    pdf_files_list.extend(glob_module.glob(os.path.join(input_path_str, "*.PDF")))
                    pdf_files_list.extend(glob_module.glob(os.path.join(input_path_str, "**", "*.pdf"), recursive=True))
                    pdf_files_list.extend(glob_module.glob(os.path.join(input_path_str, "**", "*.PDF"), recursive=True))
                    pdf_count = len(set(pdf_files_list))  # Deduplicate (matches extraction script logic)
                    with lock:
                        extraction_total[0] = pdf_count
                    
                    process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, cwd=temp_dir, env=env)
                    
                    # Monitor process and update progress from progress file
                    while process.poll() is None:
                        time.sleep(2)
                        # Read progress from file written by extraction script
                        if progress_file_path.exists():
                            try:
                                with open(progress_file_path, 'r') as f:
                                    lines = f.readlines()
                                    if len(lines) >= 2:
                                        completed = int(lines[0].strip())
                                        total = int(lines[1].strip())
                                        with lock:
                                            extraction_completed[0] = completed
                                            if total > 0:
                                                extraction_total[0] = total
                            except (ValueError, IOError) as e:
                                logger.debug(f"[Unified {job_id}] Could not read progress file: {e}")
                        update_progress()
                    
                    stdout, stderr = process.communicate()
                    
                    logger.info(f"[Unified {job_id}] Extraction stdout: {stdout}")
                    if stderr:
                        logger.info(f"[Unified {job_id}] Extraction stderr: {stderr}")
                    
                    if process.returncode != 0:
                        raise Exception(f"Extraction failed: {stderr}")
                    
                    extracted_files = list(temp_dir.glob("extracted/combined_patient_data_*.csv"))
                    if extracted_files:
                        extraction_result[0] = str(extracted_files[0])
                        with lock:
                            extraction_completed[0] = extraction_total[0]
                        update_progress()
                        logger.info(f"[Unified {job_id}] Extraction completed: {extraction_result[0]}")
                    else:
                        raise Exception("Extraction completed but no output CSV found")
                        
                except Exception as e:
                    extraction_error[0] = str(e)
                    logger.error(f"[Unified {job_id}] Extraction error: {e}")
            
            # Thread 2: ICD Vision
            def run_icd():
                try:
                    import os
                    icd_csv_path_local = str(temp_dir / "icd_predictions.csv")
                    Path(icd_csv_path_local).parent.mkdir(exist_ok=True)
                    
                    api_key = os.environ.get('OPENROUTER_API_KEY') or os.environ.get('OPENAI_API_KEY')
                    
                    def icd_progress(completed, total, message):
                        with lock:
                            icd_completed[0] = completed
                            icd_total[0] = total
                        update_progress()
                    
                    result = predict_icd_codes_from_pdfs_api(
                        pdf_folder=str(temp_dir / "input"),
                        output_file=icd_csv_path_local,
                        n_pages=icd_n_pages,
                        model=icd_vision_model,
                        api_key=api_key,
                        max_workers=icd_max_workers,
                        progress_callback=icd_progress,
                        custom_instructions=icd_custom_instructions
                    )
                    icd_result[0] = icd_csv_path_local if result else None
                    if not result:
                        raise Exception("ICD prediction returned False")
                except Exception as e:
                    icd_error[0] = str(e)
                    logger.error(f"[Unified {job_id}] ICD error: {e}")
            
            # Start both threads
            extraction_thread = threading.Thread(target=run_extraction)
            icd_thread = threading.Thread(target=run_icd)
            
            extraction_thread.start()
            icd_thread.start()
            
            # Wait for both to complete
            extraction_thread.join()
            icd_thread.join()
            
            # Check results
            if extraction_error[0]:
                raise Exception(f"Extraction failed: {extraction_error[0]}")
            if icd_error[0]:
                raise Exception(f"ICD prediction failed: {icd_error[0]}")
            
            extraction_csv_path = extraction_result[0]
            icd_csv_path = icd_result[0]
            
            if extraction_csv_path and not os.path.exists(extraction_csv_path):
                raise Exception(f"Extraction reported success but file not found: {extraction_csv_path}")
            if icd_csv_path and not os.path.exists(icd_csv_path):
                raise Exception(f"ICD prediction reported success but file not found: {icd_csv_path}")
            
            logger.info(f"[Unified {job_id}] Extraction + ICD completed in parallel")
            job.message = "Running CPT prediction on extracted data..."
            job.progress = 60
            gc.collect()
            
            # Now run CPT non-vision on the extraction results
            # (Keep the existing CPT non-vision logic here - it will execute after this block)
        
        elif enable_extraction:
            job.message = "Step 1/3: Extracting data from PDFs..."
            job.progress = 10
            logger.info(f"[Unified {job_id}] Starting data extraction")
            
            # Unzip files
            with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                zip_ref.extractall(temp_dir / "input")
            
            # Copy Excel file to temp directory
            excel_dest = temp_dir / "instructions" / excel_filename
            excel_dest.parent.mkdir(exist_ok=True)
            shutil.copy2(excel_path, excel_dest)
            
            # Set up environment for the processing script
            env = os.environ.copy()
            env['PYTHONPATH'] = str(Path(__file__).parent / "current")
            env['OPENBLAS_NUM_THREADS'] = '12'
            env['OMP_NUM_THREADS'] = '12'
            env['MKL_NUM_THREADS'] = '12'
            
            # Run the extraction script
            script_path = Path(__file__).parent / "current" / "2-extract_info.py"
            
            job.message = f"Extracting data (first {extraction_n_pages} pages per patient)..."
            job.progress = 15
            
            try:
                cmd = [
                    sys.executable,
                    str(script_path),
                    str(temp_dir / "input"),
                    str(excel_dest),
                    str(extraction_n_pages),
                    str(extraction_max_workers),  # Configurable max_workers
                    extraction_model
                ]
                
                if worktracker_group:
                    cmd.append(worktracker_group)
                else:
                    cmd.append("")
                    
                if worktracker_batch:
                    cmd.append(worktracker_batch)
                else:
                    cmd.append("")
                
                if extract_csn:
                    cmd.append("true")
                else:
                    cmd.append("false")
                
                # Add empty progress_file (not needed for sequential processing)
                cmd.append("")
                
                process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True, cwd=temp_dir, env=env)
                stdout, stderr = process.communicate(timeout=1800)
                
                logger.info(f"[Unified {job_id}] Extraction stdout: {stdout}")
                if stderr:
                    logger.info(f"[Unified {job_id}] Extraction stderr: {stderr}")
                
                if process.returncode != 0:
                    raise Exception(f"Extraction failed: {stderr}")
                    
            except subprocess.TimeoutExpired:
                if process:
                    logger.warning(f"[Unified {job_id}] Extraction timed out")
                    kill_process_tree(process)
                    gc.collect()
                raise Exception("Extraction timed out after 30 minutes")
            
            # Find the extraction output CSV
            extracted_files = list(temp_dir.glob("extracted/combined_patient_data_*.csv"))
            if extracted_files:
                extraction_csv_path = str(extracted_files[0])
                logger.info(f"[Unified {job_id}] Found extraction CSV: {extraction_csv_path}")
            else:
                raise Exception("Extraction completed but no output CSV found")
                
            job.message = "Data extraction completed"
            job.progress = 30
            gc.collect()
        
        # ==================== STEP 2 & 3: CPT and ICD Code Predictions ====================
        # Check if we can run CPT and ICD vision predictions in parallel
        # Both must be enabled, CPT must be in vision mode, and extraction must be completed or disabled
        # AND we haven't already processed them in the parallel paths above
        run_cpt_icd_parallel = (
            enable_cpt and cpt_vision_mode and 
            enable_icd and
            (not enable_extraction or (enable_extraction and extraction_csv_path is not None)) and
            not run_all_three_parallel and
            not run_extraction_icd_parallel
        )
        
        if run_cpt_icd_parallel:
            # Run CPT and ICD vision predictions in parallel for maximum speed
            job.message = "Step 2/2: Predicting CPT and ICD codes in parallel..."
            job.progress = 35
            logger.info(f"[Unified {job_id}] Running CPT and ICD vision predictions in parallel")
            
            # Ensure PDFs are unzipped
            if not (temp_dir / "input").exists():
                with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                    zip_ref.extractall(temp_dir / "input")
            
            # Import prediction functions
            general_coding_path = Path(__file__).parent / "general-coding"
            sys.path.insert(0, str(general_coding_path))
            from predict_general import predict_codes_from_pdfs_api, predict_icd_codes_from_pdfs_api, pdf_pages_to_base64_images
            
            # OPTIMIZATION: Pre-extract and cache PDF images to share between CPT and ICD
            pdf_image_cache = {}  # filename -> list of base64 images
            max_pages_needed = max(cpt_vision_pages, icd_n_pages)
            logger.info(f"[Unified {job_id}] ⚡ SPEED OPTIMIZATION: Pre-extracting PDF images (max {max_pages_needed} pages) to share between CPT and ICD...")
            
            # Match extraction script's pattern: search root + recursive, both cases, then deduplicate
            import glob as glob_module
            input_path_str = str(temp_dir / "input")
            pdf_files_list = glob_module.glob(os.path.join(input_path_str, "*.pdf"))
            pdf_files_list.extend(glob_module.glob(os.path.join(input_path_str, "*.PDF")))
            pdf_files_list.extend(glob_module.glob(os.path.join(input_path_str, "**", "*.pdf"), recursive=True))
            pdf_files_list.extend(glob_module.glob(os.path.join(input_path_str, "**", "*.PDF"), recursive=True))
            pdf_files = [Path(f) for f in set(pdf_files_list)]  # Deduplicate (matches extraction script logic)
            pdf_files = sorted(pdf_files, key=lambda x: x.name)
            
            # Pre-extract images in parallel
            from concurrent.futures import ThreadPoolExecutor, as_completed
            if pdf_files:
                with ThreadPoolExecutor(max_workers=min(20, len(pdf_files))) as executor:
                    future_to_pdf = {executor.submit(pdf_pages_to_base64_images, str(pdf_path), max_pages_needed): pdf_path for pdf_path in pdf_files}
                    for future in as_completed(future_to_pdf):
                        pdf_path = future_to_pdf[future]
                        try:
                            images = future.result()
                            if images:
                                pdf_image_cache[pdf_path.name] = images
                        except Exception as e:
                            logger.warning(f"[Unified {job_id}] Failed to pre-extract images from {pdf_path.name}: {e}")
            
            logger.info(f"[Unified {job_id}] ✅ Pre-extracted images for {len(pdf_image_cache)} PDFs (shared cache)")
            
            # Create output paths
            cpt_csv_path = str(temp_dir / "cpt_predictions.csv")
            icd_csv_path = str(temp_dir / "icd_predictions.csv")
            Path(cpt_csv_path).parent.mkdir(exist_ok=True)
            Path(icd_csv_path).parent.mkdir(exist_ok=True)
            
            # Get OpenRouter API key
            api_key = os.environ.get('OPENROUTER_API_KEY') or os.environ.get('OPENAI_API_KEY')
            
            # Thread-safe progress tracking
            import threading
            cpt_completed = [0]
            icd_completed = [0]
            cpt_total = [0]
            icd_total = [0]
            lock = threading.Lock()
            
            def cpt_progress(completed, total, message):
                with lock:
                    cpt_completed[0] = completed
                    cpt_total[0] = total
                    total_completed = cpt_completed[0] + icd_completed[0]
                    total_tasks = cpt_total[0] + icd_total[0]
                    if total_tasks > 0:
                        progress_pct = 35 + int((total_completed / total_tasks) * 50)
                        job.progress = min(progress_pct, 85)
                        job.message = f"CPT: {cpt_completed[0]}/{cpt_total[0]}, ICD: {icd_completed[0]}/{icd_total[0]}"
            
            def icd_progress(completed, total, message):
                with lock:
                    icd_completed[0] = completed
                    icd_total[0] = total
                    total_completed = cpt_completed[0] + icd_completed[0]
                    total_tasks = cpt_total[0] + icd_total[0]
                    if total_tasks > 0:
                        progress_pct = 35 + int((total_completed / total_tasks) * 50)
                        job.progress = min(progress_pct, 85)
                        job.message = f"CPT: {cpt_completed[0]}/{cpt_total[0]}, ICD: {icd_completed[0]}/{icd_total[0]}"
            
            # Run both predictions in parallel using threads
            import threading
            general_coding_path = Path(__file__).parent / "general-coding"
            sys.path.insert(0, str(general_coding_path))
            from predict_general import is_gemini_model
            
            cpt_result = [None]
            icd_result = [None]
            cpt_error = [None]
            icd_error = [None]
            
            def run_cpt():
                try:
                    import os
                    # Check if using Gemini model and select appropriate API key
                    using_gemini_cpt = is_gemini_model(cpt_vision_model)
                    if using_gemini_cpt:
                        cpt_api_key = os.environ.get('GOOGLE_API_KEY')
                    else:
                        cpt_api_key = os.environ.get('OPENROUTER_API_KEY') or os.environ.get('OPENAI_API_KEY')
                    
                    # Use shared image cache (optimization)
                    result = predict_codes_from_pdfs_api(
                        pdf_folder=str(temp_dir / "input"),
                        output_file=cpt_csv_path,
                        n_pages=cpt_vision_pages,
                        model=cpt_vision_model,  # Use selected vision model
                        api_key=cpt_api_key,
                        max_workers=cpt_max_workers,
                        progress_callback=cpt_progress,
                        custom_instructions=cpt_custom_instructions,
                        include_code_list=cpt_include_code_list,
                        image_cache=pdf_image_cache
                    )
                    cpt_result[0] = result
                except Exception as e:
                    cpt_error[0] = str(e)
            
            def run_icd():
                try:
                    import os
                    # Check if using Gemini model and select appropriate API key
                    using_gemini_icd = is_gemini_model(icd_vision_model)
                    if using_gemini_icd:
                        icd_api_key = os.environ.get('GOOGLE_API_KEY')
                    else:
                        icd_api_key = os.environ.get('OPENROUTER_API_KEY') or os.environ.get('OPENAI_API_KEY')
                    
                    # Use shared image cache (optimization)
                    result = predict_icd_codes_from_pdfs_api(
                        pdf_folder=str(temp_dir / "input"),
                        output_file=icd_csv_path,
                        n_pages=icd_n_pages,
                        model=icd_vision_model,  # Use selected vision model
                        api_key=icd_api_key,
                        max_workers=icd_max_workers,
                        progress_callback=icd_progress,
                        custom_instructions=icd_custom_instructions,
                        image_cache=pdf_image_cache
                    )
                    icd_result[0] = result
                except Exception as e:
                    icd_error[0] = str(e)
            
            # Start both threads
            cpt_thread = threading.Thread(target=run_cpt)
            icd_thread = threading.Thread(target=run_icd)
            
            cpt_thread.start()
            icd_thread.start()
            
            # Wait for both to complete
            cpt_thread.join()
            icd_thread.join()
            
            # Check results
            if cpt_error[0]:
                raise Exception(f"CPT prediction failed: {cpt_error[0]}")
            if icd_error[0]:
                raise Exception(f"ICD prediction failed: {icd_error[0]}")
            if not cpt_result[0]:
                raise Exception("CPT prediction failed")
            if not icd_result[0]:
                raise Exception("ICD prediction failed")
            
            # Verify files were created
            if not os.path.exists(cpt_csv_path):
                raise Exception(f"CPT prediction reported success but file not found: {cpt_csv_path}")
            if not os.path.exists(icd_csv_path):
                raise Exception(f"ICD prediction reported success but file not found: {icd_csv_path}")
            
            logger.info(f"[Unified {job_id}] CPT and ICD predictions completed in parallel")
            job.message = "CPT and ICD predictions completed"
            job.progress = 85
            gc.collect()
        
        # ==================== STEP 2: CPT Code Prediction (if not already done) ====================
        if enable_cpt and not cpt_csv_path:
            # CPT wasn't handled in parallel, so run it now
            if run_extraction_icd_parallel:
                # CPT non-vision runs after extraction+ICD parallel
                job.message = "Step 2/2: Predicting CPT codes from extracted data..."
                job.progress = 60
            else:
                # Sequential CPT processing
                job.message = "Step 2/3: Predicting CPT codes..."
                job.progress = 35
            logger.info(f"[Unified {job_id}] Starting CPT prediction (vision_mode={cpt_vision_mode}, client={cpt_client})")
            
            if cpt_vision_mode:
                # Use vision model with PDFs (always uses GPT-5)
                job.message = "Predicting CPT codes from PDF images..."
                
                # Ensure PDFs are unzipped
                if not (temp_dir / "input").exists():
                    with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                        zip_ref.extractall(temp_dir / "input")
                
                # Import prediction function
                general_coding_path = Path(__file__).parent / "general-coding"
                sys.path.insert(0, str(general_coding_path))
                from predict_general import predict_codes_from_pdfs_api, is_gemini_model
                
                # Create output path with .csv extension
                cpt_csv_path = str(temp_dir / "cpt_predictions.csv")
                Path(cpt_csv_path).parent.mkdir(exist_ok=True)
                
                # Check if using Gemini model and select appropriate API key
                using_gemini = is_gemini_model(cpt_vision_model)
                if using_gemini:
                    api_key = os.environ.get('GOOGLE_API_KEY')
                else:
                    api_key = os.environ.get('OPENROUTER_API_KEY') or os.environ.get('OPENAI_API_KEY')
                
                # Progress callback
                def cpt_progress(completed, total, message):
                    progress_pct = 35 + int((completed / total) * 20)
                    job.progress = min(progress_pct, 55)
                    job.message = f"CPT prediction: {message}"
                
                # Run CPT prediction with vision (uses selected model)
                success = predict_codes_from_pdfs_api(
                    pdf_folder=str(temp_dir / "input"),
                    output_file=cpt_csv_path,
                    n_pages=cpt_vision_pages,
                    model=cpt_vision_model,  # Use selected vision model
                    api_key=api_key,
                    max_workers=cpt_max_workers,
                    progress_callback=cpt_progress,
                    custom_instructions=cpt_custom_instructions,
                    include_code_list=cpt_include_code_list
                )
                
                if success:
                    # Verify file was actually created
                    if not os.path.exists(cpt_csv_path):
                        raise Exception(f"CPT prediction reported success but file not found: {cpt_csv_path}")
                    logger.info(f"[Unified {job_id}] CPT prediction completed: {cpt_csv_path}")
                else:
                    raise Exception("CPT prediction failed")
                    
            else:
                # Use CSV-based prediction (requires extraction first)
                if not extraction_csv_path:
                    raise Exception("CPT prediction requires data extraction to be enabled when not using vision mode")
                
                job.message = f"Predicting CPT codes from CSV using {cpt_client} model..."
                
                # Determine which prediction method to use based on client
                if cpt_client == "tan-esc":
                    # Use custom trained model
                    custom_coding_path = Path(__file__).parent / "custom-coding"
                    sys.path.insert(0, str(custom_coding_path))
                    from predict import predict_codes_api
                    
                    cpt_csv_path = str(temp_dir / "with_cpt_codes.csv")
                    
                    success = predict_codes_api(
                        input_file=extraction_csv_path,
                        output_file=cpt_csv_path,
                        model_dir=str(custom_coding_path),
                        confidence_threshold=0.5
                    )
                    
                elif cpt_client == "general":
                    # Use OpenAI general model
                    general_coding_path = Path(__file__).parent / "general-coding"
                    sys.path.insert(0, str(general_coding_path))
                    from predict_general import predict_codes_general_api
                    
                    cpt_csv_path = str(temp_dir / "with_cpt_codes.csv")
                    api_key = os.environ.get('OPENAI_API_KEY') or os.environ.get('OPENROUTER_API_KEY')
                    
                    # Dynamic progress based on whether we ran parallel extraction+ICD
                    progress_start = 60 if run_extraction_icd_parallel else 35
                    progress_range = 25 if run_extraction_icd_parallel else 20
                    progress_end = 85 if run_extraction_icd_parallel else 55
                    
                    def cpt_progress(completed, total, message):
                        progress_pct = progress_start + int((completed / total) * progress_range)
                        job.progress = min(progress_pct, progress_end)
                        job.message = f"CPT prediction: {message}"
                    
                    success = predict_codes_general_api(
                        input_file=extraction_csv_path,
                        output_file=cpt_csv_path,
                        model="gpt5",
                        api_key=api_key,
                        max_workers=cpt_max_workers,
                        progress_callback=cpt_progress,
                        custom_instructions=cpt_custom_instructions
                    )
                    
                else:
                    # Use client-specific prediction (UNI, SIO-STL, GAP-FIN, APO-UTP)
                    custom_coding_path = Path(__file__).parent / "custom-coding"
                    sys.path.insert(0, str(custom_coding_path))
                    from predict import predict_codes_api
                    
                    cpt_csv_path = str(temp_dir / "with_cpt_codes.csv")
                    
                    # Dynamic progress based on whether we ran parallel extraction+ICD
                    progress_start = 60 if run_extraction_icd_parallel else 35
                    progress_range = 25 if run_extraction_icd_parallel else 20
                    progress_end = 85 if run_extraction_icd_parallel else 55
                    
                    # Progress callback for custom model
                    def cpt_progress(completed, total, message):
                        progress_pct = progress_start + int((completed / total) * progress_range)
                        job.progress = min(progress_pct, progress_end)
                        job.message = f"CPT prediction: {message}"
                    
                    success = predict_codes_api(
                        input_file=extraction_csv_path,
                        output_file=cpt_csv_path,
                        model_dir=str(custom_coding_path),
                        confidence_threshold=0.5
                    )
                
                if not success:
                    raise Exception("CPT prediction failed")
                
                # Verify file was actually created
                if not os.path.exists(cpt_csv_path):
                    raise Exception(f"CPT prediction reported success but file not found: {cpt_csv_path}")
                
                logger.info(f"[Unified {job_id}] CPT prediction completed: {cpt_csv_path}")
            
            job.message = "CPT prediction completed"
            # Set final progress based on what was run in parallel
            if run_extraction_icd_parallel:
                job.progress = 85  # After extraction+ICD parallel
            else:
                job.progress = 55  # Sequential mode
            gc.collect()
        
        # ==================== STEP 3: ICD Code Prediction (if not already done) ====================
        if enable_icd and not icd_csv_path:
            # ICD wasn't handled in parallel, so run it now
            job.message = "Step 3/3: Predicting ICD codes..."
            job.progress = 60
            logger.info(f"[Unified {job_id}] Starting ICD prediction")
            
            # Ensure PDFs are unzipped
            if not (temp_dir / "input").exists():
                with zipfile.ZipFile(zip_path, 'r') as zip_ref:
                    zip_ref.extractall(temp_dir / "input")
            
            # Import prediction function
            general_coding_path = Path(__file__).parent / "general-coding"
            sys.path.insert(0, str(general_coding_path))
            from predict_general import predict_icd_codes_from_pdfs_api
            
            # Create output path with .csv extension
            icd_csv_path = str(temp_dir / "icd_predictions.csv")
            Path(icd_csv_path).parent.mkdir(exist_ok=True)
            
            # Check if using Gemini model and select appropriate API key
            general_coding_path = Path(__file__).parent / "general-coding"
            sys.path.insert(0, str(general_coding_path))
            from predict_general import is_gemini_model
            using_gemini = is_gemini_model(icd_vision_model)
            if using_gemini:
                api_key = os.environ.get('GOOGLE_API_KEY')
            else:
                api_key = os.environ.get('OPENROUTER_API_KEY') or os.environ.get('OPENAI_API_KEY')
            
            # Progress callback
            def icd_progress(completed, total, message):
                progress_pct = 60 + int((completed / total) * 25)
                job.progress = min(progress_pct, 85)
                job.message = f"ICD prediction: {message}"
            
            # Run ICD prediction (uses selected vision model)
            success = predict_icd_codes_from_pdfs_api(
                pdf_folder=str(temp_dir / "input"),
                output_file=icd_csv_path,
                n_pages=icd_n_pages,
                model=icd_vision_model,  # Use selected vision model
                api_key=api_key,
                max_workers=icd_max_workers,
                progress_callback=icd_progress,
                custom_instructions=icd_custom_instructions
            )
            
            if success:
                # Verify file was actually created
                if not os.path.exists(icd_csv_path):
                    raise Exception(f"ICD prediction reported success but file not found: {icd_csv_path}")
                logger.info(f"[Unified {job_id}] ICD prediction completed: {icd_csv_path}")
            else:
                raise Exception("ICD prediction failed")
            
            job.message = "ICD prediction completed"
            job.progress = 85
            gc.collect()
        
        # ==================== STEP 4: Merge Results ====================
        job.message = "Merging results..."
        job.progress = 90
        logger.info(f"[Unified {job_id}] Merging results")
        
        import pandas as pd
        
        # Determine the base dataframe
        if extraction_csv_path:
            # Verify file exists before reading
            if not os.path.exists(extraction_csv_path):
                raise Exception(f"Extraction CSV file not found: {extraction_csv_path}")
            # Start with extraction data
            base_df = pd.read_csv(extraction_csv_path, dtype=str)
            logger.info(f"[Unified {job_id}] Base: extraction CSV with {len(base_df)} rows and columns: {list(base_df.columns)[:10]}")
        elif cpt_csv_path and cpt_vision_mode:
            # Verify file exists before reading
            if not os.path.exists(cpt_csv_path):
                raise Exception(f"CPT CSV file not found: {cpt_csv_path}")
            # If no extraction but CPT vision mode, use CPT as base
            base_df = pd.read_csv(cpt_csv_path, dtype=str)
            logger.info(f"[Unified {job_id}] Base: CPT CSV with {len(base_df)} rows")
        elif icd_csv_path:
            # Verify file exists before reading
            if not os.path.exists(icd_csv_path):
                raise Exception(f"ICD CSV file not found: {icd_csv_path}")
            # If only ICD, use ICD as base
            base_df = pd.read_csv(icd_csv_path, dtype=str)
            logger.info(f"[Unified {job_id}] Base: ICD CSV with {len(base_df)} rows")
        else:
            raise Exception("No data to merge - at least one processing step must be enabled")
        
        # Merge CPT results if needed
        if enable_cpt and cpt_csv_path:
            # Verify file exists before reading
            if not os.path.exists(cpt_csv_path):
                raise Exception(f"CPT predictions file not found: {cpt_csv_path}")
            
            if cpt_vision_mode and not extraction_csv_path:
                # CPT is already the base
                pass
            elif cpt_vision_mode:
                # Merge CPT vision results with extraction
                cpt_df = pd.read_csv(cpt_csv_path, dtype=str)
                logger.info(f"[Unified {job_id}] CPT vision CSV columns: {list(cpt_df.columns)}")
                logger.info(f"[Unified {job_id}] Base DF columns: {list(base_df.columns)[:10]}")
                
                # Check for both 'Patient Filename' and 'Filename' column names
                filename_col = None
                if 'Patient Filename' in cpt_df.columns:
                    filename_col = 'Patient Filename'
                elif 'Filename' in cpt_df.columns:
                    filename_col = 'Filename'
                
                if filename_col and 'source_file' in base_df.columns:
                    # Extract CPT-related columns (actual column names from predict_codes_from_pdfs_api)
                    cpt_cols_to_merge = ['ASA Code', 'Procedure Code', 'Model Source', 'Error Message']
                    cpt_cols_available = [col for col in cpt_cols_to_merge if col in cpt_df.columns]
                    
                    # Include filename column for merging
                    merge_cols = [filename_col] + cpt_cols_available
                    cpt_df_merge = cpt_df[merge_cols]
                    
                    # Rename filename column to match for merge
                    cpt_df_merge = cpt_df_merge.rename(columns={filename_col: 'source_file'})
                    
                    base_df = base_df.merge(cpt_df_merge, on='source_file', how='left')
                    logger.info(f"[Unified {job_id}] Merged CPT vision results with columns: {cpt_cols_available}")
                else:
                    logger.warning(f"[Unified {job_id}] Cannot merge CPT: filename_col={filename_col}, source_file in base={('source_file' in base_df.columns)}")
            else:
                # CPT was run on extraction CSV, so it should already have all extraction columns
                base_df = pd.read_csv(cpt_csv_path, dtype=str)
                logger.info(f"[Unified {job_id}] Using CPT CSV as it contains extraction + CPT data")
        
        # Merge ICD results if needed
        if enable_icd and icd_csv_path:
            # Verify file exists before reading
            if not os.path.exists(icd_csv_path):
                raise Exception(f"ICD predictions file not found: {icd_csv_path}")
            icd_df = pd.read_csv(icd_csv_path, dtype=str)
            logger.info(f"[Unified {job_id}] ICD CSV columns: {list(icd_df.columns)}")
            logger.info(f"[Unified {job_id}] Base DF columns before ICD merge: {list(base_df.columns)[:10]}")
            
            # IMPORTANT: Remove any existing ICD columns from extraction data BEFORE merging
            # This prevents pandas from adding _x/_y suffixes
            icd_columns_to_remove = ['ICD1', 'ICD1 Reasoning', 'ICD2', 'ICD2 Reasoning', 'ICD3', 'ICD3 Reasoning', 'ICD4', 'ICD4 Reasoning']
            existing_icd_cols = [col for col in icd_columns_to_remove if col in base_df.columns]
            if existing_icd_cols:
                logger.info(f"[Unified {job_id}] Removing existing ICD columns from extraction data: {existing_icd_cols}")
                base_df = base_df.drop(columns=existing_icd_cols, errors='ignore')
            
            # Check for both 'Patient Filename' and 'Filename' column names
            filename_col = None
            if 'Patient Filename' in icd_df.columns:
                filename_col = 'Patient Filename'
            elif 'Filename' in icd_df.columns:
                filename_col = 'Filename'
            
            # Extract ICD columns to merge (including reasoning columns)
            icd_cols_to_merge = ['ICD1', 'ICD1 Reasoning', 'ICD2', 'ICD2 Reasoning', 'ICD3', 'ICD3 Reasoning', 'ICD4', 'ICD4 Reasoning', 'Model Source', 'Tokens Used', 'Cost (USD)', 'Error Message']
            icd_cols_available = [col for col in icd_cols_to_merge if col in icd_df.columns]
            
            if filename_col and 'source_file' in base_df.columns:
                # Merge on source_file
                merge_cols = [filename_col] + icd_cols_available
                icd_df_merge = icd_df[merge_cols]
                # Rename filename column to match for merge
                icd_df_merge = icd_df_merge.rename(columns={filename_col: 'source_file'})
                base_df = base_df.merge(icd_df_merge, on='source_file', how='left', suffixes=('', '_drop'))
                # Drop any columns with _drop suffix (shouldn't happen now, but just in case)
                base_df = base_df.drop(columns=[col for col in base_df.columns if col.endswith('_drop')], errors='ignore')
                logger.info(f"[Unified {job_id}] Merged ICD results with columns: {icd_cols_available}")
            elif filename_col and filename_col in base_df.columns:
                # Both have Filename column (or Patient Filename)
                merge_cols = [filename_col] + icd_cols_available
                icd_df_merge = icd_df[merge_cols]
                base_df = base_df.merge(icd_df_merge, on=filename_col, how='left', suffixes=('', '_drop'))
                # Drop any columns with _drop suffix (shouldn't happen now, but just in case)
                base_df = base_df.drop(columns=[col for col in base_df.columns if col.endswith('_drop')], errors='ignore')
                logger.info(f"[Unified {job_id}] Merged ICD results on {filename_col}")
            else:
                logger.warning(f"[Unified {job_id}] Cannot merge ICD: filename_col={filename_col}, source_file in base={('source_file' in base_df.columns)}")
        
        # Clean up any remaining merge suffixes (_x, _y) that might have been created
        # This is a safety measure in case any columns slipped through
        columns_to_clean = []
        for col in base_df.columns:
            if col.endswith('_x') or col.endswith('_y'):
                # Check if there's a base version without suffix
                base_col = col.rsplit('_', 1)[0]
                if base_col in base_df.columns:
                    # Keep the one without suffix, drop the suffixed one
                    columns_to_clean.append(col)
                else:
                    # Rename the suffixed column to remove the suffix
                    base_df = base_df.rename(columns={col: base_col})
        
        if columns_to_clean:
            logger.info(f"[Unified {job_id}] Cleaning up merge suffix columns: {columns_to_clean}")
            base_df = base_df.drop(columns=columns_to_clean, errors='ignore')
        
        # Reorder columns to put ICD columns at the end (after all other data)
        icd_cols = ['ICD1', 'ICD2', 'ICD3', 'ICD4']
        icd_cols_present = [col for col in icd_cols if col in base_df.columns]
        if icd_cols_present:
            # Get all non-ICD columns
            other_cols = [col for col in base_df.columns if col not in icd_cols_present]
            # Reorder: other columns first, then ICD columns
            base_df = base_df[other_cols + icd_cols_present]
            logger.info(f"[Unified {job_id}] Reordered columns to put ICD columns at the end")
        
        # Log final columns before saving
        logger.info(f"[Unified {job_id}] Final merged dataframe has {len(base_df)} rows and columns: {list(base_df.columns)}")
        
        # Save final result
        result_base = Path(f"/tmp/results/{job_id}_unified_result")
        result_base.parent.mkdir(exist_ok=True)
        
        # Save as both CSV and XLSX
        result_csv = f"{result_base}.csv"
        result_xlsx = f"{result_base}.xlsx"
        
        base_df.to_csv(result_csv, index=False)
        
        # Convert to XLSX with ID column protection
        try:
            # Explicitly set ID columns as text to prevent scientific notation in Excel
            id_columns = ['Primary Subsc ID', 'Secondary Subsc ID', 'MRN', 'CSN']
            for col in id_columns:
                if col in base_df.columns:
                    # Only convert non-empty values to string to avoid 'nan' text
                    base_df[col] = base_df[col].apply(lambda x: str(x) if x != '' and pd.notna(x) else '')
            
            base_df.to_excel(result_xlsx, index=False, engine='openpyxl')
        except Exception as e:
            logger.warning(f"[Unified {job_id}] Failed to create XLSX: {e}")
            result_xlsx = None
        
        job.result_file = result_csv
        job.result_file_xlsx = result_xlsx
        job.status = "completed"
        job.message = "Unified processing completed successfully"
        job.progress = 100
        
        logger.info(f"[Unified {job_id}] Processing completed: {result_csv}")
        
        # Save result metadata to database
        try:
            from db_utils import save_unified_result
            import os
            
            file_size = os.path.getsize(result_csv) if os.path.exists(result_csv) else None
            filename = output_filename if output_filename else f"unified_result_{job_id}"
            
            save_unified_result(
                job_id=job_id,
                filename=filename,
                file_path_csv=result_csv,
                file_path_xlsx=result_xlsx,
                file_size_bytes=file_size,
                row_count=len(base_df),
                enabled_extraction=enable_extraction,
                enabled_cpt=enable_cpt,
                enabled_icd=enable_icd,
                extraction_model=extraction_model if enable_extraction else None,
                cpt_vision_model=cpt_vision_model if enable_cpt and cpt_vision_mode else None,
                icd_vision_model=icd_vision_model if enable_icd else None,
                status="completed"
            )
            logger.info(f"[Unified {job_id}] Saved result metadata to database")
        except Exception as e:
            logger.warning(f"[Unified {job_id}] Failed to save result metadata: {e}")
        
        # Clean up temp directory
        try:
            shutil.rmtree(temp_dir)
        except Exception as e:
            logger.warning(f"[Unified {job_id}] Failed to clean up temp directory: {e}")
        
        # Force garbage collection
        gc.collect()
        
    except Exception as e:
        logger.error(f"[Unified {job_id}] Error: {str(e)}")
        job.status = "failed"
        job.error = str(e)
        job.message = f"Processing failed: {str(e)}"
        
        # Clean up on failure
        try:
            if temp_dir.exists():
                shutil.rmtree(temp_dir)
        except:
            pass
        
        gc.collect()


@app.post("/process-unified")
async def process_unified(
    background_tasks: BackgroundTasks,
    zip_file: UploadFile = File(None),  # Optional - used when split is disabled
    pdf_file: UploadFile = File(None),  # Optional - used when split is enabled
    excel_file: UploadFile = File(None),  # Optional - can use template instead
    template_id: Optional[int] = Form(default=None),  # Use saved template instead of excel file
    # Split parameters (Step 0)
    enable_split: bool = Form(default=False),
    split_filter_string: str = Form(default=""),
    split_method: str = Form(default="ocrspace"),  # "ocrspace" or "legacy"
    split_detection_shift: int = Form(default=0, description="Shift detections by N pages (positive = down, negative = up)"),
    # Extraction parameters
    enable_extraction: bool = Form(default=True),
    extraction_n_pages: int = Form(default=2),
    extraction_model: str = Form(default="gemini-flash-latest"),
    extraction_max_workers: int = Form(default=50),  # Configurable extraction parallelism
    worktracker_group: str = Form(default=""),
    worktracker_batch: str = Form(default=""),
    extract_csn: bool = Form(default=False),
    # CPT parameters
    enable_cpt: bool = Form(default=True),
    cpt_vision_mode: bool = Form(default=False),
    cpt_client: str = Form(default="uni"),  # For non-vision mode
    cpt_vision_pages: int = Form(default=1),  # For vision mode
    cpt_vision_model: str = Form(default="openai/gpt-5.2:online"),  # Vision model selection
    cpt_include_code_list: bool = Form(default=True),  # Include CPT code list in prompt
    cpt_max_workers: int = Form(default=50),  # Increased for better parallelism
    cpt_custom_instructions: str = Form(default=""),
    cpt_instruction_template_id: Optional[int] = Form(default=None),  # For template selection
    # ICD parameters
    enable_icd: bool = Form(default=True),
    icd_n_pages: int = Form(default=1),
    icd_vision_model: str = Form(default="openai/gpt-5.2:online"),  # Vision model selection
    icd_max_workers: int = Form(default=50),  # Increased for better parallelism
    icd_custom_instructions: str = Form(default=""),
    icd_instruction_template_id: Optional[int] = Form(default=None),  # For template selection
    output_filename: Optional[str] = Form(default=None)  # Custom output filename (optional)
):
    """
    Unified endpoint to run PDF splitting (optional) + data extraction + CPT prediction + ICD prediction
    Results are intelligently merged into a single output file
    """
    try:
        if excel_file and excel_file.filename:
            excel_filename_log = excel_file.filename
        else:
            excel_filename_log = f"template_id={template_id}" if template_id else "None"
        
        logger.info(f"Received unified processing request - split: {enable_split}, extraction={enable_extraction}, CPT={enable_cpt}, ICD={enable_icd}")
        logger.info(f"Files - PDF: {pdf_file.filename if pdf_file and pdf_file.filename else 'None'}, ZIP: {zip_file.filename if zip_file and zip_file.filename else 'None'}, Excel: {excel_filename_log}")
        
        # Validate at least one step is enabled
        if not (enable_split or enable_extraction or enable_cpt or enable_icd):
            raise HTTPException(status_code=400, detail="At least one processing step must be enabled")
        
        # Validate file inputs based on split mode
        if enable_split:
            if not pdf_file or not pdf_file.filename:
                raise HTTPException(status_code=400, detail="PDF file is required when split is enabled")
            if not pdf_file.filename.endswith('.pdf'):
                raise HTTPException(status_code=400, detail="File must be a PDF when split is enabled")
            if not split_filter_string.strip():
                raise HTTPException(status_code=400, detail="Split filter string is required when split is enabled")
        else:
            if not zip_file or not zip_file.filename:
                raise HTTPException(status_code=400, detail="ZIP file is required when split is disabled")
            if not zip_file.filename.endswith('.zip'):
                raise HTTPException(status_code=400, detail="Patient documents must be in a ZIP file when split is disabled")
        
        # Validate that either excel_file or template_id is provided
        if not excel_file and not template_id:
            raise HTTPException(status_code=400, detail="Either an Excel file or a template ID must be provided")
        
        if excel_file and excel_file.filename and not excel_file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Instructions file must be an Excel file")
        
        # Generate job ID
        job_id = str(uuid.uuid4())
        job = ProcessingJob(job_id)
        job_status[job_id] = job
        
        logger.info(f"Created unified processing job {job_id}")
        
        # Handle PDF splitting if enabled
        zip_path = None
        if enable_split:
            logger.info(f"Step 0: PDF Splitting enabled (method: {split_method})")
            job.status = "processing"
            job.message = f"Step 0: Splitting PDF with filter '{split_filter_string}'..."
            job.progress = 5
            
            # Save the PDF file
            pdf_path = f"/tmp/{job_id}_input.pdf"
            with open(pdf_path, "wb") as f:
                shutil.copyfileobj(pdf_file.file, f)
            logger.info(f"PDF saved to {pdf_path}")
            
            # Run the splitting synchronously (as part of the upload request)
            # This ensures we have the split PDFs before continuing
            try:
                current_dir = Path(__file__).parent / "current"
                output_dir = current_dir / "output" / job_id
                output_dir.mkdir(parents=True, exist_ok=True)
                
                job.message = f"Splitting PDF using {split_method} method..."
                job.progress = 10
                
                if split_method == "ocrspace":
                    # Use OCR.space splitting method
                    script_path = current_dir / "1-split_pdf_ocrspace.py"
                    if not script_path.exists():
                        raise Exception(f"OCR.space splitting script not found: {script_path}")
                    
                    import sys
                    import importlib.util
                    spec = importlib.util.spec_from_file_location("split_pdf_ocrspace", script_path)
                    split_module = importlib.util.module_from_spec(spec)
                    sys.modules["split_pdf_ocrspace"] = split_module
                    spec.loader.exec_module(split_module)
                    split_pdf_with_ocrspace = split_module.split_pdf_with_ocrspace
                    
                    created_count = split_pdf_with_ocrspace(
                        pdf_path,
                        str(output_dir),
                        [split_filter_string],
                        max_workers=7,
                        case_sensitive=False,
                        detection_shift=split_detection_shift
                    )
                else:
                    # Use legacy splitting method
                    raise HTTPException(status_code=400, detail="Legacy split method not yet supported in unified processing")
                
                logger.info(f"PDF split complete - created {created_count} files")
                job.message = f"Split complete - {created_count} PDFs created. Proceeding with processing..."
                job.progress = 20
                
                # Create a ZIP file from the split PDFs
                zip_path = f"/tmp/{job_id}_split.zip"
                import zipfile
                with zipfile.ZipFile(zip_path, 'w') as zipf:
                    for pdf_file_path in output_dir.glob("*.pdf"):
                        zipf.write(pdf_file_path, pdf_file_path.name)
                
                logger.info(f"Created ZIP from split PDFs: {zip_path}")
                
                # Clean up the PDF and split directory
                Path(pdf_path).unlink(missing_ok=True)
                
            except Exception as e:
                logger.error(f"PDF splitting failed: {str(e)}")
                job.status = "failed"
                job.error = f"PDF splitting failed: {str(e)}"
                raise HTTPException(status_code=500, detail=f"PDF splitting failed: {str(e)}")
        else:
            # Save uploaded ZIP file directly
            zip_path = f"/tmp/{job_id}_input.zip"
            with open(zip_path, "wb") as f:
                shutil.copyfileobj(zip_file.file, f)
            logger.info(f"ZIP file saved to {zip_path}")
        
        # Handle excel file or template
        excel_path = None
        excel_filename = None
        
        if template_id:
            # Export template to temporary Excel file
            from db_utils import get_template
            template = get_template(template_id)
            if not template:
                raise HTTPException(status_code=404, detail=f"Template with ID {template_id} not found")
            
            logger.info(f"Using template '{template['name']}' (ID: {template_id}) for extraction")
            
            # Create temporary Excel file from template
            import pandas as pd
            excel_path = f"/tmp/{job_id}_instructions.xlsx"
            excel_filename = f"{template['name']}.xlsx"
            
            # Convert template fields to Excel format (same format as /upload endpoint)
            fields = template['template_data'].get('fields', [])
            
            if not fields:
                raise HTTPException(status_code=400, detail="Template has no field definitions")
            
            # Create DataFrame in the expected format (field names as columns, metadata as rows)
            data = {}
            for field in fields:
                data[field['name']] = [
                    field.get('description', ''),
                    field.get('location', ''),
                    field.get('output_format', ''),
                    'YES' if field.get('priority', False) else 'NO'
                ]
            
            df = pd.DataFrame(data)
            df.to_excel(excel_path, index=False, engine='openpyxl')
            logger.info(f"Exported template to Excel: {excel_path}")
        else:
            # Use uploaded Excel file
            if not excel_file or not excel_file.filename:
                raise HTTPException(status_code=400, detail="Excel file is required when template_id is not provided")
            excel_path = f"/tmp/{job_id}_instructions{Path(excel_file.filename).suffix}"
            excel_filename = excel_file.filename
            with open(excel_path, "wb") as f:
                shutil.copyfileobj(excel_file.file, f)
        
        logger.info(f"Files saved - zip: {zip_path}, excel: {excel_path}")
        
        # Fetch instruction templates if provided
        if cpt_instruction_template_id:
            from db_utils import get_prediction_instruction
            template = get_prediction_instruction(instruction_id=cpt_instruction_template_id)
            if template:
                cpt_custom_instructions = template['instructions_text']
                logger.info(f"Using CPT instruction template '{template['name']}' for unified processing")
        
        if icd_instruction_template_id:
            from db_utils import get_prediction_instruction
            template = get_prediction_instruction(instruction_id=icd_instruction_template_id)
            if template:
                icd_custom_instructions = template['instructions_text']
                logger.info(f"Using ICD instruction template '{template['name']}' for unified processing")
        
        # Start background processing
        background_tasks.add_task(
            process_unified_background,
            job_id=job_id,
            zip_path=zip_path,
            excel_path=excel_path,
            excel_filename=excel_filename,
            output_filename=output_filename,
            enable_extraction=enable_extraction,
            extraction_n_pages=extraction_n_pages,
            extraction_model=extraction_model,
            extraction_max_workers=extraction_max_workers,
            worktracker_group=worktracker_group,
            worktracker_batch=worktracker_batch,
            extract_csn=extract_csn,
            enable_cpt=enable_cpt,
            cpt_vision_mode=cpt_vision_mode,
            cpt_client=cpt_client,
            cpt_vision_pages=cpt_vision_pages,
            cpt_vision_model=cpt_vision_model,
            cpt_include_code_list=cpt_include_code_list,
            cpt_max_workers=cpt_max_workers,
            cpt_custom_instructions=cpt_custom_instructions,
            cpt_instruction_template_id=cpt_instruction_template_id,
            enable_icd=enable_icd,
            icd_n_pages=icd_n_pages,
            icd_vision_model=icd_vision_model,
            icd_max_workers=icd_max_workers,
            icd_custom_instructions=icd_custom_instructions,
            icd_instruction_template_id=icd_instruction_template_id
        )
        
        logger.info(f"Background unified processing task started for job {job_id}")
        
        return {"job_id": job_id, "message": "Unified processing started"}
        
    except Exception as e:
        logger.error(f"Unified processing upload error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")


@app.post("/api/process-unified-with-refinement")
async def process_unified_with_refinement(
    background_tasks: BackgroundTasks,
    zip_file: UploadFile = File(...),
    excel_file: UploadFile = File(None),
    template_id: Optional[int] = Form(default=None),
    ground_truth_file: UploadFile = File(...),
    # Extraction parameters
    enable_extraction: bool = Form(default=True),
    extraction_n_pages: int = Form(default=2),
    extraction_model: str = Form(default="gemini-flash-latest"),
    extraction_max_workers: int = Form(default=50),
    worktracker_group: str = Form(default=""),
    worktracker_batch: str = Form(default=""),
    extract_csn: bool = Form(default=False),
    # CPT parameters
    enable_cpt: bool = Form(default=True),
    cpt_vision_mode: bool = Form(default=False),
    cpt_client: str = Form(default="uni"),
    cpt_vision_pages: int = Form(default=1),
    cpt_vision_model: str = Form(default="openai/gpt-5.2:online"),
    cpt_include_code_list: bool = Form(default=True),
    cpt_max_workers: int = Form(default=100),
    cpt_instruction_template_id: Optional[int] = Form(default=None),  # Required if enable_cpt is True
    # ICD parameters
    enable_icd: bool = Form(default=True),
    icd_n_pages: int = Form(default=1),
    icd_vision_model: str = Form(default="openai/gpt-5.2:online"),
    icd_max_workers: int = Form(default=100),
    icd_instruction_template_id: Optional[int] = Form(default=None),  # Required if enable_icd is True
    # Refinement parameters
    target_cpt_accuracy: float = Form(default=0.95),
    target_icd_accuracy: float = Form(default=0.95),
    max_iterations: int = Form(default=10),
    notification_email: str = Form(default="cvetkovskileon@gmail.com"),
    refinement_guidance: Optional[str] = Form(default=None),
    refinement_mode: str = Form(default="batch"),  # "batch" or "focused"
    batch_size: int = Form(default=10),  # Number of errors per batch in batch mode
    refinement_model: str = Form(default="gemini-3-flash-preview")  # Model to use for refinement
):
    """
    Unified processing with AI-powered iterative instruction refinement.
    Refines CPT instructions first, then ICD instructions, until target accuracy is reached.
    """
    try:
        logger.info(f"Received refinement request - CPT template: {cpt_instruction_template_id}, ICD template: {icd_instruction_template_id}")
        
        # Validate inputs
        if not zip_file or not zip_file.filename:
            raise HTTPException(status_code=400, detail="ZIP file is required")
        if not zip_file.filename.endswith('.zip'):
            raise HTTPException(status_code=400, detail="Patient documents must be in a ZIP file")
        
        if not ground_truth_file or not ground_truth_file.filename:
            raise HTTPException(status_code=400, detail="Ground truth file is required")
        if not ground_truth_file.filename.endswith(('.csv', '.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Ground truth file must be CSV or Excel")
        
        if not excel_file and not template_id:
            raise HTTPException(status_code=400, detail="Either an Excel file or a template ID must be provided")
        
        # Validate instruction templates exist (only if the corresponding refinement is enabled)
        from db_utils import get_prediction_instruction
        
        # Validate CPT template if CPT refinement is enabled
        if enable_cpt:
            if not cpt_instruction_template_id:
                raise HTTPException(status_code=400, detail="CPT instruction template ID is required when CPT refinement is enabled")
            cpt_template = get_prediction_instruction(instruction_id=cpt_instruction_template_id)
            if not cpt_template:
                raise HTTPException(status_code=404, detail=f"CPT instruction template {cpt_instruction_template_id} not found")
        else:
            cpt_template = None
            cpt_instruction_template_id = None  # Set to None if disabled
        
        # Validate ICD template if ICD refinement is enabled
        if enable_icd:
            if not icd_instruction_template_id:
                raise HTTPException(status_code=400, detail="ICD instruction template ID is required when ICD refinement is enabled")
            icd_template = get_prediction_instruction(instruction_id=icd_instruction_template_id)
            if not icd_template:
                raise HTTPException(status_code=404, detail=f"ICD instruction template {icd_instruction_template_id} not found")
        else:
            icd_template = None
            icd_instruction_template_id = None  # Set to None if disabled
        
        # Ensure at least one refinement type is enabled
        if not enable_cpt and not enable_icd:
            raise HTTPException(status_code=400, detail="At least one refinement type (CPT or ICD) must be enabled")
        
        # Generate job ID
        job_id = str(uuid.uuid4())
        
        logger.info(f"Created refinement job {job_id}")
        
        # Save files
        zip_path = f"/tmp/{job_id}_input.zip"
        with open(zip_path, "wb") as f:
            shutil.copyfileobj(zip_file.file, f)
        
        ground_truth_path = f"/tmp/{job_id}_ground_truth{Path(ground_truth_file.filename).suffix}"
        with open(ground_truth_path, "wb") as f:
            shutil.copyfileobj(ground_truth_file.file, f)
        
        # Handle excel file or template
        excel_path = None
        excel_filename = None
        
        if template_id:
            from db_utils import get_template
            template = get_template(template_id)
            if not template:
                raise HTTPException(status_code=404, detail=f"Template with ID {template_id} not found")
            
            import pandas as pd
            excel_path = f"/tmp/{job_id}_instructions.xlsx"
            excel_filename = f"{template['name']}.xlsx"
            
            fields = template['template_data'].get('fields', [])
            if not fields:
                raise HTTPException(status_code=400, detail="Template has no field definitions")
            
            data = {}
            for field in fields:
                data[field['name']] = [
                    field.get('description', ''),
                    field.get('location', ''),
                    field.get('output_format', ''),
                    'YES' if field.get('priority', False) else 'NO'
                ]
            
            df = pd.DataFrame(data)
            df.to_excel(excel_path, index=False, engine='openpyxl')
        else:
            if not excel_file or not excel_file.filename:
                raise HTTPException(status_code=400, detail="Excel file is required when template_id is not provided")
            excel_path = f"/tmp/{job_id}_instructions{Path(excel_file.filename).suffix}"
            excel_filename = excel_file.filename
            with open(excel_path, "wb") as f:
                shutil.copyfileobj(excel_file.file, f)
        
        # Start background refinement task
        from refinement_orchestrator import run_refinement_job
        
        background_tasks.add_task(
            run_refinement_job,
            job_id=job_id,
            zip_path=zip_path,
            excel_path=excel_path,
            excel_filename=excel_filename,
            ground_truth_path=ground_truth_path,
            enable_extraction=enable_extraction,
            extraction_n_pages=extraction_n_pages,
            extraction_model=extraction_model,
            extraction_max_workers=extraction_max_workers,
            worktracker_group=worktracker_group,
            worktracker_batch=worktracker_batch,
            extract_csn=extract_csn,
            enable_cpt=enable_cpt,
            cpt_vision_mode=cpt_vision_mode,
            cpt_client=cpt_client,
            cpt_vision_pages=cpt_vision_pages,
            cpt_vision_model=cpt_vision_model,
            cpt_include_code_list=cpt_include_code_list,
            cpt_max_workers=cpt_max_workers,
            original_cpt_template_id=cpt_instruction_template_id,
            enable_icd=enable_icd,
            icd_n_pages=icd_n_pages,
            icd_vision_model=icd_vision_model,
            icd_max_workers=icd_max_workers,
            original_icd_template_id=icd_instruction_template_id,
            target_cpt_accuracy=target_cpt_accuracy,
            target_icd_accuracy=target_icd_accuracy,
            max_iterations=max_iterations,
            notification_email=notification_email,
            refinement_guidance=refinement_guidance,
            refinement_mode=refinement_mode,
            batch_size=batch_size,
            refinement_model=refinement_model
        )
        
        logger.info(f"Background refinement task started for job {job_id} (mode: {refinement_mode}, batch_size: {batch_size})")
        
        return {"job_id": job_id, "message": "AI refinement started"}
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Refinement request error: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Request failed: {str(e)}")


@app.get("/api/refinement-status/{job_id}")
async def get_refinement_status(job_id: str):
    """
    Get current status of a refinement job.
    """
    try:
        from db_utils import get_refinement_job
        
        job = get_refinement_job(job_id)
        if not job:
            raise HTTPException(status_code=404, detail=f"Refinement job {job_id} not found")
        
        # Build status message
        phase = job.get('phase', 'cpt')
        phase_name = "CPT Refinement" if phase == "cpt" else "ICD Refinement" if phase == "icd" else "Complete"
        iteration = job.get('iteration', 0)
        status = job.get('status', 'running')
        
        if status == "running":
            message = f"Running {phase_name} iteration {iteration}..."
        elif status == "completed":
            message = "Refinement complete"
        elif status == "failed":
            message = f"Failed: {job.get('error_message', 'Unknown error')}"
        else:
            message = f"Status: {status}"
        
        return {
            "job_id": job_id,
            "status": status,
            "phase": phase,
            "iteration": iteration,
            "cpt_accuracy": job.get('cpt_accuracy'),
            "icd1_accuracy": job.get('icd1_accuracy'),
            "best_cpt_accuracy": job.get('best_cpt_accuracy'),
            "best_icd1_accuracy": job.get('best_icd1_accuracy'),
            "message": message,
            "error_message": job.get('error_message')  # Include error_message for case statuses in focused mode
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get refinement status: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to get status: {str(e)}")


@app.get("/api/refinement-history/{job_id}")
async def get_refinement_history(job_id: str):
    """
    Get full history of a refinement job including all iterations.
    Note: This currently returns the current state. Full history tracking
    would require additional database schema changes.
    """
    try:
        from db_utils import get_refinement_job, get_prediction_instruction
        
        job = get_refinement_job(job_id)
        if not job:
            raise HTTPException(status_code=404, detail=f"Refinement job {job_id} not found")
        
        # Get template details
        current_cpt_template = None
        current_icd_template = None
        best_cpt_template = None
        best_icd_template = None
        
        if job.get('current_cpt_template_id'):
            current_cpt_template = get_prediction_instruction(instruction_id=job['current_cpt_template_id'])
        if job.get('current_icd_template_id'):
            current_icd_template = get_prediction_instruction(instruction_id=job['current_icd_template_id'])
        if job.get('best_cpt_template_id'):
            best_cpt_template = get_prediction_instruction(instruction_id=job['best_cpt_template_id'])
        if job.get('best_icd_template_id'):
            best_icd_template = get_prediction_instruction(instruction_id=job['best_icd_template_id'])
        
        return {
            "job_id": job_id,
            "status": job.get('status'),
            "phase": job.get('phase'),
            "iteration": job.get('iteration'),
            "cpt_accuracy": job.get('cpt_accuracy'),
            "icd1_accuracy": job.get('icd1_accuracy'),
            "best_cpt_accuracy": job.get('best_cpt_accuracy'),
            "best_icd1_accuracy": job.get('best_icd1_accuracy'),
            "current_cpt_template": {
                "id": current_cpt_template['id'] if current_cpt_template else None,
                "name": current_cpt_template['name'] if current_cpt_template else None
            },
            "current_icd_template": {
                "id": current_icd_template['id'] if current_icd_template else None,
                "name": current_icd_template['name'] if current_icd_template else None
            },
            "best_cpt_template": {
                "id": best_cpt_template['id'] if best_cpt_template else None,
                "name": best_cpt_template['name'] if best_cpt_template else None
            },
            "best_icd_template": {
                "id": best_icd_template['id'] if best_icd_template else None,
                "name": best_icd_template['name'] if best_icd_template else None
            },
            "created_at": job.get('created_at').isoformat() if job.get('created_at') else None,
            "updated_at": job.get('updated_at').isoformat() if job.get('updated_at') else None
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to get refinement history: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to get history: {str(e)}")


@app.get("/api/unified-results")
async def list_unified_results(page: int = 1, page_size: int = 50):
    """
    List all unified processing results with pagination.
    """
    try:
        from db_utils import get_all_unified_results
        
        result = get_all_unified_results(page=page, page_size=page_size)
        
        # Format dates and file sizes
        for res in result['results']:
            if res.get('created_at'):
                res['created_at'] = res['created_at'].isoformat() if hasattr(res['created_at'], 'isoformat') else str(res['created_at'])
            if res.get('expires_at'):
                res['expires_at'] = res['expires_at'].isoformat() if hasattr(res['expires_at'], 'isoformat') else str(res['expires_at'])
            if res.get('file_size_bytes'):
                res['file_size_mb'] = round(res['file_size_bytes'] / 1024 / 1024, 2)
        
        return result
    except Exception as e:
        logger.error(f"Failed to list unified results: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to list results: {str(e)}")


@app.get("/api/unified-results/{job_id}/download")
async def download_unified_result(job_id: str, format: str = "csv"):
    """
    Download a unified processing result file.
    
    Args:
        job_id: Job identifier
        format: File format - "csv" or "xlsx" (default: "csv")
    """
    try:
        from db_utils import get_unified_result
        
        result = get_unified_result(job_id)
        if not result:
            raise HTTPException(status_code=404, detail=f"Result not found for job {job_id}")
        
        # Determine file path based on format
        if format.lower() == "xlsx" and result.get('file_path_xlsx'):
            file_path = result['file_path_xlsx']
        elif format.lower() == "csv" and result.get('file_path_csv'):
            file_path = result['file_path_csv']
        else:
            # Fallback to CSV if XLSX not available
            if format.lower() == "xlsx":
                logger.warning(f"XLSX not available for {job_id}, falling back to CSV")
            file_path = result['file_path_csv']
        
        if not os.path.exists(file_path):
            raise HTTPException(status_code=404, detail=f"File not found: {file_path}")
        
        # Determine filename
        filename = result.get('filename', f"unified_result_{job_id}")
        if not filename.endswith(f'.{format.lower()}'):
            filename = f"{filename}.{format.lower()}"
        
        # Return file
        return FileResponse(
            file_path,
            media_type="application/octet-stream",
            filename=filename
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Failed to download unified result {job_id}: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to download result: {str(e)}")


@app.post("/api/unified-results/cleanup")
async def cleanup_expired_unified_results():
    """
    Manually trigger cleanup of expired unified results (older than 3 days).
    """
    try:
        from db_utils import delete_expired_unified_results
        
        deleted_count = delete_expired_unified_results()
        
        return {
            "message": f"Cleanup completed",
            "deleted_count": deleted_count
        }
    except Exception as e:
        logger.error(f"Failed to cleanup expired results: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to cleanup: {str(e)}")


if __name__ == "__main__":
    # This is for local development only
    # Railway will use uvicorn directly via railway.json
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=PORT, log_level="info") 