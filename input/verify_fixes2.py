"""Submit 3 more unified jobs; compare against ground truth."""
import io, json, os, ssl, time, urllib.request, zipfile, uuid
from concurrent.futures import ThreadPoolExecutor

BACKEND = "https://medical-data-processor-production.up.railway.app"
MODEL = "google/gemini-3-flash-preview"
N_PAGES = 6

HERE = os.path.dirname(os.path.abspath(__file__))
ERROR_DIR = os.path.join(HERE, "error_pdfs")
OUT_DIR = os.path.join(HERE, "verify_output")
os.makedirs(OUT_DIR, exist_ok=True)

CTX = ssl.create_default_context()
CTX.check_hostname = False
CTX.verify_mode = ssl.CERT_NONE

GT = {
    "CHA-HDH_472":   {"LENNARD":("00902","K60.30"), "BRADLEY":("01967","O99.62"), "RAKER":("00670","M48.02")},
    "IAS-FVO_418":   {"WICKMAN":("01400","S83.242A"), "PATEL":("01400","S83.242A"), "GEAMAN":("01630","S46.011A")},
    "PCE-WWMG_333":  {"LUKER":("00811","D12.3"), "CAMARA":("00811","Z12.11"), "TANG":("00811","K63.5")},
}

# (bucket, test_cpt, cpt_tid, test_icd, icd_tid, extraction_tid)
JOBS = [
    ("CHA-HDH_472",   False, None, True, 57,  75),
    ("IAS-FVO_418",   False, None, True, 45,  119),
    ("PCE-WWMG_333",  True,  161,  True, 162, 100),
]


def make_zip(pdf_dir):
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for f in sorted(os.listdir(pdf_dir)):
            if f.lower().endswith(".pdf"):
                z.write(os.path.join(pdf_dir, f), arcname=f)
    return buf.getvalue()


def multipart_post(url, fields, file_name, file_bytes):
    boundary = f"----b{uuid.uuid4().hex}"
    parts = []
    for k, v in fields.items():
        parts.append(f"--{boundary}\r\nContent-Disposition: form-data; name=\"{k}\"\r\n\r\n{v}\r\n".encode())
    parts.append(
        f"--{boundary}\r\nContent-Disposition: form-data; name=\"zip_file\"; "
        f"filename=\"{file_name}\"\r\nContent-Type: application/zip\r\n\r\n".encode()
        + file_bytes + b"\r\n"
    )
    parts.append(f"--{boundary}--\r\n".encode())
    body = b"".join(parts)
    req = urllib.request.Request(
        url, data=body, method="POST",
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
    )
    with urllib.request.urlopen(req, context=CTX, timeout=120) as r:
        return json.loads(r.read())


def submit(bucket, test_cpt, cpt_tid, test_icd, icd_tid, extraction_tid):
    pdf_dir = os.path.join(ERROR_DIR, bucket)
    zbytes = make_zip(pdf_dir)
    fields = {
        "template_id": str(extraction_tid),
        "enable_extraction": "false",
        "enable_cpt": "true" if test_cpt else "false",
        "enable_icd": "true" if test_icd else "false",
    }
    if test_cpt:
        fields.update({
            "cpt_vision_mode": "true",
            "cpt_vision_pages": str(N_PAGES),
            "cpt_vision_model": MODEL,
            "cpt_max_workers": "20",
            "cpt_instruction_template_id": str(cpt_tid),
        })
    if test_icd:
        fields.update({
            "icd_n_pages": str(N_PAGES),
            "icd_vision_model": MODEL,
            "icd_max_workers": "20",
            "icd_instruction_template_id": str(icd_tid),
        })
    print(f"[{bucket}] submitting ({len(zbytes)} bytes)...")
    resp = multipart_post(f"{BACKEND}/process-unified", fields, f"{bucket}.zip", zbytes)
    return resp.get("job_id")


with ThreadPoolExecutor(max_workers=3) as ex:
    futures = {ex.submit(submit, *j): j[0] for j in JOBS}
    jobs = {}
    for fut, bucket in list(futures.items()):
        try:
            jobs[bucket] = fut.result()
            print(f"[{bucket}] job_id={jobs[bucket]}")
        except Exception as e:
            print(f"[{bucket}] ERROR: {e}"); jobs[bucket] = None


def get_status(jid):
    req = urllib.request.Request(f"{BACKEND}/status/{jid}")
    with urllib.request.urlopen(req, context=CTX, timeout=30) as r:
        return json.loads(r.read())


# Poll
pending = {b for b, j in jobs.items() if j}
start = time.time()
while pending:
    for bucket in list(pending):
        try:
            st = get_status(jobs[bucket])
            s = st.get("status"); p = st.get("progress", 0)
            print(f"  t={int(time.time()-start)}s [{bucket}] {s} {p}% — {st.get('message','')[:60]}")
            if s in ("completed", "failed"):
                pending.discard(bucket)
        except Exception as e:
            print(f"  [{bucket}] poll error: {e}")
    if pending: time.sleep(8)


def download(jid, path):
    req = urllib.request.Request(f"{BACKEND}/download/{jid}?format=xlsx")
    with urllib.request.urlopen(req, context=CTX, timeout=120) as r:
        data = r.read()
    with open(path, "wb") as fh:
        fh.write(data)
    return len(data)


for bucket, jid in jobs.items():
    if not jid: continue
    out = os.path.join(OUT_DIR, f"{bucket}.xlsx")
    try:
        n = download(jid, out)
        print(f"[{bucket}] downloaded {n} bytes -> {out}")
    except Exception as e:
        print(f"[{bucket}] download error: {e}")


# Compare
import openpyxl
def load(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return [], []
    header = [str(c).strip() if c else "" for c in rows[0]]
    return header, [dict(zip(header, r)) for r in rows[1:]]


def find(header, candidates):
    for c in candidates:
        for h in header:
            if h and c.lower() == h.lower():
                return h
    for c in candidates:
        for h in header:
            if h and c.lower() in h.lower():
                return h
    return None


print("\n" + "="*70)
print("VERIFICATION vs GROUND TRUTH")
print("="*70)
total = passed = 0
for bucket, test_cpt, _, test_icd, _, _ in JOBS:
    path = os.path.join(OUT_DIR, f"{bucket}.xlsx")
    if not os.path.exists(path):
        print(f"\n[{bucket}] no result file"); continue
    header, rows = load(path)
    c_cpt = find(header, ["procedure code", "cpt_code", "cpt", "predicted_cpt"])
    c_icd = find(header, ["icd1", "icd_1", "predicted_icd1"])
    print(f"\n--- {bucket} ({len(rows)} rows) cpt={c_cpt!r} icd={c_icd!r} ---")
    for patient, (gt_cpt, gt_icd) in GT[bucket].items():
        match = None
        for r in rows:
            row_text = " ".join(str(v or "") for v in r.values()).upper()
            if patient in row_text:
                match = r; break
        if not match:
            print(f"  {patient}: NOT FOUND"); continue
        ai_cpt = str(match.get(c_cpt, "") or "").strip() if c_cpt else ""
        ai_icd = str(match.get(c_icd, "") or "").strip() if c_icd else ""
        parts = []
        if test_cpt:
            total += 1; ok = ai_cpt == gt_cpt; passed += int(ok)
            parts.append(f"CPT AI={ai_cpt:<8} GT={gt_cpt:<8} {'PASS' if ok else 'FAIL'}")
        if test_icd:
            total += 1; ok = ai_icd == gt_icd; passed += int(ok)
            parts.append(f"ICD1 AI={ai_icd:<10} GT={gt_icd:<10} {'PASS' if ok else 'FAIL'}")
        print(f"  {patient:<10} " + " | ".join(parts))

print(f"\n{'='*70}\n  TOTAL: {passed}/{total} passed ({passed/total*100:.0f}%)" if total else "")
