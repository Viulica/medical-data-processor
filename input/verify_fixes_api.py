"""Submit 3 unified processing jobs to Railway backend; compare to ground truth."""
import io, json, os, ssl, sys, time, urllib.request, zipfile, mimetypes, uuid, csv

BACKEND = "https://medical-data-processor-production.up.railway.app"
MODEL = "google/gemini-3-flash-preview"
N_PAGES = 6

HERE = os.path.dirname(os.path.abspath(__file__))
ERROR_DIR = os.path.join(HERE, "error_pdfs")
OUT_DIR = os.path.join(HERE, "verify_output")
os.makedirs(OUT_DIR, exist_ok=True)

CTX = ssl.create_default_context()
CTX.check_hostname = False
CTX.verify_mode = ssl.CERT_NONE

GT = {
    "LOV_162": {"MORRIS":("00142","H25.12"),"DEAN":("00142","H25.12"),"FREUND":("00142","H25.12")},
    "INJE-CLIK_119": {"KIESOW":("00140","H40.9"),"SWARTZ":("00140","H40.9"),"SNYDER":("00140","H40.9")},
    "WPA_339": {"DAVIS":("00140","H40.1111"),"HEDGCOTH":("00140","H40.1121"),"SHEPARD":("00140","H40.1121")},
}

# What to test per bucket
JOBS = [
    # (bucket, test_cpt, cpt_tmpl_id, test_icd, icd_tmpl_id, extraction_tmpl_id)
    ("LOV_162",       True,  139,   False, None, 85),
    ("INJE-CLIK_119", False, None,  True,  126,  128),
    ("WPA_339",       False, None,  True,  58,   54),
]


def make_zip(pdf_dir):
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for f in sorted(os.listdir(pdf_dir)):
            if f.lower().endswith(".pdf"):
                z.write(os.path.join(pdf_dir, f), arcname=f)
    return buf.getvalue()


def multipart_post(url, fields, file_name, file_bytes):
    boundary = f"----boundary{uuid.uuid4().hex}"
    parts = []
    for k, v in fields.items():
        parts.append(f"--{boundary}\r\nContent-Disposition: form-data; name=\"{k}\"\r\n\r\n{v}\r\n".encode())
    parts.append(
        f"--{boundary}\r\nContent-Disposition: form-data; name=\"zip_file\"; "
        f"filename=\"{file_name}\"\r\nContent-Type: application/zip\r\n\r\n".encode()
        + file_bytes + b"\r\n"
    )
    parts.append(f"--{boundary}--\r\n".encode())
    body = b"".join(parts)
    req = urllib.request.Request(
        url, data=body, method="POST",
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
    )
    with urllib.request.urlopen(req, context=CTX, timeout=120) as r:
        return json.loads(r.read())


def submit(bucket, test_cpt, cpt_tid, test_icd, icd_tid, extraction_tid):
    pdf_dir = os.path.join(ERROR_DIR, bucket)
    zbytes = make_zip(pdf_dir)
    fields = {
        "template_id": str(extraction_tid),
        "enable_extraction": "false",
        "enable_cpt": "true" if test_cpt else "false",
        "enable_icd": "true" if test_icd else "false",
    }
    if test_cpt:
        fields.update({
            "cpt_vision_mode": "true",
            "cpt_vision_pages": str(N_PAGES),
            "cpt_vision_model": MODEL,
            "cpt_max_workers": "20",
            "cpt_instruction_template_id": str(cpt_tid),
        })
    if test_icd:
        fields.update({
            "icd_n_pages": str(N_PAGES),
            "icd_vision_model": MODEL,
            "icd_max_workers": "20",
            "icd_instruction_template_id": str(icd_tid),
        })
    print(f"[{bucket}] submitting ({len(zbytes)} bytes zip)...")
    resp = multipart_post(f"{BACKEND}/process-unified", fields, f"{bucket}.zip", zbytes)
    return resp.get("job_id")


def poll(job_id, label):
    start = time.time()
    while True:
        req = urllib.request.Request(f"{BACKEND}/status/{job_id}")
        with urllib.request.urlopen(req, context=CTX, timeout=30) as r:
            st = json.loads(r.read())
        status = st.get("status")
        progress = st.get("progress", 0)
        elapsed = int(time.time() - start)
        print(f"  [{label}] t={elapsed}s status={status} progress={progress}%")
        if status in ("completed", "failed"):
            return st
        if elapsed > 600:
            print("  timeout"); return st
        time.sleep(5)


def download_result(job_id, out_path):
    req = urllib.request.Request(f"{BACKEND}/api/unified-results/{job_id}/download")
    with urllib.request.urlopen(req, context=CTX, timeout=60) as r:
        data = r.read()
    with open(out_path, "wb") as fh:
        fh.write(data)
    return out_path


# Submit all 3 in parallel
from concurrent.futures import ThreadPoolExecutor
with ThreadPoolExecutor(max_workers=3) as ex:
    futures = {ex.submit(submit, *j): j[0] for j in JOBS}
    jobs = {}
    for fut in futures:
        bucket = futures[fut]
        jobs[bucket] = fut.result()
        print(f"[{bucket}] job_id={jobs[bucket]}")

# Poll each
results = {}
for bucket, job_id in jobs.items():
    if not job_id:
        print(f"[{bucket}] no job_id"); continue
    st = poll(job_id, bucket)
    results[bucket] = st
    if st.get("status") == "completed":
        out = os.path.join(OUT_DIR, f"{bucket}.xlsx")
        download_result(job_id, out)
        print(f"  downloaded -> {out}")

# Load and compare
import openpyxl

def load_xlsx_rows(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return []
    header = [str(c).strip() if c else "" for c in rows[0]]
    return [dict(zip(header, r)) for r in rows[1:]]

print("\n" + "="*70)
print("VERIFICATION vs GROUND TRUTH")
print("="*70)

total = 0; passed = 0
for bucket, test_cpt, _, test_icd, _, _ in JOBS:
    out = os.path.join(OUT_DIR, f"{bucket}.xlsx")
    if not os.path.exists(out):
        print(f"\n[{bucket}] SKIP — no result file"); continue
    rows = load_xlsx_rows(out)
    print(f"\n--- {bucket} ({len(rows)} rows) ---")
    col_cpt = None; col_icd = None; col_file = None
    if rows:
        for k in rows[0].keys():
            kl = k.lower()
            if "cpt" in kl and "desc" not in kl and not col_cpt: col_cpt = k
            if "icd1" in kl or kl == "icd 1" or kl == "icd": col_icd = col_icd or k
            if "file" in kl or "pdf" in kl or "patient" in kl: col_file = col_file or k
    print(f"  columns used: file={col_file!r}  cpt={col_cpt!r}  icd1={col_icd!r}")
    patient_gt = GT.get(bucket, {})
    for patient, (gt_cpt, gt_icd) in patient_gt.items():
        match = None
        for r in rows:
            fval = str(r.get(col_file,"") or "").upper()
            # Also check all string values for patient match
            all_vals = " ".join(str(v) for v in r.values()).upper()
            if patient in fval or patient in all_vals:
                match = r; break
        if not match:
            print(f"  {patient}: NOT FOUND in output"); continue
        ai_cpt = str(match.get(col_cpt, "") or "").strip()
        ai_icd = str(match.get(col_icd, "") or "").strip()
        checks = []
        if test_cpt:
            total += 1
            ok = ai_cpt == gt_cpt
            if ok: passed += 1
            checks.append(f"CPT: AI={ai_cpt:<8} GT={gt_cpt:<8} {'OK' if ok else 'FAIL'}")
        if test_icd:
            total += 1
            ok = ai_icd == gt_icd
            if ok: passed += 1
            checks.append(f"ICD1: AI={ai_icd:<10} GT={gt_icd:<10} {'OK' if ok else 'FAIL'}")
        print(f"  {patient:<10}  " + " | ".join(checks))

print(f"\n{'='*70}\n  TOTAL: {passed}/{total} passed  ({passed/total*100:.0f}%)")
